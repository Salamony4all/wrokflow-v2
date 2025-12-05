"""
Excel File Processor
Handles reading and extracting data from Excel files (.xls, .xlsx)
Includes image extraction from cells
"""

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as OpenpyxlImage
from PIL import Image
import os
import json
import logging
import base64
from io import BytesIO
import shutil

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Process Excel files and extract table data with images"""
    
    def __init__(self, filepath):
        """
        Initialize Excel processor
        
        Args:
            filepath: Path to Excel file
        """
        self.original_filepath = filepath
        self.original_extension = os.path.splitext(filepath)[1].lower()
        
        # Convert .xls to .xlsx for uniform processing
        if self.original_extension == '.xls':
            logger.info(f"Converting .xls file to .xlsx format: {filepath}")
            try:
                self.filepath = self._convert_xls_to_xlsx(filepath)
                self.extension = '.xlsx'
            except ValueError as ve:
                # Conversion failed - provide helpful error
                logger.error(str(ve))
                raise ValueError(f"❌ .XLS file format is not supported due to compatibility issues. Please save your file as .XLSX format in Excel/LibreOffice and upload again.")
        else:
            self.filepath = filepath
            self.extension = self.original_extension
        
        self.filename = os.path.basename(self.filepath)
        self.workbook = None
        self.images_cache = {}  # Cache for extracted images
        
    def _convert_xls_to_xlsx(self, xls_filepath):
        """
        Convert .xls file to .xlsx format for uniform processing
        
        Args:
            xls_filepath: Path to .xls file
            
        Returns:
            str: Path to converted .xlsx file
        """
        try:
            # Create output path (same directory, .xlsx extension)
            xlsx_filepath = xls_filepath.replace('.xls', '_converted.xlsx')
            
            # Try reading with xlrd - but this may fail for some .xls files
            try:
                xls_file = pd.ExcelFile(xls_filepath, engine='xlrd')
            except (AssertionError, Exception) as xlrd_error:
                logger.error(f"xlrd failed to read .xls file: {str(xlrd_error)}")
                logger.error("This .xls file has compatibility issues. Please convert it to .xlsx format manually.")
                raise ValueError("Cannot read .xls file: The file format is incompatible. Please save the file as .xlsx format and upload again.")
            
            # Create Excel writer for .xlsx
            with pd.ExcelWriter(xlsx_filepath, engine='openpyxl') as writer:
                for sheet_name in xls_file.sheet_names:
                    # Read each sheet
                    df = pd.read_excel(xls_file, sheet_name=sheet_name, header=None)
                    # Write to xlsx
                    df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            
            logger.info(f"Successfully converted .xls to .xlsx: {xlsx_filepath}")
            return xlsx_filepath
            
        except ValueError:
            # Re-raise ValueError with our custom message
            raise
        except Exception as e:
            logger.error(f"Failed to convert .xls to .xlsx: {str(e)}")
            raise ValueError(f"Cannot convert .xls file: {str(e)}. Please save as .xlsx format and upload again.")
    
    def _extract_images_from_sheet(self, sheet, output_dir):
        """
        Extract all images from an Excel sheet and map them to their cell positions
        Note: Only works with .xlsx files (openpyxl). .xls files don't support image extraction.
        
        Args:
            sheet: openpyxl worksheet object
            output_dir: Directory to save extracted images
            
        Returns:
            dict: Mapping of row numbers to image paths (row-based mapping)
        """
        images_dir = os.path.join(output_dir, 'imgs')
        os.makedirs(images_dir, exist_ok=True)
        
        row_images = {}  # Map row numbers to images
        
        if not hasattr(sheet, '_images') or not sheet._images:
            logger.info(f"No images found in sheet '{sheet.title}'")
            return row_images
        
        logger.info(f"Found {len(sheet._images)} images in sheet '{sheet.title}'")
        
        for idx, img in enumerate(sheet._images):
            try:
                # Get image anchor (position)
                anchor = None
                if hasattr(img, 'anchor'):
                    if hasattr(img.anchor, '_from'):
                        anchor = img.anchor._from
                    elif hasattr(img.anchor, 'from'):
                        anchor = img.anchor['from']
                
                if anchor:
                    # Get the row where image is anchored (1-indexed in Excel)
                    if hasattr(anchor, 'row'):
                        row_num = anchor.row + 1  # Convert to 1-indexed
                    elif hasattr(anchor, 'rowOff'):
                        row_num = anchor.rowOff + 1
                    else:
                        row_num = 0
                    
                    # Save image
                    img_filename = f"{sheet.title.replace(' ', '_')}_row{row_num}_img{idx}.png"
                    img_path = os.path.join(images_dir, img_filename)
                    
                    # Get image data and save
                    if hasattr(img, '_data'):
                        with open(img_path, 'wb') as f:
                            f.write(img._data())
                    elif hasattr(img, 'ref'):
                        # Handle embedded images
                        img_data = img.ref
                        if hasattr(img_data, '_data'):
                            with open(img_path, 'wb') as f:
                                f.write(img_data())
                    
                    # Store relative path mapped to row number
                    rel_path = f"imgs/{img_filename}"
                    
                    if row_num not in row_images:
                        row_images[row_num] = []
                    row_images[row_num].append(rel_path)
                    
                    logger.info(f"Extracted image {idx} at row {row_num}: {img_filename}")
                else:
                    logger.warning(f"Could not determine anchor for image {idx}")
                    
            except Exception as e:
                logger.error(f"Error extracting image {idx}: {e}", exc_info=True)
        
        return row_images
    
    def _get_cell_value_with_wrapping(self, cell):
        """
        Get cell value with proper text wrapping for long content
        
        Args:
            cell: openpyxl cell object
            
        Returns:
            str: Formatted cell value
        """
        if cell.value is None:
            return ''
        
        text = str(cell.value).strip()
        
        # Check if text is long and needs wrapping
        if len(text) > 50:
            # Wrap text at reasonable points (spaces, punctuation)
            words = text.split()
            lines = []
            current_line = []
            current_length = 0
            
            for word in words:
                if current_length + len(word) + 1 <= 60:  # Target line length
                    current_line.append(word)
                    current_length += len(word) + 1
                else:
                    if current_line:
                        lines.append(' '.join(current_line))
                    current_line = [word]
                    current_length = len(word)
            
            if current_line:
                lines.append(' '.join(current_line))
            
            return '<br>'.join(lines)
        
        return text
        
    def extract_all_sheets(self, output_dir=None, session_id=None, file_id=None):
        """
        Extract data from all sheets in Excel file with smart table detection and images
        
        Args:
            output_dir: Directory to save extracted images
            session_id: Session ID for image URL paths
            file_id: File ID for image URL paths
            
        Returns:
            dict: Dictionary with sheet names as keys and data as values
        """
        try:
            logger.info(f"Extracting data from Excel file: {self.filename}")
            
            # Load workbook for image extraction (.xlsx format)
            if self.workbook is None:
                self.workbook = openpyxl.load_workbook(self.filepath, data_only=False)
            
            # Read all sheets
            excel_file = pd.ExcelFile(self.filepath)
            sheet_names = excel_file.sheet_names
            
            logger.info(f"Found {len(sheet_names)} sheets: {sheet_names}")
            
            results = {}
            for sheet_name in sheet_names:
                try:
                    sheet_data = self.extract_sheet(sheet_name, output_dir=output_dir, session_id=session_id, file_id=file_id)
                    
                    # Only include sheets that have data
                    if not sheet_data['empty']:
                        results[sheet_name] = sheet_data
                        logger.info(f"Sheet '{sheet_name}': {sheet_data['shape'][0]} rows, {sheet_data['shape'][1]} columns, {sheet_data.get('image_count', 0)} images")
                    else:
                        logger.info(f"Sheet '{sheet_name}': Skipped (empty or no valid data)")
                    
                except Exception as e:
                    logger.error(f"Error processing sheet '{sheet_name}': {e}")
                    results[sheet_name] = {
                        'error': str(e),
                        'data': [],
                        'html': f'<p>Error reading sheet: {str(e)}</p>',
                        'markdown': f'Error reading sheet: {str(e)}',
                        'images': {}
                    }
            
            return results
            
        except Exception as e:
            logger.error(f"Error reading Excel file {self.filename}: {e}")
            raise
    
    def _detect_table_start(self, df):
        """
        Detect where the actual table starts by finding the header row
        Looks for rows with multiple non-empty values that look like headers
        """
        # Common header keywords for BOQ/offer tables
        header_keywords = [
            'sn', 's.n', 'serial', 'item', 'description', 'desc', 
            'quantity', 'qty', 'unit', 'rate', 'price', 'amount', 
            'total', 'location', 'image', 'indicative', 'material'
        ]
        
        for idx, row in df.iterrows():
            # Count non-null values
            non_null = row.notna().sum()
            
            # Check if this row looks like a header
            if non_null >= 3:  # At least 3 columns should have values
                row_str = ' '.join(str(val).lower() for val in row if pd.notna(val))
                
                # Check if any header keywords are present
                keyword_count = sum(1 for keyword in header_keywords if keyword in row_str)
                
                if keyword_count >= 2:  # At least 2 header keywords found
                    logger.info(f"Detected table header at row {idx}: {row.tolist()}")
                    return idx
        
        # If no clear header found, return 0
        return 0
    
    def _clean_dataframe(self, df):
        """
        Clean dataframe by removing empty rows and columns while preserving _excel_row
        """
        # Store _excel_row if it exists
        excel_rows = df['_excel_row'].copy() if '_excel_row' in df.columns else None
        
        # Store column names before cleaning
        original_columns = df.columns.tolist()
        
        # Remove _excel_row temporarily for cleaning
        if '_excel_row' in df.columns:
            df = df.drop(columns=['_excel_row'])
            original_columns.remove('_excel_row')
        
        # Remove completely empty rows
        empty_mask = df.isna().all(axis=1)
        df = df[~empty_mask]
        
        # Remove completely empty columns, BUT preserve image columns
        # Image columns often appear empty because pandas can't read embedded images
        image_keywords = ['image', 'img', 'picture', 'photo', 'indicative']
        columns_to_keep = []
        
        for col in df.columns:
            col_lower = str(col).lower()
            # Keep if it's an image column or has any non-null values
            if any(keyword in col_lower for keyword in image_keywords) or df[col].notna().any():
                columns_to_keep.append(col)
        
        df = df[columns_to_keep]
        
        # Restore _excel_row if it existed
        if excel_rows is not None:
            df['_excel_row'] = excel_rows[~empty_mask].values
        
        return df
    
    def _is_valid_table_row(self, row, headers):
        """
        Check if a row is a valid data row (not a header repetition or empty)
        """
        # Skip _excel_row column if present
        actual_values = [val for val, col in zip(row, headers) if col != '_excel_row']
        
        # Convert row to strings for comparison
        row_str = [str(val).lower().strip() for val in actual_values]
        header_str = [str(h).lower().strip() for h in headers if h != '_excel_row']
        
        # Check if row is a header repetition
        if row_str == header_str:
            return False
        
        # Check if row has at least one non-empty value (very permissive)
        non_empty = sum(1 for val in actual_values if pd.notna(val) and str(val).strip() not in ['', 'nan', 'none'])
        
        return non_empty >= 1  # At least 1 column should have a value
    
    def extract_sheet(self, sheet_name=0, output_dir=None, session_id=None, file_id=None):
        """
        Extract data from specific sheet with smart table detection and image extraction
        
        Args:
            sheet_name: Sheet name (str) or index (int). Default is first sheet (0)
            output_dir: Directory to save extracted images
            session_id: Session ID for image URL paths
            file_id: File ID for image URL paths
            
        Returns:
            dict: Extracted data in multiple formats
        """
        try:
            # Load workbook with openpyxl for image extraction
            if self.workbook is None:
                self.workbook = openpyxl.load_workbook(self.filepath, data_only=False)
            
            # Get the worksheet
            if isinstance(sheet_name, int):
                ws = self.workbook.worksheets[sheet_name]
                actual_sheet_name = ws.title
            else:
                ws = self.workbook[sheet_name]
                actual_sheet_name = sheet_name
            
            # Extract images if output_dir provided
            cell_images = {}
            if output_dir:
                cell_images = self._extract_images_from_sheet(ws, output_dir)
                
                # Update image paths to include session/file_id for web access
                if session_id and file_id and cell_images:
                    updated_images = {}
                    for row_num, img_paths in cell_images.items():
                        updated_images[row_num] = [
                            f"/outputs/{session_id}/{file_id}/{path}" for path in img_paths
                        ]
                    cell_images = updated_images
            
            # Read with header=None to get raw data first
            # Use xlrd engine for .xls files
            if self.extension == '.xls':
                df_raw = pd.read_excel(self.filepath, sheet_name=actual_sheet_name, header=None, engine='xlrd')
            else:
                df_raw = pd.read_excel(self.filepath, sheet_name=actual_sheet_name, header=None)
            
            # Detect where the table starts
            header_row = self._detect_table_start(df_raw)
            
            logger.info(f"Detected header at row {header_row}")
            
            # Re-read with proper header
            # Use xlrd engine for .xls files
            if self.extension == '.xls':
                if header_row > 0:
                    df = pd.read_excel(self.filepath, sheet_name=sheet_name, header=header_row, engine='xlrd')
                else:
                    df = pd.read_excel(self.filepath, sheet_name=sheet_name, engine='xlrd')
            else:
                if header_row > 0:
                    df = pd.read_excel(self.filepath, sheet_name=sheet_name, header=header_row)
                else:
                    df = pd.read_excel(self.filepath, sheet_name=sheet_name)
            
            # Store the actual Excel row number for each dataframe row
            # This is critical for image mapping
            df['_excel_row'] = range(header_row + 2, header_row + 2 + len(df))
            
            # Clean the dataframe
            df = self._clean_dataframe(df)
            
            if df.empty:
                logger.warning(f"Sheet '{sheet_name}' is empty after cleaning")
                return {
                    'data': [],
                    'html': '<p>No data found</p>',
                    'markdown': 'No data found',
                    'columns': [],
                    'shape': (0, 0),
                    'empty': True,
                    'sheet_name': actual_sheet_name,
                    'images': {}
                }
            
            # Get headers
            headers = df.columns.tolist()
            
            # Filter valid data rows
            valid_rows = []
            valid_excel_rows = []
            for idx, row in df.iterrows():
                if self._is_valid_table_row(row.values, headers):
                    # Store both the row data and the excel row number
                    row_without_excel = {k: v for k, v in row.items() if k != '_excel_row'}
                    valid_rows.append(row_without_excel)
                    # Keep track of excel row numbers
                    if '_excel_row' in row:
                        valid_excel_rows.append(row['_excel_row'])
                    else:
                        valid_excel_rows.append(idx + header_row + 2)
            
            if valid_rows:
                df = pd.DataFrame(valid_rows)
                # Restore _excel_row column
                df['_excel_row'] = valid_excel_rows
            
            # Reset index
            df = df.reset_index(drop=True)
            
            # Enhance data with images and preserve full text
            enhanced_data = []
            for idx, row in df.iterrows():
                row_dict = {}
                
                # Get the actual Excel row number
                excel_row = row.get('_excel_row', idx + header_row + 2)
                
                # Debug: Check if this row should have images
                if idx < 5:  # Log first 5 rows for debugging
                    logger.info(f"Row {idx}: excel_row={excel_row}, has_images={excel_row in cell_images}")
                
                for col_name, value in row.items():
                    # Skip internal columns
                    if col_name == '_excel_row':
                        continue
                    
                    # Check if this row has images
                    row_has_images = excel_row in cell_images
                    
                    # For columns that typically contain images (INDICATIVE IMAGE, IMAGE, etc.)
                    col_lower = str(col_name).lower()
                    is_image_column = any(keyword in col_lower for keyword in ['image', 'picture', 'photo', 'img'])
                    
                    if row_has_images and is_image_column:
                        # Add images from this row with click-to-enlarge functionality
                        img_html = ''.join([
                            f'<img src="{img}" class="table-thumbnail" '
                            f'style="max-width:80px; max-height:80px; cursor:pointer; margin:2px; object-fit:cover; border: 1px solid #ddd; border-radius: 4px;" '
                            f'onclick="openImageModal(this.src)" '
                            f'title="Click to enlarge" />'
                            for img in cell_images[excel_row]
                        ])
                        
                        # Combine with text if present
                        if pd.notna(value) and str(value).strip():
                            text_content = str(value).strip()
                            row_dict[col_name] = f"{img_html}<br>{text_content}"
                        else:
                            row_dict[col_name] = img_html
                    else:
                        # Preserve full text without aggressive wrapping
                        if pd.notna(value):
                            text = str(value).strip()
                            # Only wrap extremely long text (>200 chars)
                            if len(text) > 200:
                                row_dict[col_name] = self._wrap_text(text, max_length=120)
                            else:
                                row_dict[col_name] = text
                        else:
                            row_dict[col_name] = ''
                
                enhanced_data.append(row_dict)
            
            # Debug: Log sample of enhanced data
            if idx < 2:  # Log first 2 rows
                logger.info(f"Enhanced row {idx} data keys: {list(row_dict.keys())}")
                if 'INDICATIVE IMAGE' in row_dict:
                    img_val = str(row_dict['INDICATIVE IMAGE'])[:150]
                    logger.info(f"Enhanced row {idx} INDICATIVE IMAGE: {img_val}...")
            
            # Remove _excel_row from headers if present
            headers_clean = [h for h in headers if h != '_excel_row']
            
            # Create enhanced DataFrame
            df_enhanced = pd.DataFrame(enhanced_data)
            
            logger.info(f"Sheet '{actual_sheet_name}': Found {len(df_enhanced)} valid rows with {len(headers_clean)} columns")
            logger.info(f"Headers: {headers_clean}")
            logger.info(f"Extracted {len(cell_images)} images from {len(set(cell_images.keys()))} rows")
            
            # Debug: Log first and last SN values
            if not df_enhanced.empty and 'SN' in df_enhanced.columns:
                first_sn = df_enhanced['SN'].iloc[0] if len(df_enhanced) > 0 else 'N/A'
                last_sn = df_enhanced['SN'].iloc[-1] if len(df_enhanced) > 0 else 'N/A'
                logger.info(f"SN range: first={first_sn}, last={last_sn}, total_rows={len(df_enhanced)}")
            
            # Debug: Check if images are in enhanced data
            has_img_col = 'INDICATIVE IMAGE' in df_enhanced.columns
            if has_img_col and len(df_enhanced) > 0:
                first_img_val = str(df_enhanced['INDICATIVE IMAGE'].iloc[0])[:100]
                logger.info(f"First row INDICATIVE IMAGE value: {first_img_val}...")
            
            # Generate HTML with embedded images
            html = df_enhanced.to_html(index=False, classes='table table-striped', escape=False, na_rep='')
            
            return {
                'data': enhanced_data,
                'html': html,
                'markdown': df[headers_clean].to_markdown(index=False),  # Markdown without HTML
                'columns': headers_clean,
                'shape': (len(df), len(headers_clean)),
                'empty': df.empty,
                'sheet_name': actual_sheet_name,
                'images': cell_images,
                'image_count': len(cell_images)
            }
            
        except Exception as e:
            logger.error(f"Error extracting sheet '{sheet_name}' from {self.filename}: {e}")
            raise
    
    def _wrap_text(self, text, max_length=60):
        """
        Wrap long text for better display
        
        Args:
            text: Text to wrap
            max_length: Maximum characters per line
            
        Returns:
            str: Wrapped text with HTML breaks
        """
        if not text or len(str(text)) <= max_length:
            return str(text)
        
        text = str(text).strip()
        words = text.split()
        lines = []
        current_line = []
        current_length = 0
        
        for word in words:
            if current_length + len(word) + 1 <= max_length:
                current_line.append(word)
                current_length += len(word) + 1
            else:
                if current_line:
                    lines.append(' '.join(current_line))
                current_line = [word]
                current_length = len(word)
        
        if current_line:
            lines.append(' '.join(current_line))
        
        return '<br>'.join(lines)
    
    def get_sheet_names(self):
        """
        Get list of all sheet names in Excel file
        
        Returns:
            list: Sheet names
        """
        try:
            excel_file = pd.ExcelFile(self.filepath)
            return excel_file.sheet_names
        except Exception as e:
            logger.error(f"Error getting sheet names from {self.filename}: {e}")
            raise
    
    def to_json(self, sheet_name=None):
        """
        Convert Excel data to JSON format
        
        Args:
            sheet_name: Specific sheet to convert, or None for all sheets
            
        Returns:
            str: JSON string
        """
        try:
            if sheet_name:
                data = self.extract_sheet(sheet_name)
                return json.dumps(data, indent=2, default=str)
            else:
                data = self.extract_all_sheets()
                return json.dumps(data, indent=2, default=str)
        except Exception as e:
            logger.error(f"Error converting to JSON: {e}")
            raise
    
    def validate_file(self):
        """
        Validate if file exists and is a valid Excel file
        
        Returns:
            tuple: (bool: is_valid, str: error_message)
        """
        if not os.path.exists(self.filepath):
            return False, "File not found"
        
        if self.extension not in ['.xlsx']:
            return False, f"Invalid file extension after conversion: {self.extension}"
        
        try:
            # Validate the (converted) xlsx file
            excel_file = pd.ExcelFile(self.filepath)
            logger.info(f"Successfully validated Excel file with {len(excel_file.sheet_names)} sheets")
            return True, "Valid Excel file"
        except Exception as e:
            import traceback
            error_details = traceback.format_exc()
            logger.error(f"Excel validation failed: {str(e)}")
            logger.error(f"Full traceback: {error_details}")
            return False, f"Cannot read Excel file: {str(e)}"
    
    def get_file_info(self):
        """
        Get basic information about Excel file
        
        Returns:
            dict: File information
        """
        try:
            excel_file = pd.ExcelFile(self.filepath)
            sheet_names = excel_file.sheet_names
            
            file_size = os.path.getsize(self.filepath)
            
            info = {
                'filename': self.filename,
                'filepath': self.filepath,
                'extension': self.extension,
                'size_bytes': file_size,
                'size_mb': round(file_size / (1024 * 1024), 2),
                'sheet_count': len(sheet_names),
                'sheet_names': sheet_names
            }
            
            return info
            
        except Exception as e:
            logger.error(f"Error getting file info: {e}")
            raise


def process_excel_file(filepath, output_dir=None, session_id=None, file_id=None):
    """
    Convenience function to process an Excel file
    
    Args:
        filepath: Path to Excel file
        output_dir: Directory to save extracted images (optional)
        session_id: Session ID for image URL paths (optional)
        file_id: File ID for image URL paths (optional)
        
    Returns:
        dict: Complete extraction results
    """
    processor = ExcelProcessor(filepath)
    
    # Validate file first
    is_valid, message = processor.validate_file()
    if not is_valid:
        return {
            'success': False,
            'error': message,
            'filepath': filepath
        }
    
    try:
        # Get file info
        file_info = processor.get_file_info()
        
        # Extract all sheets with images
        sheets_data = processor.extract_all_sheets(output_dir=output_dir, session_id=session_id, file_id=file_id)
        
        # Count total images across all sheets
        total_images = sum(sheet.get('image_count', 0) for sheet in sheets_data.values())
        
        return {
            'success': True,
            'file_info': file_info,
            'sheets': sheets_data,
            'sheet_count': len(sheets_data),
            'image_count': total_images,
            'message': f"Successfully extracted {len(sheets_data)} sheet(s) with {total_images} image(s)"
        }
        
    except Exception as e:
        logger.error(f"Error processing Excel file: {e}")
        return {
            'success': False,
            'error': str(e),
            'filepath': filepath
        }
