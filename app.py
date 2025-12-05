from flask import Flask, render_template, request, jsonify, send_file, session, send_from_directory, url_for
import logging
from flask_session import Session
import os
import base64
import requests
import json
import re
from werkzeug.utils import secure_filename
import uuid
from datetime import datetime, timedelta
import shutil
import time
from threading import Thread
import threading
import uuid as uuid_module
from utils.excel_processor import process_excel_file

app = Flask(__name__)

# In-memory storage for scraping events/status (for real-time preview)
scraping_status = {}
scraping_status_lock = threading.Lock()
app.config['SECRET_KEY'] = 'your-secret-key-here'
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['OUTPUT_FOLDER'] = 'outputs'
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB max file size
app.config['SESSION_TYPE'] = 'filesystem'
app.config['SESSION_PERMANENT'] = True
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=4)  # Session lasts 4 hours
Session(app)

# Basic logging - configure to log to both file and console
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('server.log', encoding='utf-8'),
        logging.StreamHandler()  # Also log to console
    ]
)
logger = logging.getLogger(__name__)

# PP-StructureV3 API Configuration
# Using API URL from official documentation (PP-StructureV3_API_en documentation.txt)
API_URL = "https://wfk3ide9lcd0x0k9.aistudio-hub.baidu.com/layout-parsing"
TOKEN = "031c87b3c44d16aa4adf6928bcfa132e23393afc"

ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg', 'xls', 'xlsx'}

# Create necessary directories
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['OUTPUT_FOLDER'], exist_ok=True)
os.makedirs('flask_session', exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

def cleanup_session_files():
    """Clean up all uploaded and extracted files for current session"""
    session_id = session.get('session_id')
    if session_id:
        session_upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
        session_output_dir = os.path.join(app.config['OUTPUT_FOLDER'], session_id)
        
        if os.path.exists(session_upload_dir):
            shutil.rmtree(session_upload_dir)
        if os.path.exists(session_output_dir):
            shutil.rmtree(session_output_dir)

def cleanup_old_files(hours=24):
    """Clean up files and directories older than specified hours"""
    cutoff_time = time.time() - (hours * 3600)
    cleaned = {'uploads': 0, 'outputs': 0, 'sessions': 0}
    logger.info(f"Starting cleanup of files older than {hours} hours")
    
    # Clean old upload directories
    for session_dir in os.listdir(app.config['UPLOAD_FOLDER']):
        dir_path = os.path.join(app.config['UPLOAD_FOLDER'], session_dir)
        if os.path.isdir(dir_path) and os.path.getmtime(dir_path) < cutoff_time:
            try:
                shutil.rmtree(dir_path)
                cleaned['uploads'] += 1
                logger.info(f"Cleaned old upload directory: {session_dir}")
            except Exception as e:
                logger.error(f"Error cleaning upload directory {session_dir}: {e}")
    
    # Clean old output directories
    for session_dir in os.listdir(app.config['OUTPUT_FOLDER']):
        dir_path = os.path.join(app.config['OUTPUT_FOLDER'], session_dir)
        if os.path.isdir(dir_path) and os.path.getmtime(dir_path) < cutoff_time:
            try:
                shutil.rmtree(dir_path)
                cleaned['outputs'] += 1
                logger.info(f"Cleaned old output directory: {session_dir}")
            except Exception as e:
                logger.error(f"Error cleaning output directory {session_dir}: {e}")
    
    # Clean old flask session files
    session_dir = 'flask_session'
    if os.path.exists(session_dir):
        for session_file in os.listdir(session_dir):
            file_path = os.path.join(session_dir, session_file)
            if os.path.isfile(file_path) and os.path.getmtime(file_path) < cutoff_time:
                try:
                    os.remove(file_path)
                    cleaned['sessions'] += 1
                    logger.info(f"Cleaned old session file: {session_file}")
                except Exception as e:
                    logger.error(f"Error cleaning session file {session_file}: {e}")
    
    return cleaned

def cleanup_other_sessions(current_session_id):
    """Clean up all sessions EXCEPT the current one"""
    cleaned = {'uploads': 0, 'outputs': 0, 'sessions': 0}
    
    # Clean other session upload directories
    if os.path.exists(app.config['UPLOAD_FOLDER']):
        for session_dir in os.listdir(app.config['UPLOAD_FOLDER']):
            if session_dir != current_session_id:
                dir_path = os.path.join(app.config['UPLOAD_FOLDER'], session_dir)
                if os.path.isdir(dir_path):
                    try:
                        shutil.rmtree(dir_path)
                        cleaned['uploads'] += 1
                        logger.info(f"Cleaned other session upload directory: {session_dir}")
                    except Exception as e:
                        logger.error(f"Error cleaning upload directory {session_dir}: {e}")
    
    # Clean other session output directories
    if os.path.exists(app.config['OUTPUT_FOLDER']):
        for session_dir in os.listdir(app.config['OUTPUT_FOLDER']):
            if session_dir != current_session_id:
                dir_path = os.path.join(app.config['OUTPUT_FOLDER'], session_dir)
                if os.path.isdir(dir_path):
                    try:
                        shutil.rmtree(dir_path)
                        cleaned['outputs'] += 1
                        logger.info(f"Cleaned other session output directory: {session_dir}")
                    except Exception as e:
                        logger.error(f"Error cleaning output directory {session_dir}: {e}")
    
    # Keep flask session files - they are needed for active sessions
    # Only clean files older than 24 hours
    session_dir = 'flask_session'
    if os.path.exists(session_dir):
        cutoff_time = time.time() - (24 * 3600)
        for session_file in os.listdir(session_dir):
            file_path = os.path.join(session_dir, session_file)
            if os.path.isfile(file_path) and os.path.getmtime(file_path) < cutoff_time:
                try:
                    os.remove(file_path)
                    cleaned['sessions'] += 1
                    logger.info(f"Cleaned old session file: {session_file}")
                except Exception as e:
                    logger.error(f"Error cleaning session file {session_file}: {e}")
    
    return cleaned

def cleanup_all_sessions():
    """Clean up ALL session data (aggressive cleanup on startup)"""
    cleaned = {'uploads': 0, 'outputs': 0, 'sessions': 0}
    
    # Clean ALL upload directories
    if os.path.exists(app.config['UPLOAD_FOLDER']):
        for session_dir in os.listdir(app.config['UPLOAD_FOLDER']):
            dir_path = os.path.join(app.config['UPLOAD_FOLDER'], session_dir)
            if os.path.isdir(dir_path):
                try:
                    shutil.rmtree(dir_path)
                    cleaned['uploads'] += 1
                    logger.info(f"Cleaned upload directory: {session_dir}")
                except Exception as e:
                    logger.error(f"Error cleaning upload directory {session_dir}: {e}")
    
    # Clean ALL output directories
    if os.path.exists(app.config['OUTPUT_FOLDER']):
        for item in os.listdir(app.config['OUTPUT_FOLDER']):
            dir_path = os.path.join(app.config['OUTPUT_FOLDER'], item)
            if os.path.isdir(dir_path):
                try:
                    shutil.rmtree(dir_path)
                    cleaned['outputs'] += 1
                    logger.info(f"Cleaned output directory: {item}")
                except Exception as e:
                    logger.error(f"Error cleaning output directory {item}: {e}")
    
    # Clean ALL flask session files
    session_dir = 'flask_session'
    if os.path.exists(session_dir):
        for session_file in os.listdir(session_dir):
            file_path = os.path.join(session_dir, session_file)
            if os.path.isfile(file_path):
                try:
                    os.remove(file_path)
                    cleaned['sessions'] += 1
                    logger.info(f"Cleaned session file: {session_file}")
                except Exception as e:
                    logger.error(f"Error cleaning session file {session_file}: {e}")
    
    return cleaned

def periodic_cleanup():
    """Run cleanup periodically in background"""
    while True:
        time.sleep(3600)  # Run every hour
        try:
            cleaned = cleanup_old_files(hours=24)
            logger.info(f"Periodic cleanup: {cleaned}")
        except Exception as e:
            logger.error(f"Error in periodic cleanup: {e}")

@app.before_request
def before_request():
    """Initialize session and ensure session directories exist"""
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
        # Only initialize uploaded_files if it doesn't exist
        if 'uploaded_files' not in session:
            session['uploaded_files'] = []
        logger.info(f"New session created: {session['session_id']}, Files in session: {len(session.get('uploaded_files', []))}")
        
        # Don't clean up on every new session - let periodic cleanup handle it
        # This prevents deleting files that users are still working with
        logger.info(f"New session started. Periodic cleanup will handle old files (runs every hour).")
    else:
        # For existing sessions, ensure uploaded_files exists
        if 'uploaded_files' not in session:
            session['uploaded_files'] = []
            logger.info(f"Initialized uploaded_files for existing session: {session['session_id']}")
        else:
            logger.debug(f"Existing session: {session['session_id']}, Files: {len(session['uploaded_files'])}")
    
    # Create session-specific directories
    session_id = session['session_id']
    os.makedirs(os.path.join(app.config['UPLOAD_FOLDER'], session_id), exist_ok=True)
    os.makedirs(os.path.join(app.config['OUTPUT_FOLDER'], session_id), exist_ok=True)

@app.route('/fix-session')
def fix_session_page():
    """Serve the session fix page"""
    return send_file('fix_session.html')

@app.route('/api/session-files', methods=['GET'])
def get_session_files():
    """Get list of files in current session"""
    try:
        uploaded_files = session.get('uploaded_files', [])
        session_id = session.get('session_id', 'none')
        
        # Return minimal file info for UI sync
        files_info = [{
            'id': f.get('id'),
            'name': f.get('original_name', 'Unknown'),
            'status': f.get('status', 'unknown'),
            'upload_time': f.get('upload_time', '')
        } for f in uploaded_files]
        
        logger.info(f"Session files request - Session: {session_id}, Files: {len(files_info)}")
        
        return jsonify({
            'success': True,
            'session_id': session_id,
            'files': files_info,
            'count': len(files_info)
        })
    except Exception as e:
        logger.exception("Error getting session files")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/debug/session')
def debug_session():
    """Debug page to show current session state"""
    try:
        uploaded_files = session.get('uploaded_files', [])
        session_id = session.get('session_id', 'No session ID')
        
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Session Debug</title>
            <style>
                body {{ font-family: Arial, sans-serif; padding: 20px; background: #f5f5f5; }}
                .container {{ max-width: 800px; margin: 0 auto; background: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                h1 {{ color: #333; }}
                .info {{ background: #e3f2fd; padding: 15px; border-radius: 4px; margin: 10px 0; }}
                .file {{ background: #f5f5f5; padding: 10px; margin: 10px 0; border-left: 3px solid #4CAF50; }}
                .empty {{ background: #fff3cd; padding: 15px; border-left: 3px solid #ffc107; color: #856404; }}
                code {{ background: #f4f4f4; padding: 2px 6px; border-radius: 3px; }}
                .btn {{ display: inline-block; padding: 10px 20px; background: #2196F3; color: white; text-decoration: none; border-radius: 4px; margin: 10px 5px 10px 0; }}
                .btn:hover {{ background: #0b7dda; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>🔍 Session Debug Information</h1>
                
                <div class="info">
                    <strong>Session ID:</strong> <code>{session_id}</code>
                </div>
                
                <h2>Uploaded Files ({len(uploaded_files)})</h2>
        """
        
        if uploaded_files:
            for idx, f in enumerate(uploaded_files):
                html += f"""
                <div class="file">
                    <strong>File #{idx + 1}</strong><br>
                    <strong>ID:</strong> <code>{f.get('id', 'N/A')}</code><br>
                    <strong>Name:</strong> {f.get('original_name', 'N/A')}<br>
                    <strong>Status:</strong> {f.get('status', 'N/A')}<br>
                    <strong>Upload Time:</strong> {f.get('upload_time', 'N/A')}<br>
                    <strong>Has Extraction:</strong> {'Yes' if 'extraction_result' in f else 'No'}
                </div>
                """
        else:
            html += '<div class="empty">⚠️ No files in session. Please upload a file.</div>'
        
        html += """
                <h2>Actions</h2>
                <a href="/" class="btn">Go to Main App</a>
                <a href="/debug/session" class="btn" onclick="location.reload(); return false;">Refresh</a>
                <a href="/cleanup" class="btn" onclick="if(confirm('Clear all files?')) {{ fetch('/cleanup', {{method: 'POST'}}).then(() => location.reload()); }} return false;">Clear Session</a>
                
                <p style="margin-top: 30px; color: #666; font-size: 0.9em;">
                    <strong>Tip:</strong> If you see files here but they don't appear in the main app, 
                    go back to the main app and refresh the page (Ctrl+F5).
                </p>
            </div>
        </body>
        </html>
        """
        
        return html
    except Exception as e:
        return f"<h1>Error</h1><pre>{str(e)}</pre>", 500

@app.route('/landing')
def landing():
    """Modern landing page"""
    return render_template('landing.html')

@app.route('/app')
def main_app():
    """Main application page"""
    # Clean up other sessions on every page load
    try:
        session_id = session.get('session_id')
        if session_id:
            cleaned = cleanup_other_sessions(session_id)
            if any(cleaned.values()):
                logger.info(f"Page load cleanup (removed other sessions): {cleaned}")
    except Exception as e:
        logger.error(f"Error in page load cleanup: {e}")
    
    return render_template('index.html', now=int(time.time()))

@app.route('/')
def index():
    """Home page - show landing page by default"""
    # Check if user wants to go directly to app
    if request.args.get('workflow') or request.args.get('app'):
        return render_template('index.html', now=int(time.time()))
    # Otherwise show landing page
    return render_template('landing.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    """Handle file upload"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        session_id = session['session_id']
        
        # Create unique filename
        unique_filename = f"{uuid.uuid4()}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], session_id, unique_filename)
        
        file.save(filepath)
        
        # Add to uploaded files list
        uploaded_files = session.get('uploaded_files', [])
        file_info = {
            'id': str(uuid.uuid4()),
            'original_name': filename,
            'unique_name': unique_filename,
            'filepath': filepath,
            'upload_time': datetime.now().isoformat(),
            'status': 'uploaded'
        }
        uploaded_files.append(file_info)
        session['uploaded_files'] = uploaded_files
        session.modified = True
        
        logger.info(f"File uploaded: {filename}, Session ID: {session_id}, Total files in session: {len(uploaded_files)}")
        
        return jsonify({
            'success': True,
            'file_id': file_info['id'],
            'filename': filename,
            'message': 'File uploaded successfully'
        })
    
    return jsonify({'error': 'Invalid file type'}), 400

@app.route('/upload-and-extract', methods=['POST'])
def upload_and_extract():
    """Handle file upload and automatically extract"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No selected file'}), 400
    
    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        session_id = session['session_id']
        
        # Create unique filename
        unique_filename = f"{uuid.uuid4()}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], session_id, unique_filename)
        
        file.save(filepath)
        
        # Get extraction settings from form data
        extraction_settings = {}
        if 'extraction_settings' in request.form:
            try:
                extraction_settings = json.loads(request.form['extraction_settings'])
            except:
                pass
        
        # Get extraction method from settings (default to pdfplumber for backward compatibility)
        extraction_method = extraction_settings.get('extractionMethod', 'pdfplumber')
        logger.info(f"Extraction method selected: {extraction_method}")
        
        # Add to uploaded files list
        uploaded_files = session.get('uploaded_files', [])
        
        # Check for duplicate uploads (same filename within last 10 seconds)
        current_time = datetime.now()
        is_duplicate = False
        existing_file_index = -1
        for idx, existing_file in enumerate(uploaded_files):
            if existing_file.get('original_name') == filename:
                upload_time_str = existing_file.get('upload_time', '')
                try:
                    upload_time = datetime.fromisoformat(upload_time_str)
                    time_diff = (current_time - upload_time).total_seconds()
                    if time_diff < 10:  # Within 10 seconds
                        is_duplicate = True
                        existing_file_index = idx
                        logger.warning(f"Duplicate upload detected for {filename} within {time_diff:.1f}s, removing old file and creating new one")
                        break
                except:
                    pass
        
        # If duplicate found, remove the old one to avoid confusion
        if is_duplicate and existing_file_index >= 0:
            old_file = uploaded_files.pop(existing_file_index)
            logger.info(f"Removed old duplicate file_id: {old_file['id']}")
        
        # Always create new file info for each upload
        file_info = {
            'id': str(uuid.uuid4()),
            'original_name': filename,
            'unique_name': unique_filename,
            'filepath': filepath,
            'upload_time': datetime.now().isoformat(),
            'status': 'uploaded'
        }
        uploaded_files.append(file_info)
        session['uploaded_files'] = uploaded_files
        session.modified = True
        logger.info(f"File uploaded: {filename}, file_id: {file_info['id']}, Session ID: {session_id}, Total files: {len(uploaded_files)}")
        
        # Automatically extract the file using selected method (PP-Structure API or pdfplumber) or Excel processor
        extraction_result = None
        extraction_error = None
        try:
            logger.info(f"Starting automatic extraction for file: {filename}, file_id: {file_info['id']} using {extraction_method}")
            
            # Determine file type
            file_extension = filename.rsplit('.', 1)[1].lower()
            
            # Setup output directory
            session_id = session['session_id']
            output_dir = os.path.join(app.config['OUTPUT_FOLDER'], session_id, file_info['id'])
            images_dir = os.path.join(output_dir, 'imgs')
            os.makedirs(output_dir, exist_ok=True)
            os.makedirs(images_dir, exist_ok=True)
            
            # Handle Excel files differently
            if file_extension in ['xls', 'xlsx']:
                logger.info(f"Processing Excel file: {filename}")
                
                # Process Excel file with output directory for images
                excel_result = process_excel_file(filepath, output_dir=output_dir, session_id=session_id, file_id=file_info['id'])
                
                if not excel_result['success']:
                    raise Exception(excel_result.get('error', 'Excel processing failed'))
                
                # Convert Excel result to extraction_result format
                sheets_data = excel_result['sheets']
                
                # Create layoutParsingResults-like structure for compatibility
                layoutParsingResults = []
                
                for sheet_name, sheet_data in sheets_data.items():
                    # Save markdown for this sheet
                    md_filename = os.path.join(output_dir, f"{sheet_name.replace(' ', '_')}.md")
                    
                    markdown_content = f"# {sheet_name}\n\n"
                    if 'error' in sheet_data:
                        markdown_content += f"**Error:** {sheet_data['error']}\n\n"
                    else:
                        markdown_content += f"**Rows:** {sheet_data['shape'][0]}  \n"
                        markdown_content += f"**Columns:** {sheet_data['shape'][1]}  \n"
                        markdown_content += f"**Images:** {sheet_data.get('image_count', 0)}  \n\n"
                        if sheet_data.get('markdown'):
                            markdown_content += sheet_data['markdown']
                    
                    with open(md_filename, 'w', encoding='utf-8') as f:
                        f.write(markdown_content)
                    
                    # Add to layoutParsingResults
                    layoutParsingResults.append({
                        'sheet_name': sheet_name,
                        'markdown': {
                            'text': markdown_content,
                            'images': sheet_data.get('images', {})
                        },
                        'html': sheet_data.get('html', ''),
                        'shape': sheet_data.get('shape', [0, 0]),
                        'image_count': sheet_data.get('image_count', 0)
                    })
                
                # Create extraction result
                extraction_result = {
                    'layoutParsingResults': layoutParsingResults,
                    'file_type': 'excel',
                    'sheet_count': excel_result['sheet_count'],
                    'file_info': excel_result['file_info'],
                    'image_count': excel_result.get('image_count', 0)
                }
                
                logger.info(f"Excel extraction completed for file: {filename} with {excel_result.get('image_count', 0)} images")
            else:
                # Check if we should use PP-Structure API or pdfplumber
                if extraction_method == 'pp-structure':
                    # Use PP-Structure API for extraction
                    logger.info(f"Using PP-Structure API for extraction")
                    
                    # Call the extract API endpoint internally
                    # Read file and encode to base64
                    with open(filepath, 'rb') as file:
                        file_bytes = file.read()
                        file_data = base64.b64encode(file_bytes).decode('ascii')
                    
                    # Determine file type
                    file_type = 0 if file_extension == 'pdf' else 1
                    
                    headers = {
                        "Authorization": f"token {TOKEN}",
                        "Content-Type": "application/json"
                    }
                    
                    # Build payload with extraction settings
                    payload = {
                        "file": file_data,
                        "fileType": file_type,
                        "useDocPreprocessor": extraction_settings.get("useDocPreprocessor", False),
                        "useSealRecognition": extraction_settings.get("useSealRecognition", True),
                        "useTableRecognition": extraction_settings.get("useTableRecognition", True),
                        "useFormulaRecognition": extraction_settings.get("useFormulaRecognition", True),
                        "useChartRecognition": extraction_settings.get("useChartRecognition", True),
                        "useRegionDetection": extraction_settings.get("useRegionDetection", True),
                        "formatBlockContent": extraction_settings.get("formatBlockContent", True),
                        "useTextlineOrientation": extraction_settings.get("useTextlineOrientation", False),
                        "useDocOrientationClassify": extraction_settings.get("useDocOrientationClassify", False),
                        "useDocUnwarping": extraction_settings.get("useDocUnwarping", False),
                        "visualize": extraction_settings.get("visualize", True)
                    }
                    
                    # Add optional settings only if they are provided (not None)
                    optional_settings = [
                        "useWiredTableCellsTransToHtml", "useWirelessTableCellsTransToHtml",
                        "useTableOrientationClassify", "useOcrResultsWithTableCells",
                        "useE2eWiredTableRecModel", "useE2eWirelessTableRecModel",
                        "layoutThreshold", "layoutNms", "layoutUnclipRatio", "layoutMergeBboxesMode",
                        "textDetLimitSideLen", "textDetLimitType", "textDetThresh",
                        "textDetBoxThresh", "textDetUnclipRatio", "textRecScoreThresh",
                        "sealDetLimitSideLen", "sealDetLimitType", "sealDetThresh",
                        "sealDetBoxThresh", "sealDetUnclipRatio", "sealRecScoreThresh"
                    ]
                    
                    for setting in optional_settings:
                        value = extraction_settings.get(setting)
                        if value is not None:
                            payload[setting] = value
                    
                    logger.info(f'PP-Structure API extraction settings: {json.dumps({k: v for k, v in payload.items() if k != "file"}, indent=2)}')
                    
                    # Call API with retry logic
                    max_retries = 3
                    retry_delay = 2
                    last_error = None
                    
                    for attempt in range(max_retries):
                        try:
                            # Set timeout based on file size - increased for API processing time
                            file_size = os.path.getsize(filepath)
                            if file_size > 15 * 1024 * 1024:
                                timeout = 900  # 15 minutes for very large files
                            elif file_size > 10 * 1024 * 1024:
                                timeout = 720  # 12 minutes
                            elif file_size > 5 * 1024 * 1024:
                                timeout = 600  # 10 minutes
                            else:
                                timeout = 480  # 8 minutes for smaller files
                            
                            logger.info(f'Calling PP-Structure API (attempt {attempt + 1}/{max_retries}, timeout: {timeout}s)...')
                            
                            response = requests.post(
                                API_URL, 
                                json=payload, 
                                headers=headers, 
                                timeout=timeout,
                                stream=False
                            )
                            
                            # If we get a response, break out of retry loop
                            break
                            
                        except requests.exceptions.Timeout:
                            if attempt < max_retries - 1:
                                wait_time = retry_delay * (2 ** attempt)
                                logger.warning(f'Request timeout. Retrying in {wait_time} seconds... (attempt {attempt + 1}/{max_retries})')
                                time.sleep(wait_time)
                                continue
                            else:
                                raise
                        except requests.exceptions.ConnectionError as ce:
                            last_error = ce
                            error_str = str(ce)
                            if 'ConnectionResetError' in error_str or 'forcibly closed' in error_str or '10054' in error_str:
                                if attempt < max_retries - 1:
                                    wait_time = retry_delay * (2 ** attempt)
                                    logger.warning(f'Connection reset by server. Retrying in {wait_time} seconds... (attempt {attempt + 1}/{max_retries})')
                                    time.sleep(wait_time)
                                    continue
                                else:
                                    raise
                            else:
                                raise
                    
                    # Check if we have a response
                    if 'response' not in locals() or response is None:
                        raise Exception('No response received from PP-Structure API after retries')
                    
                    # Process API response
                    logger.info(f'PP-Structure API response status: {response.status_code}')
                    
                    if response.status_code == 200:
                        result = response.json().get("result")
                        
                        # Download and save images from API response
                        for i, res in enumerate(result.get("layoutParsingResults", [])):
                            # Save markdown
                            md_filename = os.path.join(output_dir, f"doc_{i}.md")
                            
                            # Check for images in markdown
                            markdown_data = res.get("markdown", {})
                            markdown_text = markdown_data.get("text", "")
                            images_dict = markdown_data.get("images", {})
                            
                            # Track successfully downloaded images for this page
                            page_downloaded_images = {}
                            
                            if images_dict:
                                logger.info(f"Page {i} has {len(images_dict)} images to download")
                                # Download images
                                for img_path, img_url in images_dict.items():
                                    # Retry logic for image downloads
                                    max_img_retries = 3
                                    download_success = False
                                    
                                    for img_attempt in range(max_img_retries):
                                        try:
                                            logger.info(f'Downloading image (attempt {img_attempt + 1}/{max_img_retries}): {img_url}')
                                            img_response = requests.get(img_url, timeout=60, stream=True)
                                            
                                            if img_response.status_code == 200:
                                                # Save image locally
                                                local_img_path = os.path.join(images_dir, os.path.basename(img_path))
                                                with open(local_img_path, 'wb') as img_file:
                                                    # Write in chunks to handle large images
                                                    for chunk in img_response.iter_content(chunk_size=8192):
                                                        if chunk:
                                                            img_file.write(chunk)
                                                
                                                # Verify file was saved
                                                if os.path.exists(local_img_path) and os.path.getsize(local_img_path) > 0:
                                                    # Create URL-safe path for serving
                                                    relative_img_path = f"imgs/{os.path.basename(img_path)}"
                                                    local_url = url_for('serve_output', session_id=session_id, filename=f"{file_info['id']}/{relative_img_path}")
                                                    
                                                    # Track successful download
                                                    page_downloaded_images[img_path] = local_url
                                                    
                                                    # Replace remote URL with local URL in markdown
                                                    markdown_text = markdown_text.replace(img_path, local_url)
                                                    
                                                    logger.info(f'✓ Downloaded image: {img_path} -> {local_img_path} ({os.path.getsize(local_img_path)} bytes)')
                                                    download_success = True
                                                    break
                                                else:
                                                    logger.warning(f'Image file is empty or not saved: {local_img_path}')
                                            else:
                                                logger.warning(f'Failed to download image {img_url}: HTTP {img_response.status_code}')
                                                if img_attempt < max_img_retries - 1:
                                                    time.sleep(1)
                                                    continue
                                        except requests.exceptions.Timeout:
                                            logger.warning(f'Image download timeout for {img_url} (attempt {img_attempt + 1}/{max_img_retries})')
                                            if img_attempt < max_img_retries - 1:
                                                time.sleep(2)
                                                continue
                                        except Exception as e:
                                            logger.error(f'Failed to download image {img_url}: {str(e)}')
                                            if img_attempt < max_img_retries - 1:
                                                time.sleep(1)
                                                continue
                                    
                                    if not download_success:
                                        logger.error(f'Failed to download image after {max_img_retries} attempts: {img_url}')
                            
                            # Update markdown with local image paths
                            res["markdown"]["text"] = markdown_text
                            res["markdown"]["images"] = page_downloaded_images
                            
                            # Save markdown to file
                            with open(md_filename, 'w', encoding='utf-8') as f:
                                f.write(markdown_text)
                        
                        # Store the result
                        extraction_result = {
                            'layoutParsingResults': result.get("layoutParsingResults", []),
                            'file_type': 'pdf' if file_type == 0 else 'image',
                            'extraction_method': 'pp-structure'
                        }
                        
                        logger.info(f"PP-Structure API extraction completed for file: {filename}")
                    else:
                        # API error
                        error_msg = f"PP-Structure API returned error {response.status_code}"
                        logger.error(error_msg)
                        raise Exception(error_msg)
                else:
                    # Use improved_table_extractor (pdfplumber) for PDF/images
                    from utils.improved_table_extractor import ImprovedTableExtractor
                    extractor = ImprovedTableExtractor()
                    
                    # Extract tables with images using pdfplumber
                    extraction_result = extractor.extract_tables(
                        file_path=filepath,
                        file_extension=file_extension,
                        output_dir=output_dir,
                        bordered_method='pdfplumber',
                        borderless_method='pdfplumber',
                        ai_strategy='auto'
                    )
                    
                    logger.info(f"pdfplumber extraction completed for file: {filename}")
            
            logger.info(f"Extraction completed using {extraction_method}")
            logger.info(f"Extraction result structure - keys: {list(extraction_result.keys()) if isinstance(extraction_result, dict) else 'Not a dict'}")
            
            # Log image URLs for debugging
            if 'layoutParsingResults' in extraction_result:
                for page_idx, page_result in enumerate(extraction_result['layoutParsingResults']):
                    markdown_data = page_result.get('markdown', {})
                    if isinstance(markdown_data, dict):
                        markdown_text = markdown_data.get('text', '')
                        # Count image tags
                        import re
                        img_tags = re.findall(r'<img[^>]+src="([^"]+)"', markdown_text)
                        if img_tags:
                            logger.info(f"Page {page_idx + 1}: Found {len(img_tags)} image(s)")
                            for img_url in img_tags[:3]:  # Log first 3 images
                                logger.info(f"  Using local image: {img_url.replace('/outputs/' + session_id + '/' + file_info['id'] + '/', '')} -> {img_url}")
            
            # Store output directory for later use
            file_info['output_dir'] = output_dir
            
            # Update file info with extraction result
            file_info['status'] = 'extracted'
            file_info['extraction_result'] = extraction_result
            
            # Update in session
            for i, f in enumerate(uploaded_files):
                if f['id'] == file_info['id']:
                    uploaded_files[i] = file_info
                    break
            session['uploaded_files'] = uploaded_files
            session.modified = True
            
            logger.info(f"Extraction successful for file: {filename}")
            
        except Exception as e:
            logger.exception(f"Error during automatic extraction for {filename}")
            extraction_error = str(e)
            # File is still uploaded, just extraction failed
            file_info['status'] = 'uploaded'
            file_info['extraction_error'] = extraction_error
        
        return jsonify({
            'success': True,
            'file_id': file_info['id'],
            'filename': filename,
            'message': 'File uploaded successfully',
            'extraction_success': extraction_result is not None,
            'extraction_error': extraction_error
        })
    
    return jsonify({'error': 'Invalid file type'}), 400

def convert_costed_data_to_html(costed_data):
    """Convert costed_data (headers/rows format) to HTML table format"""
    try:
        from bs4 import BeautifulSoup
        
        tables = costed_data.get('tables', [])
        if not tables:
            return ''
        
        # Use first table (or combine all tables)
        table_data = tables[0] if tables else {}
        headers = table_data.get('headers', [])
        rows = table_data.get('rows', [])
        
        if not headers or not rows:
            return ''
        
        # Build HTML table
        html = '<table border="1" style="width: 100%; border-collapse: collapse; background: white;">'
        html += '<thead><tr>'
        
        for header in headers:
            # Skip Action column
            if header.lower() in ['action', 'actions', 'product selection', 'productselection']:
                continue
            html += f'<th style="background: #1a365d; color: #d4af37; padding: 12px; border: 1px solid #d4af37; text-align: center;">{header}</th>'
        
        html += '</tr></thead><tbody>'
        
        for row in rows:
            html += '<tr>'
            for header in headers:
                # Skip Action column
                if header.lower() in ['action', 'actions', 'product selection', 'productselection']:
                    continue
                cell_value = row.get(header, '')
                # If cell value contains HTML (like images), use it directly, otherwise escape
                if '<img' in str(cell_value) or '<' in str(cell_value):
                    html += f'<td style="padding: 8px; border: 1px solid #ddd;">{cell_value}</td>'
                else:
                    # Escape HTML
                    escaped_value = str(cell_value).replace('<', '&lt;').replace('>', '&gt;').replace('&', '&amp;')
                    html += f'<td style="padding: 8px; border: 1px solid #ddd;">{escaped_value}</td>'
            html += '</tr>'
        
        html += '</tbody></table>'
        return html
    except Exception as e:
        logger.exception('Error converting costed_data to HTML')
        return ''

@app.route('/api/files/list', methods=['GET'])
def list_files():
    """List uploaded files for current session (for multi-budget feature)"""
    try:
        uploaded_files = session.get('uploaded_files', [])
        # Filter out sensitive data and only return file metadata with stitched_table
        # Exclude multibudget export files (they are generated, not uploaded)
        files_list = []
        for f in uploaded_files:
            # Skip multibudget export files - these are generated by the app, not user uploads
            if f.get('multibudget', False):
                continue
            
            file_data = {
                'id': f.get('id'),
                'original_name': f.get('original_name'),
                'upload_time': f.get('upload_time'),
                'has_extraction': 'extraction_result' in f,
                'has_stitched': 'stitched_table' in f,
                'has_costed': 'costed_data' in f
            }
            # Include stitched_table HTML if available
            if 'stitched_table' in f:
                file_data['stitched_table'] = {
                    'html': f['stitched_table'].get('html', '')
                }
            # Include costed_data with HTML conversion if available
            if 'costed_data' in f:
                costed_data = f['costed_data']
                # Convert costed_data (headers/rows format) to HTML table
                costed_html = convert_costed_data_to_html(costed_data)
                file_data['costed_table'] = {
                    'html': costed_html
                }
            files_list.append(file_data)
        
        return jsonify({
            'success': True,
            'files': files_list
        })
    except Exception as e:
        logger.exception('Error listing files')
        return jsonify({'error': str(e)}), 500

@app.route('/files', methods=['GET'])
def get_uploaded_files():
    """Get list of uploaded files (excludes multibudget export files)"""
    try:
        uploaded_files = session.get('uploaded_files', [])
        logger.info(f"Session ID: {session.get('session_id', 'N/A')}, Found {len(uploaded_files)} files in session")
        
        # Filter out multibudget export files - these are generated, not uploaded
        user_uploaded_files = [f for f in uploaded_files if not f.get('multibudget', False)]
        
        # Also check if files still exist on disk and validate them
        validated_files = []
        session_id = session.get('session_id')
        if session_id:
            upload_dir = os.path.join(app.config['UPLOAD_FOLDER'], session_id)
            if os.path.exists(upload_dir):
                for file_info in user_uploaded_files:
                    filepath = file_info.get('filepath')
                    if filepath and os.path.exists(filepath):
                        validated_files.append(file_info)
                    else:
                        logger.warning(f"File not found on disk: {filepath}")
        
        logger.info(f"Returning {len(validated_files)} validated files")
        return jsonify({'files': validated_files})
    except Exception as e:
        logger.exception('Error getting uploaded files')
        return jsonify({'files': [], 'error': str(e)}), 500

@app.route('/delete-file/<file_id>', methods=['DELETE'])
def delete_file(file_id):
    """Delete a specific uploaded file"""
    uploaded_files = session.get('uploaded_files', [])
    file_to_delete = None
    
    for file_info in uploaded_files:
        if file_info['id'] == file_id:
            file_to_delete = file_info
            break
    
    if file_to_delete:
        # Delete physical file
        if os.path.exists(file_to_delete['filepath']):
            os.remove(file_to_delete['filepath'])
        
        # Remove from session
        uploaded_files.remove(file_to_delete)
        session['uploaded_files'] = uploaded_files
        
        return jsonify({'success': True, 'message': 'File deleted'})
    
    return jsonify({'error': 'File not found'}), 404

@app.route('/cleanup', methods=['POST'])
def cleanup():
    """Clean up all session files"""
    cleanup_session_files()
    session['uploaded_files'] = []
    return jsonify({'success': True, 'message': 'All files cleaned up'})

@app.route('/clear-session', methods=['POST'])
def clear_session():
    """Force clear all session data - useful after cleanup issues"""
    try:
        session_id = session.get('session_id')
        if session_id:
            cleanup_session_files()
        session.clear()
        # Generate new session ID
        session['session_id'] = str(uuid.uuid4())
        logger.info(f'Session cleared and reset. New session ID: {session["session_id"]}')
        return jsonify({'success': True, 'message': 'Session cleared successfully', 'new_session_id': session['session_id']})
    except Exception as e:
        logger.exception('Error clearing session')
        return jsonify({'error': str(e)}), 500

@app.route('/preprocess/<file_id>', methods=['POST'])
def preprocess_file(file_id):
    """Preprocess PDF to detect and stitch tables"""
    from utils.pdf_processor import PDFProcessor
    
    uploaded_files = session.get('uploaded_files', [])
    file_info = None
    
    for f in uploaded_files:
        if f['id'] == file_id:
            file_info = f
            break
    
    if not file_info:
        return jsonify({'error': 'File not found'}), 404
    
    try:
        processor = PDFProcessor()
        result = processor.preprocess_pdf(file_info['filepath'], session['session_id'])

        # Convert local output paths to URLs that the frontend can fetch
        session_id = session['session_id']
        # stitched_image
        stitched_local = result.get('stitched_image')
        if stitched_local:
            # stitched_local is like outputs/<session_id>/preprocessing/stitched_table.jpg
            stitched_rel = os.path.relpath(stitched_local, os.path.join(app.config['OUTPUT_FOLDER'], session_id))
            result['stitched_image_url'] = url_for('serve_output', session_id=session_id, filename=stitched_rel)

        # thumbnails
        thumbs = result.get('thumbnails', [])
        thumb_urls = []
        for t in thumbs:
            thumb_local = t.get('path')
            if thumb_local:
                thumb_rel = os.path.relpath(thumb_local, os.path.join(app.config['OUTPUT_FOLDER'], session_id))
                t['path_url'] = url_for('serve_output', session_id=session_id, filename=thumb_rel)
            thumb_urls.append(t)

        result['thumbnails'] = thumb_urls

        # Update file status
        file_info['status'] = 'preprocessed'
        file_info['preprocessed_data'] = result
        session.modified = True

        return jsonify({
            'success': True,
            'result': result,
            'message': 'Preprocessing completed'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/outputs/<session_id>/<path:filename>')
def serve_output(session_id, filename):
    """Serve files from the outputs directory for the given session."""
    base_dir = os.path.join(app.config['OUTPUT_FOLDER'], session_id)
    # Security: prevent path traversal by resolving real path
    full_path = os.path.realpath(os.path.join(base_dir, filename))
    if not full_path.startswith(os.path.realpath(base_dir)):
        return jsonify({'error': 'Invalid file path'}), 400
    if not os.path.exists(full_path):
        return jsonify({'error': 'File not found'}), 404
    # send_from_directory expects directory and filename relative to it
    rel_dir = os.path.dirname(filename)
    rel_file = os.path.basename(filename)
    return send_from_directory(os.path.join(app.config['OUTPUT_FOLDER'], session_id, rel_dir), rel_file)

def handle_excel_extraction(file_id, file_info):
    """
    Handle Excel file extraction using excel_processor
    
    Args:
        file_id: Unique file identifier
        file_info: File information dictionary
        
    Returns:
        JSON response with extraction results
    """
    try:
        # Create output directory first
        session_id = session['session_id']
        output_dir = os.path.join(app.config['OUTPUT_FOLDER'], session_id, file_id)
        os.makedirs(output_dir, exist_ok=True)
        
        # Create imgs subdirectory for extracted images
        imgs_dir = os.path.join(output_dir, 'imgs')
        os.makedirs(imgs_dir, exist_ok=True)
        
        # Process Excel file with image extraction
        result = process_excel_file(file_info['filepath'], output_dir=output_dir, session_id=session_id, file_id=file_id)
        
        if not result['success']:
            logger.error(f"Excel processing failed: {result.get('error')}")
            return jsonify({
                'error': 'Excel processing failed',
                'message': result.get('error'),
                'filepath': file_info['filepath']
            }), 500
        
        # Generate markdown for each sheet
        sheets_data = result['sheets']
        markdown_files = []
        
        for sheet_name, sheet_data in sheets_data.items():
            # Create markdown file for this sheet
            md_filename = os.path.join(output_dir, f"{sheet_name.replace(' ', '_')}.md")
            
            markdown_content = f"# {sheet_name}\n\n"
            
            if 'error' in sheet_data:
                markdown_content += f"**Error:** {sheet_data['error']}\n\n"
            else:
                # Add sheet info
                markdown_content += f"**Rows:** {sheet_data['shape'][0]}  \n"
                markdown_content += f"**Columns:** {sheet_data['shape'][1]}  \n"
                markdown_content += f"**Images:** {sheet_data.get('image_count', 0)}  \n\n"
                
                # Add the table in markdown format
                if sheet_data.get('markdown'):
                    markdown_content += sheet_data['markdown']
            
            # Save markdown file
            with open(md_filename, 'w', encoding='utf-8') as f:
                f.write(markdown_content)
            
            markdown_files.append({
                'sheet_name': sheet_name,
                'filename': os.path.basename(md_filename),
                'path': md_filename
            })
        
        # Save complete JSON result
        json_filename = os.path.join(output_dir, 'extraction_result.json')
        with open(json_filename, 'w', encoding='utf-8') as f:
            json.dump(result, f, indent=2, default=str)
        
        # Update file status in session
        file_info['status'] = 'extracted'
        file_info['extracted_data'] = {
            'sheet_count': result['sheet_count'],
            'sheets': list(sheets_data.keys()),
            'markdown_files': markdown_files,
            'file_info': result['file_info'],
            'image_count': result.get('image_count', 0)
        }
        session.modified = True
        
        # Return response in format compatible with frontend
        return jsonify({
            'success': True,
            'file_type': 'excel',
            'sheet_count': result['sheet_count'],
            'sheets': sheets_data,
            'file_info': result['file_info'],
            'markdown_files': markdown_files,
            'image_count': result.get('image_count', 0),
            'message': result['message'],
            'extraction_type': 'Excel Direct Read'
        })
        
    except Exception as e:
        logger.exception(f'Error in Excel extraction: {e}')
        return jsonify({
            'error': 'Excel extraction failed',
            'message': str(e),
            'filepath': file_info['filepath']
        }), 500

@app.route('/extract/<file_id>', methods=['POST'])
def extract_table(file_id):
    """Extract table using PP-StructureV3 API or Excel processor"""
    logger.info(f'Extract request received for file_id: {file_id}')
    
    try:
        uploaded_files = session.get('uploaded_files', [])
        logger.info(f'Found {len(uploaded_files)} files in session')
        
        file_info = None
        for f in uploaded_files:
            if f['id'] == file_id:
                file_info = f
                logger.info(f'File found: {file_info.get("original_name", "unknown")}')
                break
        
        if not file_info:
            logger.error(f'File not found in session: {file_id}')
            return jsonify({'error': 'File not found'}), 404
        
        if not os.path.exists(file_info['filepath']):
            logger.error(f'File path does not exist: {file_info["filepath"]}')
            return jsonify({'error': 'File not found on disk'}), 404
        
        logger.info(f'Starting extraction for file: {file_info["original_name"]} at {file_info["filepath"]}')
        
        # Determine file extension
        file_extension = file_info['original_name'].rsplit('.', 1)[1].lower()
        
        # Handle Excel files differently
        if file_extension in ['xls', 'xlsx']:
            logger.info(f'Processing Excel file: {file_info["original_name"]}')
            return handle_excel_extraction(file_id, file_info)
        
        # For PDF and image files, continue with PP-StructureV3 API
        # Read file and encode to base64
        file_size = os.path.getsize(file_info['filepath'])
        file_size_mb = file_size / (1024 * 1024)
        
        # Check file size - base64 encoding increases size by ~33%
        # Warn if file is larger than 30MB (base64 encoded will be ~40MB)
        if file_size > 30 * 1024 * 1024:
            logger.warning(f'Large file detected: {file_size_mb:.2f}MB')
        
        with open(file_info['filepath'], 'rb') as file:
            file_bytes = file.read()
            file_data = base64.b64encode(file_bytes).decode('ascii')
        
        # Log payload size for debugging
        payload_size = len(file_data.encode('utf-8'))
        payload_size_mb = payload_size / (1024 * 1024)
        logger.info(f'File size: {file_size_mb:.2f}MB, Payload size: {payload_size_mb:.2f}MB')
        
        # Determine file type
        file_extension = file_info['original_name'].rsplit('.', 1)[1].lower()
        file_type = 0 if file_extension == 'pdf' else 1
        
        headers = {
            "Authorization": f"token {TOKEN}",
            "Content-Type": "application/json"
        }
        
        # Get extraction settings from request or use defaults from example JSON
        # Default settings from borderless.png_by_PP-StructrueV3.json example
        settings = request.json if request.is_json else {}
        payload = {
            "file": file_data,
            "fileType": file_type,
            # Default settings from example JSON that produced clean extraction
            "useDocPreprocessor": settings.get("useDocPreprocessor", False),
            "useSealRecognition": settings.get("useSealRecognition", True),
            "useTableRecognition": settings.get("useTableRecognition", True),
            "useFormulaRecognition": settings.get("useFormulaRecognition", True),
            "useChartRecognition": settings.get("useChartRecognition", True),
            "useRegionDetection": settings.get("useRegionDetection", True),
            "formatBlockContent": settings.get("formatBlockContent", True),
            "useTextlineOrientation": settings.get("useTextlineOrientation", False),
            "useDocOrientationClassify": settings.get("useDocOrientationClassify", False),
            "useDocUnwarping": settings.get("useDocUnwarping", False),
            "visualize": settings.get("visualize", True)
        }
        
        # Add optional settings only if they are provided (not None)
        optional_settings = [
            "useWiredTableCellsTransToHtml", "useWirelessTableCellsTransToHtml",
            "useTableOrientationClassify", "useOcrResultsWithTableCells",
            "useE2eWiredTableRecModel", "useE2eWirelessTableRecModel",
            "layoutThreshold", "layoutNms", "layoutUnclipRatio", "layoutMergeBboxesMode",
            "textDetLimitSideLen", "textDetLimitType", "textDetThresh",
            "textDetBoxThresh", "textDetUnclipRatio", "textRecScoreThresh",
            "sealDetLimitSideLen", "sealDetLimitType", "sealDetThresh",
            "sealDetBoxThresh", "sealDetUnclipRatio", "sealRecScoreThresh"
        ]
        
        for setting in optional_settings:
            value = settings.get(setting)
            if value is not None:
                payload[setting] = value
        
        logger.info(f'Extraction settings: {json.dumps({k: v for k, v in payload.items() if k != "file"}, indent=2)}')
        
        # Retry logic with exponential backoff
        max_retries = 3
        retry_delay = 2  # seconds
        last_error = None
        
        for attempt in range(max_retries):
            try:
                # Increase timeout significantly for large files - give API enough time
                if file_size > 15 * 1024 * 1024:  # > 15MB
                    timeout = 900  # 15 minutes
                elif file_size > 10 * 1024 * 1024:  # > 10MB
                    timeout = 720  # 12 minutes
                elif file_size > 5 * 1024 * 1024:  # > 5MB
                    timeout = 600  # 10 minutes
                else:
                    timeout = 480  # 8 minutes for smaller files
                
                logger.info(f'Using timeout: {timeout}s for file size: {file_size / (1024*1024):.2f}MB')
                
                # Use simple post request matching the API example
                # Avoid session pooling which might cause connection resets
                logger.info(f'Attempting API request (attempt {attempt + 1}/{max_retries})...')
                logger.info(f'Payload size: {len(json.dumps(payload)) / 1024:.2f}KB')
                
                response = requests.post(
                    API_URL, 
                    json=payload, 
                    headers=headers, 
                    timeout=timeout,
                    stream=False  # Don't stream for JSON requests
                )
                
                # If we get a response, break out of retry loop
                break
                
            except requests.exceptions.ConnectionError as ce:
                last_error = ce
                error_str = str(ce)
                
                # Check if it's a connection reset error
                if 'ConnectionResetError' in error_str or 'forcibly closed' in error_str or '10054' in error_str:
                    if attempt < max_retries - 1:
                        wait_time = retry_delay * (2 ** attempt)
                        logger.warning(f'Connection reset by server. Retrying in {wait_time} seconds... (attempt {attempt + 1}/{max_retries})')
                        time.sleep(wait_time)
                        continue
                    else:
                        # Last attempt failed
                        logger.error(f'PP-StructureV3 API connection reset after {max_retries} attempts')
                        return jsonify({
                            'error': 'Connection reset by server',
                            'message': f'The API server closed the connection. This may be due to file size ({file_size_mb:.2f}MB) or server limits.',
                            'details': error_str,
                            'file_size_mb': round(file_size_mb, 2),
                            'payload_size_mb': round(payload_size_mb, 2),
                            'api_url': API_URL,
                            'hint': 'Try splitting large files or reducing file size. The API may have size limits.',
                            'attempts': max_retries
                        }), 502
                else:
                    # Other connection errors
                    raise
            except requests.exceptions.Timeout:
                if attempt < max_retries - 1:
                    wait_time = retry_delay * (2 ** attempt)
                    logger.warning(f'Request timeout. Retrying in {wait_time} seconds... (attempt {attempt + 1}/{max_retries})')
                    time.sleep(wait_time)
                    continue
                else:
                    raise
        else:
            # If we exhausted all retries
            if last_error:
                logger.error(f'All retry attempts failed. Last error: {last_error}')
                raise last_error
        
        # Check if we have a response object
        if 'response' not in locals() or response is None:
            logger.error('No response received from API after retries')
            return jsonify({
                'error': 'No response from API',
                'message': 'The API server did not respond after multiple attempts.',
                'hint': 'Please check your internet connection and API status.'
            }), 502
        
        # If we get here, we have a response - continue processing

        # Log status and small preview of response for debugging
        logger.info('PP-StructureV3 response status: %s', response.status_code)
        logger.info('PP-StructureV3 response content-type: %s', response.headers.get('Content-Type', 'unknown'))
        resp_text = None
        try:
            resp_text = response.text[:2000]
            logger.info('PP-StructureV3 response body (truncated): %s', resp_text)
        except Exception as e:
            logger.warning(f'Could not read response text: {e}')

        # Check if response is HTML (error page)
        content_type = response.headers.get('Content-Type', '')
        if 'text/html' in content_type:
            logger.error('API returned HTML instead of JSON - likely an error page or invalid endpoint')
            return jsonify({
                'error': 'API returned HTML instead of JSON',
                'message': 'The API server returned an HTML page instead of JSON. This may indicate an authentication error or invalid endpoint.',
                'status_code': response.status_code,
                'content_type': content_type,
                'body_preview': resp_text[:500] if resp_text else 'No response body',
                'hint': 'Check API_URL and TOKEN are correct. The API may be down or the endpoint may have changed.'
            }), 502

        if response.status_code == 200:
            try:
                result = response.json().get("result")
            except Exception as je:
                logger.exception('Failed to decode JSON from PP-StructureV3 response')
                return jsonify({'error': 'Invalid JSON from API', 'status_code': response.status_code, 'body': resp_text}), 502

            # Save results
            session_id = session['session_id']
            output_dir = os.path.join(app.config['OUTPUT_FOLDER'], session_id, file_id)
            images_dir = os.path.join(output_dir, 'imgs')
            os.makedirs(output_dir, exist_ok=True)
            os.makedirs(images_dir, exist_ok=True)

            # Download and save images from API response
            # Track successfully downloaded images
            successfully_downloaded_images = {}
            
            for i, res in enumerate(result.get("layoutParsingResults", [])):
                # Save markdown
                md_filename = os.path.join(output_dir, f"doc_{i}.md")
                
                # Check for images in markdown
                markdown_data = res.get("markdown", {})
                markdown_text = markdown_data.get("text", "")
                images_dict = markdown_data.get("images", {})
                
                logger.info(f'Page {i+1}: Found {len(images_dict)} images to download')
                
                # Download images and replace URLs with local paths
                for img_path, img_url in images_dict.items():
                    # Skip if already downloaded
                    if img_path in successfully_downloaded_images:
                        logger.info(f'Image already downloaded: {img_path}')
                        continue
                    
                    # Retry logic for image downloads
                    max_retries = 3
                    download_success = False
                    
                    for attempt in range(max_retries):
                        try:
                            # Increase timeout to 60 seconds for large images
                            logger.info(f'Downloading image (attempt {attempt + 1}/{max_retries}): {img_url}')
                            img_response = requests.get(img_url, timeout=60, stream=True)
                            
                            if img_response.status_code == 200:
                                # Save image locally
                                local_img_path = os.path.join(images_dir, os.path.basename(img_path))
                                with open(local_img_path, 'wb') as img_file:
                                    # Write in chunks to handle large images
                                    for chunk in img_response.iter_content(chunk_size=8192):
                                        if chunk:
                                            img_file.write(chunk)
                                
                                # Verify file was saved
                                if os.path.exists(local_img_path) and os.path.getsize(local_img_path) > 0:
                                    # Create URL-safe path for serving
                                    relative_img_path = f"imgs/{os.path.basename(img_path)}"
                                    local_url = url_for('serve_output', session_id=session_id, filename=f"{file_id}/{relative_img_path}")
                                    
                                    # Track successful download
                                    successfully_downloaded_images[img_path] = local_url
                                    
                                    # Replace remote URL with local URL in markdown
                                    markdown_text = markdown_text.replace(img_path, local_url)
                                    
                                    logger.info(f'✓ Downloaded image: {img_path} -> {local_img_path} ({os.path.getsize(local_img_path)} bytes)')
                                    download_success = True
                                    break
                                else:
                                    logger.warning(f'Image file is empty or not saved: {local_img_path}')
                            else:
                                logger.warning(f'Failed to download image {img_url}: HTTP {img_response.status_code}')
                                if attempt < max_retries - 1:
                                    time.sleep(1)  # Wait before retry
                                    continue
                        except requests.exceptions.Timeout:
                            logger.warning(f'Image download timeout for {img_url} (attempt {attempt + 1}/{max_retries})')
                            if attempt < max_retries - 1:
                                time.sleep(2)  # Wait before retry
                                continue
                        except requests.exceptions.ConnectionError as ce:
                            logger.warning(f'Image download connection error for {img_url}: {ce}')
                            if attempt < max_retries - 1:
                                time.sleep(2)  # Wait before retry
                                continue
                        except Exception as e:
                            logger.error(f'Failed to download image {img_url}: {str(e)}')
                            if attempt < max_retries - 1:
                                time.sleep(1)  # Wait before retry
                                continue
                    
                    if not download_success:
                        logger.error(f'✗ Failed to download image after {max_retries} attempts: {img_url}')
                
                # Also update block_content in prunedResult if it exists
                # ONLY replace URLs for successfully downloaded images
                pruned_result = res.get("prunedResult", {})
                parsing_res_list = pruned_result.get("parsing_res_list", [])
                for block in parsing_res_list:
                    if block.get("block_content"):
                        block_content = block["block_content"]
                        # Replace image paths in block content - ONLY for successfully downloaded images
                        for img_path, local_url in successfully_downloaded_images.items():
                            # Replace both the path and any remote URLs
                            block_content = block_content.replace(img_path, local_url)
                            # Also try to replace the remote URL if it exists
                            if img_path in images_dict:
                                remote_url = images_dict[img_path]
                                block_content = block_content.replace(remote_url, local_url)
                        block["block_content"] = block_content
                
                # Save updated markdown with UTF-8 encoding to handle Unicode characters
                with open(md_filename, "w", encoding='utf-8') as md_file:
                    md_file.write(markdown_text)

            # Log image download summary
            total_images_found = len(images_dict) if 'images_dict' in locals() else 0
            images_downloaded = len(successfully_downloaded_images)
            logger.info(f'Image download summary: {images_downloaded}/{total_images_found} images downloaded successfully')

            # Update file status
            file_info['status'] = 'extracted'
            file_info['extraction_result'] = result
            file_info['output_dir'] = output_dir
            file_info['images_downloaded'] = images_downloaded
            file_info['total_images'] = total_images_found
            session.modified = True

            return jsonify({
                'success': True,
                'result': result,
                'message': f'Extraction completed successfully. Downloaded {images_downloaded}/{total_images_found} images.',
                'images_downloaded': images_downloaded,
                'total_images': total_images_found
            })
        else:
            # Try to parse body for helpful details
            err_body = None
            try:
                err_body = response.json()
                logger.error('PP-StructureV3 API returned error %s: %s', response.status_code, err_body)
            except Exception as je:
                err_body = resp_text
                logger.error('PP-StructureV3 API returned error %s. Could not parse JSON. Body: %s', response.status_code, resp_text[:500] if resp_text else 'No response body')

            return jsonify({
                'error': 'API error',
                'message': f'The API server returned an error status code: {response.status_code}',
                'status_code': response.status_code,
                'body': err_body,
                'hint': 'The API request failed. Check the API status and your request parameters.'
            }), 502
        
    except requests.exceptions.Timeout:
        logger.error('PP-StructureV3 API request timed out after retries')
        return jsonify({
            'error': 'Request timeout',
            'message': 'The API request took too long. Please try again or check your file size.',
            'details': 'Timeout after retries'
        }), 504
    except requests.exceptions.ConnectionError as ce:
        error_str = str(ce)
        # Connection errors should already be handled in retry logic, but catch any that escape
        logger.error(f'PP-StructureV3 API connection error: {ce}')
        file_size_mb = os.path.getsize(file_info['filepath']) / (1024 * 1024) if 'file_info' in locals() else 0
        return jsonify({
            'error': 'Connection error',
            'message': 'Unable to connect to the API server. Please check your internet connection.',
            'details': error_str,
            'api_url': API_URL,
            'file_size_mb': round(file_size_mb, 2) if file_size_mb > 0 else None
        }), 502
    except requests.exceptions.HTTPError as he:
        logger.error(f'PP-StructureV3 API HTTP error: {he}')
        return jsonify({
            'error': 'HTTP error',
            'message': 'The API server returned an error.',
            'details': str(he),
            'status_code': getattr(he.response, 'status_code', None) if hasattr(he, 'response') else None
        }), 502
    except requests.exceptions.RequestException as re:
        logger.exception('Request to PP-StructureV3 API failed')
        error_type = type(re).__name__
        return jsonify({
            'error': 'Request error',
            'message': f'An error occurred while contacting the API: {error_type}',
            'details': str(re),
            'api_url': API_URL,
            'hint': 'Please check your internet connection and API configuration.'
        }), 502
    except Exception as e:
        logger.exception(f'Unexpected error during extraction: {type(e).__name__}: {str(e)}')
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f'Full traceback:\n{error_trace}')
        return jsonify({
            'error': 'Internal server error',
            'message': f'An unexpected error occurred: {str(e)}',
            'error_type': type(e).__name__,
            'hint': 'Please check the server logs for more details.'
        }), 500

@app.route('/stitch-tables/<file_id>', methods=['POST'])
def stitch_tables(file_id):
    """Stitch tables from multiple pages, keeping only one header and removing duplicates"""
    # Ensure session has an ID
    if 'session_id' not in session:
        session['session_id'] = str(uuid.uuid4())
        session.modified = True
    
    uploaded_files = session.get('uploaded_files', [])
    file_info = None
    
    # Debug logging
    logger.info(f"Stitch request for file_id: {file_id}")
    logger.info(f"Session ID: {session.get('session_id')}")
    logger.info(f"Total files in session: {len(uploaded_files)}")
    if uploaded_files:
        logger.info(f"Available file IDs: {[f.get('id') for f in uploaded_files]}")
    
    for f in uploaded_files:
        if f['id'] == file_id:
            file_info = f
            break
    
    if not file_info:
        logger.error(f"File not found: {file_id}. Available IDs: {[f.get('id') for f in uploaded_files]}")
        error_response = {
            'error': 'File not found in session',
            'details': 'The file may have been removed due to session expiration or cleanup. Please re-upload and extract the file.',
            'available_files': [{'id': f.get('id'), 'name': f.get('name', 'Unknown')} for f in uploaded_files] if uploaded_files else []
        }
        return jsonify(error_response), 404
    
    if 'extraction_result' not in file_info:
        return jsonify({'error': 'Please extract the tables first'}), 400
    
    try:
        result = file_info['extraction_result']
        
        # Import re for pattern matching
        import re
        
        # Log the structure we received for debugging
        logger.info(f"Extraction result type: {type(result)}")
        if isinstance(result, dict):
            logger.info(f"Extraction result keys: {list(result.keys())}")
        elif isinstance(result, list):
            logger.info(f"Extraction result is a list with {len(result)} items")
        
        # The API might return result wrapped in 'result' key, or directly
        if isinstance(result, dict) and 'result' in result:
            result = result['result']
            logger.info("Unwrapped 'result' key from extraction_result")
            if isinstance(result, dict):
                logger.info(f"After unwrapping, keys: {list(result.keys())}")
        
        # Try to get layoutParsingResults or tables data
        layout_parsing_results = []
        extraction_method = result.get('extraction_method', 'unknown')
        
        logger.info(f"Extraction method: {extraction_method}")
        
        # Handle pdfplumber extraction format (tables, images, etc.)
        if isinstance(result, dict) and 'tables' in result and extraction_method in ['pdfplumber', 'builtin']:
            logger.info(f"Detected pdfplumber format with 'tables' key")
            tables_data = result.get('tables', [])
            
            logger.info(f"Number of tables in extraction result: {len(tables_data)}")
            
            if not tables_data:
                return jsonify({
                    'error': 'No tables found to stitch',
                    'details': 'The extraction found no tables in the document.',
                    'hint': 'Make sure the PDF contains tables and try extracting again.'
                }), 400
            
            # Convert pdfplumber tables to HTML format for stitching
            all_tables = []
            main_header = None
            
            for table_idx, table in enumerate(tables_data):
                if not isinstance(table, dict):
                    logger.warning(f'Table {table_idx} is not a dict: {type(table)}')
                    continue
                
                logger.info(f'Table {table_idx} keys: {list(table.keys())}')
                
                # pdfplumber tables have 'headers' and 'rows', not 'html'
                headers = table.get('headers', [])
                rows = table.get('rows', [])
                
                if not headers or not rows:
                    logger.warning(f'Table {table_idx} missing headers or rows')
                    continue
                
                logger.info(f'Table {table_idx}: {len(headers)} columns, {len(rows)} rows')
                
                # Generate HTML from headers and rows
                if main_header is None:
                    # Create header row with <th> tags
                    header_cells = ''.join([f'<th>{h}</th>' for h in headers])
                    main_header = f'<tr>{header_cells}</tr>'
                    logger.info(f'Set main header from table {table_idx}')
                
                # Create data rows with <td> tags
                for row in rows:
                    # Ensure row has same number of cells as headers
                    while len(row) < len(headers):
                        row.append('')
                    
                    row_cells = ''.join([f'<td>{cell}</td>' for cell in row])
                    all_tables.append(f'<tr>{row_cells}</tr>')
            
            if not all_tables and not main_header:
                return jsonify({
                    'error': 'No table rows found',
                    'details': 'Could not extract table rows from the extraction result.',
                    'hint': 'The tables may be in an unexpected format. Try re-extracting the file.'
                }), 400
            
            # Don't remove duplicates - preserve all rows as they come from extraction
            # Each row should be unique based on its position in the table
            # Duplicate detection was causing rows with images to be incorrectly removed
            filtered_tables = all_tables
            
            # Create stitched HTML table
            stitched_html = f'<table class="table table-bordered editable-table">{main_header}{"".join(filtered_tables)}</table>'
            
            # Store in session
            stitched_filename = f"stitched_table_{file_id}.html"
            if 'shared_tables' not in session:
                session['shared_tables'] = {}
            session['shared_tables']['stitched_table'] = {
                'file_id': file_id,
                'html': stitched_html,
                'filepath': stitched_filename,
                'row_count': len(filtered_tables),
                'timestamp': datetime.now().isoformat()
            }
            
            session.modified = True
            
            logger.info(f'Stitched {len(filtered_tables)} rows from {len(tables_data)} tables (pdfplumber format)')
            
            return jsonify({
                'success': True,
                'stitched_html': stitched_html,
                'row_count': len(filtered_tables),
                'table_count': len(tables_data),
                'message': f'Successfully stitched {len(filtered_tables)} rows from {len(tables_data)} tables'
            })
        
        # Handle API-based extraction format (layoutParsingResults)
        if isinstance(result, dict):
            # Check if this is an Excel file - use HTML directly instead of parsing markdown
            if result.get('file_type') == 'excel':
                logger.info("Detected Excel file type - using HTML directly without markdown parsing")
                layout_parsing_results = result.get('layoutParsingResults', [])
                
                if not layout_parsing_results:
                    return jsonify({'error': 'No sheets found to stitch'}), 400
                
                # For Excel, use the HTML from each sheet directly
                all_html_rows = []
                main_header = None
                
                for sheet_idx, sheet_result in enumerate(layout_parsing_results):
                    sheet_html = sheet_result.get('html', '')
                    if sheet_html:
                        # Extract rows from HTML table
                        import re
                        # Find all <tr> tags
                        tr_pattern = r'<tr[^>]*>(.*?)</tr>'
                        rows = re.findall(tr_pattern, sheet_html, re.DOTALL | re.IGNORECASE)
                        
                        if rows:
                            logger.info(f"Sheet {sheet_idx + 1}: Found {len(rows)} rows in HTML")
                            # First row is typically header
                            if main_header is None and rows:
                                main_header = f'<tr>{rows[0]}</tr>'
                                all_html_rows.extend([f'<tr>{row}</tr>' for row in rows[1:]])
                                logger.info(f"Set main header from sheet {sheet_idx + 1}, added {len(rows) - 1} data rows")
                            else:
                                # Skip header row in subsequent sheets
                                all_html_rows.extend([f'<tr>{row}</tr>' for row in rows[1:]])
                                logger.info(f"Added {len(rows) - 1} data rows from sheet {sheet_idx + 1}")
                
                if not main_header:
                    return jsonify({'error': 'No table data found in Excel sheets'}), 400
                
                # Create stitched HTML table
                stitched_html = f'<table class="table table-bordered editable-table">{main_header}{"".join(all_html_rows)}</table>'
                
                # Store in session
                stitched_filename = f"stitched_table_{file_id}.html"
                if 'shared_tables' not in session:
                    session['shared_tables'] = {}
                session['shared_tables']['stitched_table'] = {
                    'file_id': file_id,
                    'html': stitched_html,
                    'filepath': stitched_filename,
                    'row_count': len(all_html_rows),
                    'timestamp': datetime.now().isoformat()
                }
                
                session.modified = True
                
                logger.info(f'Stitched {len(all_html_rows)} Excel rows from {len(layout_parsing_results)} sheets using HTML')
                
                return jsonify({
                    'success': True,
                    'stitched_html': stitched_html,
                    'row_count': len(all_html_rows),
                    'table_count': len(layout_parsing_results),
                    'message': f'Successfully stitched {len(all_html_rows)} rows from {len(layout_parsing_results)} sheet(s)'
                })
            
            layout_parsing_results = result.get('layoutParsingResults', [])
            # Also try alternative key names
            if not layout_parsing_results:
                layout_parsing_results = result.get('layout_parsing_results', [])
            if not layout_parsing_results:
                layout_parsing_results = result.get('pages', [])
            if not layout_parsing_results:
                layout_parsing_results = result.get('results', [])
        elif isinstance(result, list):
            # Result might be directly a list of layout parsing results
            layout_parsing_results = result
            logger.info("Using result as direct list of layout parsing results")
        
        if not layout_parsing_results:
            logger.error(f"No layoutParsingResults found. Result type: {type(result)}")
            if isinstance(result, dict):
                logger.error(f"Available keys: {list(result.keys())}")
                # Log a sample of the structure for debugging
                for key, value in list(result.items())[:5]:
                    logger.error(f"  {key}: {type(value)} - {str(value)[:100] if not isinstance(value, (dict, list)) else '...'}")
            return jsonify({
                'error': 'No tables found to stitch',
                'details': f'No layoutParsingResults found in extraction result. Result type: {type(result).__name__}',
                'available_keys': list(result.keys()) if isinstance(result, dict) else None,
                'hint': 'The extraction may not have found any tables, or the result structure is different than expected. Please check the extraction was successful and try extracting again.'
            }), 400
        
        # Extract all tables from all pages
        all_tables = []
        main_header = None
        
        logger.info(f'Processing {len(layout_parsing_results)} pages for stitching')
        
        for page_idx, layout_result in enumerate(layout_parsing_results):
            if not isinstance(layout_result, dict):
                logger.warning(f'Page {page_idx + 1} layout_result is not a dict: {type(layout_result)}')
                continue
            
            # Also check markdown for tables (some APIs return tables in markdown format)
            markdown_data = layout_result.get('markdown', {})
            markdown_text = markdown_data.get('text', '') if isinstance(markdown_data, dict) else ''
            
            # Try to extract HTML tables from markdown first (PP-StructureV3 may embed HTML tables in markdown)
            if markdown_text:
                import re
                # Look for HTML tables in markdown
                html_table_pattern = r'<table[^>]*>(.*?)</table>'
                html_tables = re.findall(html_table_pattern, markdown_text, re.DOTALL | re.IGNORECASE)
                
                if html_tables:
                    logger.info(f'Page {page_idx + 1} has {len(html_tables)} HTML table(s) in markdown')
                    for table_idx, table_content in enumerate(html_tables):
                        # Extract rows from HTML table
                        rows = re.findall(r'<tr[^>]*>(.*?)</tr>', table_content, re.DOTALL | re.IGNORECASE)
                        if rows:
                            logger.info(f'Found HTML table {table_idx + 1} with {len(rows)} rows in markdown')
                            # Check if first row is header
                            first_row = rows[0]
                            is_header = '<th' in first_row.lower() or is_header_row(first_row)
                            
                            if main_header is None:
                                if is_header:
                                    main_header = f'<tr>{first_row}</tr>'
                                    logger.info(f'Set main header from HTML table in markdown (page {page_idx + 1})')
                                    all_tables.extend([f'<tr>{row}</tr>' for row in rows[1:]])
                                else:
                                    all_tables.extend([f'<tr>{row}</tr>' for row in rows])
                                logger.info(f'Added {len(rows)} rows from HTML table in markdown')
                            else:
                                # Skip header if present
                                start_idx = 1 if is_header and len(rows) > 1 else 0
                                all_tables.extend([f'<tr>{row}</tr>' for row in rows[start_idx:]])
                                logger.info(f'Added {len(rows) - start_idx} data rows from HTML table in markdown')
                
                # Also try markdown table format (| separators)
                if '|' in markdown_text:
                    logger.info(f'Page {page_idx + 1} has markdown text with potential markdown tables')
                    # Extract markdown tables - improved logic
                    lines = markdown_text.split('\n')
                    markdown_table_lines = []
                    in_table = False
                    
                    for line in lines:
                        line = line.strip()
                        # Check if line looks like a markdown table row
                        if '|' in line and line.count('|') >= 2:
                            # Skip separator rows (like |---|---|)
                            if not re.match(r'^[\|\s\-:]+$', line):
                                markdown_table_lines.append(line)
                                in_table = True
                        elif in_table and markdown_table_lines:
                            # End of table - process it
                            if len(markdown_table_lines) > 1:
                                logger.info(f'Found markdown table on page {page_idx + 1} with {len(markdown_table_lines)} rows')
                                try:
                                    from utils.costing_engine import CostingEngine
                                    engine = CostingEngine()
                                    table_data = engine.markdown_table_to_dict('\n'.join(markdown_table_lines))
                                    if table_data and 'rows' in table_data:
                                        # Convert to HTML rows
                                        rows_html = []
                                        # Header row
                                        if 'headers' in table_data:
                                            header_cells = ''.join([f'<th>{h}</th>' for h in table_data['headers']])
                                            rows_html.append(f'<tr>{header_cells}</tr>')
                                        # Data rows
                                        for row in table_data.get('rows', []):
                                            if isinstance(row, dict):
                                                row_cells = ''.join([f'<td>{row.get(h, "")}</td>' for h in table_data.get('headers', [])])
                                            else:
                                                row_cells = ''.join([f'<td>{cell}</td>' for cell in row])
                                            rows_html.append(f'<tr>{row_cells}</tr>')
                                        
                                        if rows_html:
                                            if main_header is None and 'headers' in table_data:
                                                main_header = rows_html[0]
                                            all_tables.extend(rows_html[1:] if len(rows_html) > 1 else rows_html)
                                            logger.info(f'Added {len(rows_html)} rows from markdown table on page {page_idx + 1}')
                                except Exception as e:
                                    logger.warning(f'Error parsing markdown table: {e}')
                            markdown_table_lines = []  # Reset for next table
                            in_table = False
                    
                    # Process last table if file ends with table
                    if in_table and markdown_table_lines and len(markdown_table_lines) > 1:
                        try:
                            from utils.costing_engine import CostingEngine
                            engine = CostingEngine()
                            table_data = engine.markdown_table_to_dict('\n'.join(markdown_table_lines))
                            if table_data and 'rows' in table_data:
                                rows_html = []
                                if 'headers' in table_data:
                                    header_cells = ''.join([f'<th>{h}</th>' for h in table_data['headers']])
                                    rows_html.append(f'<tr>{header_cells}</tr>')
                                for row in table_data.get('rows', []):
                                    if isinstance(row, dict):
                                        row_cells = ''.join([f'<td>{row.get(h, "")}</td>' for h in table_data.get('headers', [])])
                                    else:
                                        row_cells = ''.join([f'<td>{cell}</td>' for cell in row])
                                    rows_html.append(f'<tr>{row_cells}</tr>')
                                
                                if rows_html:
                                    if main_header is None and 'headers' in table_data:
                                        main_header = rows_html[0]
                                    all_tables.extend(rows_html[1:] if len(rows_html) > 1 else rows_html)
                                    logger.info(f'Added {len(rows_html)} rows from final markdown table on page {page_idx + 1}')
                        except Exception as e:
                            logger.warning(f'Error parsing final markdown table: {e}')
                
            pruned_result = layout_result.get('prunedResult', {})
            if not pruned_result:
                logger.warning(f'Page {page_idx + 1} has no prunedResult. Keys: {list(layout_result.keys())}')
                # Continue to next page even if no prunedResult (we might have found tables in markdown)
                continue
                
            parsing_res_list = pruned_result.get('parsing_res_list', [])
            
            logger.info(f'Processing page {page_idx + 1} with {len(parsing_res_list)} blocks')
            
            if not parsing_res_list:
                logger.warning(f'Page {page_idx + 1} has no parsing_res_list')
                continue
            
            # Log all block labels found on this page for debugging
            block_labels_found = []
            block_types_found = []
            for block in parsing_res_list:
                if isinstance(block, dict):
                    block_labels_found.append(block.get('block_label', 'unknown'))
                    block_types_found.append(block.get('block_type', 'unknown'))
            logger.info(f'Page {page_idx + 1} block labels: {block_labels_found}')
            logger.info(f'Page {page_idx + 1} block types: {block_types_found}')
            
            # Also log a sample of markdown text to see what's there
            if markdown_text:
                markdown_preview = markdown_text[:500].replace('\n', '\\n')
                logger.info(f'Page {page_idx + 1} markdown preview (first 500 chars): {markdown_preview}')
            
            for block_idx, block in enumerate(parsing_res_list):
                if not isinstance(block, dict):
                    logger.warning(f'Page {page_idx + 1}, block {block_idx} is not a dict')
                    continue
                    
                block_label = block.get('block_label', '')
                block_content = block.get('block_content', '')
                
                logger.debug(f'Page {page_idx + 1}, block {block_idx}: label={block_label}, has_content={bool(block_content)}')
                
                # Try multiple ways to detect tables
                is_table = False
                block_type = block.get('block_type', '')
                
                # Check block_label variations
                if block_label.lower() in ['table', 'table_block', 'table_cell'] and block_content:
                    is_table = True
                    logger.info(f'Found table with label: {block_label}')
                # Check block_type field (some APIs use this instead of block_label)
                elif block_type.lower() in ['table', 'table_block'] and block_content:
                    is_table = True
                    logger.info(f'Found table with type: {block_type}')
                # Check for HTML table content
                elif block_content and ('<table' in block_content.lower() or '<tr>' in block_content.lower() or '<td>' in block_content.lower()):
                    is_table = True
                    logger.info(f'Found table-like HTML content with label: {block_label}, type: {block_type}')
                # Check for markdown table content
                elif block_content and '|' in block_content and block_content.count('|') > 2:
                    is_table = True
                    logger.info(f'Found markdown table content with label: {block_label}, type: {block_type}')
                # Check if block has table-related keys
                elif 'table' in str(block).lower() and block_content:
                    # Last resort: check if any key contains "table"
                    table_keys = [k for k in block.keys() if 'table' in k.lower()]
                    if table_keys:
                        is_table = True
                        logger.info(f'Found table-related keys {table_keys} with label: {block_label}')
                
                if is_table:
                    table_html = block_content
                    
                    # Parse the table to extract header and rows
                    import re
                    from html.parser import HTMLParser
                    
                    # Extract table rows - try multiple patterns
                    rows = []
                    
                    # Try tbody pattern first
                    tbody_match = re.search(r'<tbody>(.*?)</tbody>', table_html, re.DOTALL)
                    if tbody_match:
                        tbody_content = tbody_match.group(1)
                        rows = re.findall(r'<tr>(.*?)</tr>', tbody_content, re.DOTALL)
                    else:
                        # Try direct tr pattern (no tbody)
                        rows = re.findall(r'<tr>(.*?)</tr>', table_html, re.DOTALL)
                        
                        logger.info(f'Found {len(rows)} rows in table on page {page_idx + 1}')
                        
                        if rows:
                            # First row is typically the header
                            first_row = rows[0]
                            
                            # Check if this is a header row (contains <th> or looks like a header)
                            is_header = '<th>' in first_row or is_header_row(first_row)
                            
                            if main_header is None:
                                # First table - keep header and all data rows
                                if is_header:
                                    main_header = first_row
                                    logger.info(f'Set main header from page {page_idx + 1}')
                                # Add all rows from first table
                                all_tables.extend(rows)
                                logger.info(f'Added {len(rows)} rows from first table (page {page_idx + 1})')
                            else:
                                # Subsequent tables - skip header, add only data rows
                                start_idx = 0
                                if is_header and len(rows) > 1:
                                    # Skip the header row
                                    start_idx = 1
                                    logger.info(f'Skipping header row on page {page_idx + 1}')
                                
                                # Add data rows
                                data_rows = rows[start_idx:]
                                all_tables.extend(data_rows)
                                logger.info(f'Added {len(data_rows)} data rows from page {page_idx + 1}')
        
        if not all_tables:
            logger.error(f'No tables found after processing {len(layout_parsing_results)} pages')
            return jsonify({
                'error': 'No tables found to stitch',
                'details': f'Processed {len(layout_parsing_results)} pages but found no table blocks with content',
                'hint': 'The document may not contain tables, or tables were not detected by the extraction API. Try checking the extraction settings or the document format.'
            }), 400
        
        logger.info(f'Total rows before filtering: {len(all_tables)}')
        
        # Filter out empty rows (rows with only whitespace or empty cells)
        import re
        filtered_tables = []
        for row in all_tables:
            # Remove HTML tags and check if there's actual content
            row_text = re.sub(r'<[^>]+>', '', row).strip()
            # Remove common whitespace characters and check if anything remains
            row_text_clean = re.sub(r'[\s\xa0\u00a0\u200b\u200c\u200d\ufeff]+', '', row_text)
            
            # Only include rows that have actual content
            if row_text_clean:
                filtered_tables.append(row)
            else:
                logger.debug(f'Filtered out empty row: {row[:100]}...')
        
        logger.info(f'Total rows after filtering empty rows: {len(filtered_tables)} (removed {len(all_tables) - len(filtered_tables)} empty rows)')
        
        # Build the stitched table HTML
        stitched_html = '''
<div style="text-align: center;">
    <html>
        <body>
            <table border="1">
'''
        
        # Add header row if available
        if main_header:
            stitched_html += '                <thead>\n'
            stitched_html += f'                    {main_header}\n'
            stitched_html += '                </thead>\n'
            logger.info('Added header row to stitched table')
        
        stitched_html += '                <tbody>\n'
        
        for row in filtered_tables:
            stitched_html += f'                    <tr>{row}</tr>\n'
        
        stitched_html += '''                </tbody>
            </table>
        </body>
    </html>
</div>
'''
        
        # Save stitched table
        session_id = session['session_id']
        output_dir = file_info.get('output_dir', os.path.join(app.config['OUTPUT_FOLDER'], session_id, file_id))
        os.makedirs(output_dir, exist_ok=True)
        
        stitched_filename = os.path.join(output_dir, 'stitched_table.html')
        with open(stitched_filename, 'w', encoding='utf-8') as f:
            f.write(stitched_html)
        
        # Update file info
        file_info['stitched_table'] = {
            'html': stitched_html,
            'filepath': stitched_filename,
            'row_count': len(filtered_tables)
        }
        
        # Also store in global shared session for cross-app access
        if 'shared_tables' not in session:
            session['shared_tables'] = {}
        session['shared_tables']['stitched_table'] = {
            'file_id': file_id,
            'html': stitched_html,
            'filepath': stitched_filename,
            'row_count': len(filtered_tables),
            'timestamp': datetime.now().isoformat()
        }
        
        session.modified = True
        
        logger.info(f'Stitched {len(filtered_tables)} rows from {len(layout_parsing_results)} pages')
        
        return jsonify({
            'success': True,
            'stitched_html': stitched_html,
            'row_count': len(filtered_tables),
            'page_count': len(layout_parsing_results),
            'message': f'Successfully stitched {len(filtered_tables)} rows from {len(layout_parsing_results)} pages'
        })
        
    except Exception as e:
        logger.exception('Error stitching tables')
        return jsonify({'error': str(e)}), 500

def is_header_row(row_html):
    """Check if a row is likely a header row"""
    row_text = re.sub(r'<[^>]+>', '', row_html).strip().lower()
    header_keywords = ['si.no', 'item', 'description', 'qty', 'unit', 'rate', 'amount', 'price', 'total', 'image', 'ref']
    return any(keyword in row_text for keyword in header_keywords)

@app.route('/apply-zero-costing/<file_id>', methods=['POST'])
def apply_zero_costing(file_id):
    """Apply zero costing factors to stitched table for direct document generation"""
    try:
        data = request.json
        table_data = data.get('table_data')  # Get table data from DOM
        
        if not table_data:
            return jsonify({'error': 'Table data is required'}), 400
        
        # Define zero factors (no markup/fees)
        zero_factors = {
            'net_margin': 0,
            'freight': 0,
            'customs': 0,
            'installation': 0,
            'exchange_rate': 1.0,
            'additional': 0
        }
        
        from utils.costing_engine import CostingEngine
        engine = CostingEngine()
        result = engine.apply_factors(file_id, zero_factors, session, table_data)
        
        return jsonify({
            'success': True,
            'result': result,
            'message': 'Zero costing applied successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/costing', methods=['GET', 'POST'])
def costing():
    """Costing card functionality"""
    if request.method == 'GET':
        return render_template('costing.html')
    
    # Apply costing factors
    data = request.json
    file_id = data.get('file_id')
    factors = data.get('factors', {})
    table_data = data.get('table_data')  # Get table data from DOM
    
    try:
        from utils.costing_engine import CostingEngine
        engine = CostingEngine()
        result = engine.apply_factors(file_id, factors, session, table_data)
        
        return jsonify({
            'success': True,
            'result': result,
            'message': 'Costing applied successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/generate-offer/<file_id>', methods=['POST'])
def generate_offer(file_id):
    """Generate offer with costing factors"""
    try:
        from utils.offer_generator import OfferGenerator
        generator = OfferGenerator()
        result = generator.generate(file_id, session)
        
        return jsonify({
            'success': True,
            'file_path': result,
            'message': 'Offer generated successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/multibudget/store-table', methods=['POST'])
def store_multibudget_table():
    """Store multi-budget table data for export (excludes Product Selection and Actions columns)"""
    try:
        data = request.get_json()
        tier = data.get('tier', 'budgetary')
        table_html = data.get('table_html', '')
        product_selections_data = data.get('product_selections', [])  # Optional: product selections from frontend
        
        if not table_html:
            return jsonify({'error': 'Table HTML is required'}), 400
        
        # Parse HTML and extract product selections before removing columns
        from bs4 import BeautifulSoup
        soup = BeautifulSoup(table_html, 'html.parser')
        table = soup.find('table')
        
        if not table:
            return jsonify({'error': 'Invalid table HTML'}), 400
        
        # Extract product selections and image URLs before removing Product Selection column
        product_selections = []
        
        # Use provided product selections if available (from frontend)
        if product_selections_data:
            for idx, selection in enumerate(product_selections_data):
                if selection.get('brand') and selection.get('category') and selection.get('model'):
                    product_info = {
                        'row_index': idx,
                        'brand': selection.get('brand'),
                        'category': selection.get('category'),
                        'subcategory': selection.get('subcategory', 'general'),
                        'model': selection.get('model')
                    }
                    # Get image URL from brand data
                    from utils.image_helper import get_product_image_url
                    image_url = get_product_image_url(
                        product_info['brand'], 
                        product_info['category'], 
                        product_info['subcategory'], 
                        product_info['model'], 
                        tier
                    )
                    if image_url:
                        product_info['image_url'] = image_url
                    product_selections.append(product_info)
        else:
            # Fallback: try to extract from HTML
            rows = table.find_all('tr')[1:]  # Skip header row
            for row_idx, row in enumerate(rows):
                cells = row.find_all('td')
                product_info = {}
                
                # Find Product Selection cell
                for cell in cells:
                    product_dropdowns = cell.find(class_='product-selection-dropdowns')
                    if product_dropdowns:
                        # Extract selected values from dropdowns
                        brand_select = product_dropdowns.find('select', class_='brand-dropdown')
                        category_select = product_dropdowns.find('select', class_='category-dropdown')
                        subcategory_select = product_dropdowns.find('select', class_='subcategory-dropdown')
                        model_select = product_dropdowns.find('select', class_='model-dropdown')
                        
                        if brand_select and category_select and model_select:
                            # Get selected values - check value attribute first, then selected option
                            brand = brand_select.get('value', '').strip()
                            if not brand:
                                selected_option = brand_select.find('option', selected=True) or \
                                                brand_select.find('option', {'selected': 'selected'}) or \
                                                brand_select.find('option', value=lambda v: v and v.strip())
                                if selected_option:
                                    brand = selected_option.get('value', '').strip() or selected_option.get_text(strip=True)
                            
                            category = None
                            if category_select:
                                category = category_select.get('value', '').strip()
                                if not category:
                                    selected_option = category_select.find('option', selected=True) or \
                                                    category_select.find('option', {'selected': 'selected'}) or \
                                                    category_select.find('option', value=lambda v: v and v.strip())
                                    if selected_option:
                                        category = selected_option.get('value', '').strip() or selected_option.get_text(strip=True)
                            
                            subcategory = 'general'
                            if subcategory_select:
                                subcategory = subcategory_select.get('value', '').strip() or 'general'
                                if not subcategory or subcategory == 'Select Sub-Category':
                                    selected_option = subcategory_select.find('option', selected=True) or \
                                                    subcategory_select.find('option', {'selected': 'selected'}) or \
                                                    subcategory_select.find('option', value=lambda v: v and v.strip())
                                    if selected_option:
                                        subcategory = selected_option.get('value', '').strip() or selected_option.get_text(strip=True) or 'general'
                            
                            model = None
                            if model_select:
                                model = model_select.get('value', '').strip()
                                if not model or model == 'Select Model':
                                    selected_option = model_select.find('option', selected=True) or \
                                                    model_select.find('option', {'selected': 'selected'}) or \
                                                    model_select.find('option', value=lambda v: v and v.strip())
                                    if selected_option:
                                        model_text = selected_option.get('value', '').strip() or selected_option.get_text(strip=True)
                                        # Extract model name (remove price info like "(Contact for price)")
                                        model = re.sub(r'\s*\([^)]+\)\s*$', '', model_text).strip()
                            
                            if brand and category and model:
                                product_info = {
                                    'row_index': row_idx,
                                    'brand': brand,
                                    'category': category,
                                    'subcategory': subcategory,
                                    'model': model
                                }
                                
                                # Get image URL from brand data
                                from utils.image_helper import get_product_image_url
                                image_url = get_product_image_url(brand, category, subcategory, model, tier)
                                if image_url:
                                    product_info['image_url'] = image_url
                                
                                product_selections.append(product_info)
                        break
        
        # Remove Product Selection and Actions columns
        for row in table.find_all('tr'):
            cells = row.find_all(['th', 'td'])
            cells_to_remove = []
            
            for cell in cells:
                text = cell.get_text(strip=True).lower()
                # Check if cell contains Product Selection or Actions
                if 'product selection' in text or 'actions' in text:
                    cells_to_remove.append(cell)
                # Check if cell has dropdowns or buttons
                elif cell.find(class_='product-selection-dropdowns') or cell.find('button'):
                    cells_to_remove.append(cell)
            
            for cell in cells_to_remove:
                cell.decompose()
        
        # Add image column if we have product selections with images
        if product_selections:
            # Find Image column or add it
            header_row = table.find('tr')
            if header_row:
                headers = [th.get_text(strip=True).lower() for th in header_row.find_all(['th', 'td'])]
                has_image_col = any('image' in h for h in headers)
                
                if not has_image_col:
                    # Add Image header
                    image_header = soup.new_tag('th')
                    image_header.string = 'Image'
                    header_row.insert(1, image_header)  # Insert after Sl.No
                
                # Add image cells to data rows
                rows = table.find_all('tr')[1:]
                for row_idx, row in enumerate(rows):
                    product_info = next((p for p in product_selections if p['row_index'] == row_idx), None)
                    if product_info and product_info.get('image_url'):
                        # Check if image cell already exists
                        cells = row.find_all('td')
                        has_image_cell = False
                        for cell in cells:
                            if cell.find('img'):
                                has_image_cell = True
                                break
                        
                        if not has_image_cell:
                            # Create image cell
                            image_cell = soup.new_tag('td')
                            img_tag = soup.new_tag('img', src=product_info['image_url'], style='width: 80px; height: 80px; object-fit: cover;')
                            image_cell.append(img_tag)
                            row.insert(1, image_cell)  # Insert after Sl.No
        
        # Store filtered table in session
        if 'multibudget_tables' not in session:
            session['multibudget_tables'] = {}
        
        session['multibudget_tables'][tier] = {
            'html': str(table),
            'timestamp': datetime.now().isoformat(),
            'product_selections': product_selections  # Store for export generators
        }
        session.modified = True
        
        # Create a temporary file_id for this table
        import uuid
        file_id = str(uuid.uuid4())
        
        # Store in uploaded_files for compatibility with existing export functions
        if 'uploaded_files' not in session:
            session['uploaded_files'] = []
        
        session['uploaded_files'].append({
            'id': file_id,
            'original_name': f'multibudget_{tier}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html',
            'stitched_table': {
                'html': str(table)
            },
            'multibudget': True,
            'tier': tier
        })
        session.modified = True
        
        return jsonify({
            'success': True,
            'file_id': file_id,
            'message': 'Table stored successfully for export'
        })
    except Exception as e:
        logger.exception('Error storing multibudget table')
        return jsonify({'error': str(e)}), 500

@app.route('/api/multibudget/export/<tier>', methods=['POST'])
def export_multibudget(tier):
    """Export multi-budget table (excludes Product Selection and Actions columns)"""
    try:
        data = request.get_json()
        export_type = data.get('type', 'offer')  # offer, presentation, mas
        format_type = data.get('format', 'pdf')  # pdf, excel, pptx
        
        # Get stored table from session
        if 'multibudget_tables' not in session or tier not in session['multibudget_tables']:
            return jsonify({'error': 'No table data found for this tier. Please generate a table first.'}), 404
        
        table_html = session['multibudget_tables'][tier]['html']
        
        # Create temporary file entry for export
        import uuid
        file_id = str(uuid.uuid4())
        
        # Store in session for export functions
        if 'uploaded_files' not in session:
            session['uploaded_files'] = []
        
        session['uploaded_files'].append({
            'id': file_id,
            'original_name': f'multibudget_{tier}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.html',
            'stitched_table': {
                'html': table_html
            },
            'multibudget': True,
            'tier': tier
        })
        session.modified = True
        
        # Route to appropriate export function
        if export_type == 'offer':
            if format_type == 'pdf':
                from utils.offer_generator import OfferGenerator
                generator = OfferGenerator()
                result = generator.generate(file_id, session)
                return send_file(result, as_attachment=True, download_name=f'offer_{tier}.pdf')
            elif format_type in ['excel', 'xlsx']:
                from utils.download_manager import DownloadManager
                manager = DownloadManager()
                file_path = manager.prepare_download(file_id, 'offer', format_type, session)
                return send_file(file_path, as_attachment=True, download_name=f'offer_{tier}.xlsx')
        elif export_type == 'presentation':
            if format_type == 'pdf':
                from utils.presentation_generator import PresentationGenerator
                generator = PresentationGenerator()
                result = generator.generate(file_id, session, 'pdf')
                return send_file(result, as_attachment=True, download_name=f'presentation_{tier}.pdf')
            elif format_type == 'pptx':
                from utils.presentation_generator import PresentationGenerator
                generator = PresentationGenerator()
                result = generator.generate(file_id, session, 'pptx')
                return send_file(result, as_attachment=True, download_name=f'presentation_{tier}.pptx')
        elif export_type == 'mas':
            from utils.mas_generator import MASGenerator
            generator = MASGenerator()
            result = generator.generate(file_id, session)
            return send_file(result, as_attachment=True, download_name=f'mas_{tier}.pdf')
        
        return jsonify({'error': 'Invalid export type or format'}), 400
        
    except Exception as e:
        logger.exception('Error exporting multibudget table')
        return jsonify({'error': str(e)}), 500

@app.route('/generate-presentation/<file_id>', methods=['POST'])
def generate_presentation(file_id):
    """Generate technical presentation"""
    try:
        data = request.json or {}
        format_type = data.get('format', 'pdf')
        
        from utils.presentation_generator import PresentationGenerator
        generator = PresentationGenerator()
        result = generator.generate(file_id, session, format_type)
        
        return jsonify({
            'success': True,
            'file_path': result,
            'message': f'Presentation generated successfully as {format_type.upper()}'
        })
    except Exception as e:
        logger.exception('Error generating presentation')
        import traceback
        error_details = traceback.format_exc()
        logger.error(f"Full traceback: {error_details}")
        return jsonify({'error': str(e), 'details': error_details}), 500

@app.route('/generate-mas/<file_id>', methods=['POST'])
def generate_mas(file_id):
    """Generate Material Approval Sheets"""
    try:
        # Debug logging
        uploaded_files = session.get('uploaded_files', [])
        logger.info(f'Generate MAS called for file_id: {file_id}')
        logger.info(f'Number of uploaded files: {len(uploaded_files)}')
        
        file_info = None
        for f in uploaded_files:
            if f['id'] == file_id:
                file_info = f
                logger.info(f'Found file_info with keys: {file_info.keys()}')
                logger.info(f'Has costed_data: {"costed_data" in file_info}')
                logger.info(f'Has stitched_table: {"stitched_table" in file_info}')
                logger.info(f'Has extraction_result: {"extraction_result" in file_info}')
                break
        
        if not file_info:
            logger.error(f'File not found. Available file IDs: {[f.get("id") for f in uploaded_files]}')
        
        from utils.mas_generator import MASGenerator
        generator = MASGenerator()
        result = generator.generate(file_id, session)
        
        return jsonify({
            'success': True,
            'file_path': result,
            'message': 'MAS generated successfully'
        })
    except Exception as e:
        logger.exception('Error generating MAS')
        return jsonify({'error': str(e)}), 500

@app.route('/value-engineering/<file_id>', methods=['POST'])
def value_engineering(file_id):
    """Generate value-engineered alternatives"""
    data = request.json
    budget_option = data.get('budget_option', 'mid_range')
    
    try:
        from utils.value_engineering import ValueEngineer
        engineer = ValueEngineer()
        result = engineer.generate_alternatives(file_id, budget_option, session)
        
        return jsonify({
            'success': True,
            'alternatives': result,
            'message': 'Alternatives generated successfully'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/tiers', methods=['GET'])
def get_tiers():
    """Get available budget tiers"""
    try:
        from utils.value_engineering import ValueEngineer
        engineer = ValueEngineer()
        tiers = engineer.get_tiers()
        
        return jsonify({
            'success': True,
            'tiers': tiers
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/list', methods=['GET'])
def get_brands_list():
    """Get brands for a specific tier (loads from brands_data folder)"""
    tier = request.args.get('tier', 'mid_range')
    category = request.args.get('category', None)  # Optional category filter
    
    try:
        # Normalize tier name
        tier_map = {
            'budgetary': 'budgetary',
            'mid-range': 'mid_range',
            'mid_range': 'mid_range',
            'high-end': 'high_end',
            'high_end': 'high_end'
        }
        tier = tier_map.get(tier.lower(), tier.lower())
        
        brands = []
        brands_data_dir = 'brands_data'
        
        if not os.path.exists(brands_data_dir):
            return jsonify({
                'success': True,
                'brands': [],
                'tier': tier,
                'category': category
            })
        
        
        # Load all brand files for this tier
        import json
        import re
        pattern = re.compile(rf'^.+_{re.escape(tier)}\.json$', re.I)
        
        for filename in os.listdir(brands_data_dir):
            if pattern.match(filename):
                try:
                    filepath = os.path.join(brands_data_dir, filename)
                    with open(filepath, 'r', encoding='utf-8') as f:
                        brand_data = json.load(f)
                    
                    brand_name = brand_data.get('brand', 'Unknown')
                    brand_categories = brand_data.get('categories', {})
                    
                    # Filter by category if specified
                    if category:
                        category_lower = category.lower()
                        has_category = any(
                            cat.lower() == category_lower or 
                            any(subcat.lower() == category_lower for subcat in cats.keys())
                            for cat, cats in brand_categories.items()
                        )
                        if not has_category:
                            continue
                    
                    brands.append({
                        'name': brand_name,
                        'website': brand_data.get('website', ''),
                        'country': brand_data.get('country', 'Unknown'),
                        'tier': tier,
                        'categories': list(brand_categories.keys())
                    })
                except Exception as e:
                    logger.warning(f"Error loading brand file {filename}: {e}")
                    continue
        
        return jsonify({
            'success': True,
            'brands': brands,
            'tier': tier,
            'category': category
        })
    except Exception as e:
        logger.exception('Error getting brands list')
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/models', methods=['GET'])
def get_brand_models_api():
    """Get models for a specific brand (loads from brands_data folder)"""
    tier = request.args.get('tier', 'mid_range')
    category = request.args.get('category', None)
    brand = request.args.get('brand', '')
    subcategory = request.args.get('subcategory', None)
    
    if not brand:
        return jsonify({
            'success': True,
            'models': []
        })
    
    try:
        # Normalize tier name
        tier_map = {
            'budgetary': 'budgetary',
            'mid-range': 'mid_range',
            'mid_range': 'mid_range',
            'high-end': 'high_end',
            'high_end': 'high_end'
        }
        tier = tier_map.get(tier.lower(), tier.lower())
        
        # Load brand data from brands_data folder
        import json
        import re
        safe_brand_name = re.sub(r'[^\w\-_]', '', brand.replace(' ', '_'))
        filename = f"{safe_brand_name}_{tier}.json"
        filepath = os.path.join('brands_data', filename)
        
        if not os.path.exists(filepath):
            # Try case-insensitive search
            brands_data_dir = 'brands_data'
            if os.path.exists(brands_data_dir):
                for f in os.listdir(brands_data_dir):
                    if f.lower() == filename.lower():
                        filepath = os.path.join(brands_data_dir, f)
                        break
        
        if not os.path.exists(filepath):
            return jsonify({
                'success': True,
                'models': []
            })
        
        with open(filepath, 'r', encoding='utf-8') as f:
            brand_data = json.load(f)
        
        models = []
        categories_data = brand_data.get('categories', {})
        
        # Try category_tree format (from old scrapers)
        if not categories_data and 'category_tree' in brand_data:
            categories_data = brand_data.get('category_tree', {})
        
        # Handle collections format - convert to categories format (from Firecrawl)
        if not categories_data and 'collections' in brand_data:
            categories_data = {}
            for collection_name, collection_data in brand_data.get('collections', {}).items():
                # Handle "Category > Subcategory" format
                parts = collection_name.split('>')
                if len(parts) == 2:
                    cat = parts[0].strip()
                    subcat = parts[1].strip()
                    if cat not in categories_data:
                        categories_data[cat] = {'subcategories': {}}
                    if 'subcategories' not in categories_data[cat]:
                        categories_data[cat]['subcategories'] = {}
                    
                    # Convert products to list
                    products_list = []
                    for product in collection_data.get('products', []):
                        model_entry = {
                            'model': product.get('name', 'Unknown'),
                            'image_url': product.get('image_url', ''),
                            'source_url': product.get('source_url', ''),
                            'description': product.get('description', ''),
                            'price': product.get('price'),
                            'price_range': product.get('price_range', 'Contact for price'),
                            'features': product.get('features', [])
                        }
                        products_list.append(model_entry)
                    
                    categories_data[cat]['subcategories'][subcat] = {'products': products_list}
                else:
                    # Single-level category - use 'general' subcategory
                    clean_name = collection_name.split('\n')[0].strip()
                    if clean_name not in categories_data:
                        categories_data[clean_name] = {'subcategories': {}}
                    
                    # Add 'general' subcategory with products
                    products_list = []
                    for product in collection_data.get('products', []):
                        model_entry = {
                            'model': product.get('name', 'Unknown'),
                            'image_url': product.get('image_url', ''),
                            'source_url': product.get('source_url', ''),
                            'description': product.get('description', ''),
                            'price': product.get('price'),
                            'price_range': product.get('price_range', 'Contact for price'),
                            'features': product.get('features', [])
                        }
                        products_list.append(model_entry)
                    
                    categories_data[clean_name]['subcategories']['general'] = {'products': products_list}
        
        # Filter by category if specified
        if category:
            category_lower = category.lower()
            matching_categories = {
                cat: cats for cat, cats in categories_data.items()
                if cat.lower() == category_lower or category_lower in cat.lower()
            }
        else:
            matching_categories = categories_data
        
        # Get models from matching categories and subcategories
        for cat_name, cat_data in matching_categories.items():
            # Extract subcategories - handle both old and new formats
            if isinstance(cat_data, dict) and 'subcategories' in cat_data:
                subcategories = cat_data['subcategories']
            else:
                subcategories = cat_data if isinstance(cat_data, dict) else {}
            
            if subcategory:
                subcategory_lower = subcategory.lower()
                matching_subcats = {
                    subcat: subcat_data 
                    for subcat, subcat_data in subcategories.items()
                    if subcat.lower() == subcategory_lower or subcategory_lower in subcat.lower()
                }
            else:
                matching_subcats = subcategories
            
            for subcat_name, subcat_data in matching_subcats.items():
                # Extract products list - handle nested structure
                if isinstance(subcat_data, dict) and 'products' in subcat_data:
                    models_list = subcat_data['products']
                elif isinstance(subcat_data, list):
                    models_list = subcat_data
                else:
                    continue
                
                for model in models_list:
                    # Handle both model dict format and product format
                    model_name = model.get('model') or model.get('name', 'Unknown')
                    models.append({
                        'model': model_name,
                        'price': model.get('price'),
                        'price_range': model.get('price_range', 'Contact for price'),
                        'description': model.get('description') or model.get('name', ''),
                        'image_url': model.get('image_url', ''),
                        'features': model.get('features', []),
                        'source_url': model.get('source_url') or model.get('url', ''),
                        'category': cat_name,
                        'subcategory': subcat_name
                    })
        
        # Optional: Enrich products with missing images/descriptions
        enrich = request.args.get('enrich', 'false').lower() == 'true'
        if enrich and models:
            try:
                from utils.product_enricher import ProductEnricher
                enricher = ProductEnricher()
                logger.info(f"Enriching {len(models)} products for {brand}...")
                models = enricher.enrich_product_selection_data(models, use_selenium=False)
                logger.info(f"Enrichment complete for {brand}")
            except Exception as e:
                logger.error(f"Error enriching products: {e}")
                # Continue without enrichment
        
        return jsonify({
            'success': True,
            'models': models,
            'brand': brand,
            'tier': tier,
            'category': category,
            'subcategory': subcategory,
            'enriched': enrich
        })
    except Exception as e:
        logger.exception('Error getting models')
        return jsonify({'error': str(e)}), 500

@app.route('/api/products/enrich', methods=['POST'])
def enrich_products():
    """Enrich product data with images and descriptions"""
    try:
        data = request.get_json()
        products = data.get('products', [])
        use_selenium = data.get('use_selenium', False)
        
        if not products:
            return jsonify({'error': 'No products provided'}), 400
        
        from utils.product_enricher import ProductEnricher
        enricher = ProductEnricher()
        
        logger.info(f"Enriching {len(products)} products...")
        enriched_products = enricher.enrich_product_selection_data(products, use_selenium)
        logger.info(f"Enrichment complete")
        
        return jsonify({
            'success': True,
            'products': enriched_products,
            'count': len(enriched_products)
        })
    except Exception as e:
        logger.exception('Error enriching products')
        return jsonify({'error': str(e)}), 500


@app.route('/api/brands/categories', methods=['GET'])
def get_brands_categories():
    """Get categories for a specific brand and tier (loads from brands_data folder)"""
    tier = request.args.get('tier', 'mid_range')
    brand = request.args.get('brand', None)
    
    if not brand:
        # Return all categories across all brands in this tier
        brands_data_dir = 'brands_data'
        all_categories = set()
        
        if os.path.exists(brands_data_dir):
            import json
            import re
            tier_map = {
                'budgetary': 'budgetary',
                'mid-range': 'mid_range',
                'mid_range': 'mid_range',
                'high-end': 'high_end',
                'high_end': 'high_end'
            }
            tier = tier_map.get(tier.lower(), tier.lower())
            pattern = re.compile(rf'^.+_{re.escape(tier)}\.json$', re.I)
            
            for filename in os.listdir(brands_data_dir):
                if pattern.match(filename):
                    try:
                        filepath = os.path.join(brands_data_dir, filename)
                        with open(filepath, 'r', encoding='utf-8') as f:
                            brand_data = json.load(f)
                        # Handle both collections and categories formats
                        cats = brand_data.get('categories', {})
                        if not cats and 'collections' in brand_data:
                            # Extract collection names as categories
                            for collection_name in brand_data.get('collections', {}).keys():
                                clean_name = collection_name.split('\n')[0].strip()
                                all_categories.add(clean_name)
                        else:
                            all_categories.update(cats.keys())
                    except:
                        continue
        
        return jsonify({
            'success': True,
            'categories': sorted(list(all_categories))
        })
    
    # Load specific brand's categories
    try:
        import json
        import re
        tier_map = {
            'budgetary': 'budgetary',
            'mid-range': 'mid_range',
            'mid_range': 'mid_range',
            'high-end': 'high_end',
            'high_end': 'high_end'
        }
        tier = tier_map.get(tier.lower(), tier.lower())
        
        safe_brand_name = re.sub(r'[^\w\-_]', '', brand.replace(' ', '_'))
        filename = f"{safe_brand_name}_{tier}.json"
        filepath = os.path.join('brands_data', filename)
        
        if not os.path.exists(filepath):
            # Try case-insensitive
            brands_data_dir = 'brands_data'
            if os.path.exists(brands_data_dir):
                for f in os.listdir(brands_data_dir):
                    if f.lower() == filename.lower():
                        filepath = os.path.join(brands_data_dir, f)
                        break
        
        if not os.path.exists(filepath):
            return jsonify({
                'success': True,
                'categories': []
            })
        
        with open(filepath, 'r', encoding='utf-8') as f:
            brand_data = json.load(f)
        
        # Handle multiple data formats: categories, collections, category_tree
        categories_data = brand_data.get('categories', {})
        
        # Try category_tree format (from old scrapers)
        if not categories_data and 'category_tree' in brand_data:
            categories_data = brand_data.get('category_tree', {})
        
        # Try collections format (from Firecrawl)
        if not categories_data and 'collections' in brand_data:
            # Convert collections to categories format
            categories_data = {}
            for collection_name, collection_data in brand_data.get('collections', {}).items():
                # Handle "Category > Subcategory" format
                parts = collection_name.split('>')
                if len(parts) == 2:
                    cat = parts[0].strip()
                    subcat = parts[1].strip()
                    if cat not in categories_data:
                        categories_data[cat] = {}
                    if 'subcategories' not in categories_data[cat]:
                        categories_data[cat]['subcategories'] = {}
                    categories_data[cat]['subcategories'][subcat] = collection_data
                else:
                    # Single-level category
                    clean_name = collection_name.split('\n')[0].strip()
                    if clean_name not in categories_data:
                        categories_data[clean_name] = {}
        
        categories = list(categories_data.keys()) if categories_data else []
        
        return jsonify({
            'success': True,
            'categories': categories
        })
    except Exception as e:
        logger.exception('Error getting categories')
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/subcategories', methods=['GET'])
def get_subcategories_api():
    """Get subcategories for a specific brand and category (loads from brands_data folder)"""
    tier = request.args.get('tier', 'mid_range')
    brand = request.args.get('brand', None)
    category = request.args.get('category', None)
    
    if not brand or not category:
        return jsonify({
            'success': True,
            'subcategories': []
        })
    
    try:
        # Load brand data
        import json
        import re
        tier_map = {
            'budgetary': 'budgetary',
            'mid-range': 'mid_range',
            'mid_range': 'mid_range',
            'high-end': 'high_end',
            'high_end': 'high_end'
        }
        tier = tier_map.get(tier.lower(), tier.lower())
        
        safe_brand_name = re.sub(r'[^\w\-_]', '', brand.replace(' ', '_'))
        filename = f"{safe_brand_name}_{tier}.json"
        filepath = os.path.join('brands_data', filename)
        
        if not os.path.exists(filepath):
            # Try case-insensitive
            brands_data_dir = 'brands_data'
            if os.path.exists(brands_data_dir):
                for f in os.listdir(brands_data_dir):
                    if f.lower() == filename.lower():
                        filepath = os.path.join(brands_data_dir, f)
                        break
        
        if not os.path.exists(filepath):
            return jsonify({
                'success': True,
                'subcategories': []
            })
        
        with open(filepath, 'r', encoding='utf-8') as f:
            brand_data = json.load(f)
        
        # Handle multiple data formats: categories, collections, category_tree
        categories_data = brand_data.get('categories', {})
        
        # Try category_tree format (from old scrapers)
        if not categories_data and 'category_tree' in brand_data:
            categories_data = brand_data.get('category_tree', {})
        
        # Try collections format (from Firecrawl)
        if not categories_data and 'collections' in brand_data:
            # Convert collections to categories format
            categories_data = {}
            for collection_name, collection_data in brand_data.get('collections', {}).items():
                # Handle "Category > Subcategory" format
                parts = collection_name.split('>')
                if len(parts) == 2:
                    cat = parts[0].strip()
                    subcat = parts[1].strip()
                    if cat not in categories_data:
                        categories_data[cat] = {'subcategories': {}}
                    if 'subcategories' not in categories_data[cat]:
                        categories_data[cat]['subcategories'] = {}
                    categories_data[cat]['subcategories'][subcat] = collection_data
                else:
                    # Single-level category
                    clean_name = collection_name.split('\n')[0].strip()
                    if clean_name not in categories_data:
                        categories_data[clean_name] = {}
        
        # Find matching category (case-insensitive)
        matching_category = None
        category_lower = category.lower()
        for cat_name, cat_data in categories_data.items():
            if cat_name.lower() == category_lower or category_lower in cat_name.lower():
                # Check if it has subcategories key (category_tree format)
                if isinstance(cat_data, dict) and 'subcategories' in cat_data:
                    matching_category = cat_data['subcategories']
                else:
                    matching_category = cat_data
                break
        
        if not matching_category:
            return jsonify({
                'success': True,
                'subcategories': []
            })
        
        subcategories = list(matching_category.keys()) if matching_category else []
        
        return jsonify({
            'success': True,
            'subcategories': subcategories,
            'category': category
        })
    except Exception as e:
        logger.exception('Error getting subcategories')
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/scrape', methods=['POST'])
def scrape_brand():
    """Scrape a brand's website to discover products"""
    try:
        data = request.get_json()
        brand_name = data.get('brand_name')
        website = data.get('website')
        
        if not brand_name or not website:
            return jsonify({'error': 'Brand name and website are required'}), 400
        
        from utils.universal_brand_scraper import UniversalBrandScraper
        scraper = UniversalBrandScraper()
        
        logger.info(f"Starting scrape for {brand_name} ({website}) using UniversalBrandScraper")
        scraped_data = scraper.scrape_brand_website(website, brand_name)
        
        if 'error' in scraped_data:
            return jsonify({'error': scraped_data['error']}), 400
        
        return jsonify({
            'success': True,
            'data': scraped_data,
            'message': f'Successfully scraped {len(scraped_data.get("products", []))} products'
        })
    except Exception as e:
        logger.exception('Error in brand scraping')
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/add', methods=['POST'])
def add_brand():
    """Add a new brand to the database"""
    try:
        data = request.get_json()
        brand_name = data.get('brand_name')
        website = data.get('website')
        country = data.get('country', 'Unknown')
        tier = data.get('tier', 'mid_range')
        categories = data.get('categories', {})
        
        if not brand_name or not website:
            return jsonify({'error': 'Brand name and website are required'}), 400
        
        # Load existing brands
        from utils.brand_database import BRAND_DATABASE
        import json
        
        # Check if brand already exists
        for t in BRAND_DATABASE:
            for cat in BRAND_DATABASE[t]:
                for brand in BRAND_DATABASE[t][cat]:
                    if brand['name'].lower() == brand_name.lower():
                        return jsonify({'error': 'Brand already exists'}), 400
        
        # Create new brand entry
        new_brand = {
            'name': brand_name,
            'website': website,
            'country': country,
            'models': categories
        }
        
        # Add to database (in memory for now)
        if tier not in BRAND_DATABASE:
            return jsonify({'error': f'Invalid tier: {tier}'}), 400
        
        # Add to first available category or create new one
        if categories:
            for category in categories:
                if category not in BRAND_DATABASE[tier]:
                    BRAND_DATABASE[tier][category] = []
                
                # Check if brand already in this category
                brand_exists = False
                for brand in BRAND_DATABASE[tier][category]:
                    if brand['name'].lower() == brand_name.lower():
                        brand_exists = True
                        break
                
                if not brand_exists:
                    BRAND_DATABASE[tier][category].append(new_brand)
        else:
            # Default to general category
            if 'general' not in BRAND_DATABASE[tier]:
                BRAND_DATABASE[tier]['general'] = []
            BRAND_DATABASE[tier]['general'].append(new_brand)
        
        # Save to file
        save_brand_database(BRAND_DATABASE)
        
        return jsonify({
            'success': True,
            'message': f'Brand {brand_name} added successfully',
            'brand': new_brand
        })
    except Exception as e:
        logger.exception('Error adding brand')
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/scrape-and-add', methods=['POST'])
def scrape_and_add_brand():
    """
    Unified scraper: Always uses Requests with automatic Selenium fallback
    Supports: Requests (default) → Selenium fallback, Firecrawl (optional)
    """
    try:
        data = request.get_json()
        brand_name = data.get('brand_name')
        website = data.get('website')
        country = data.get('country', 'Unknown')
        tier = data.get('tier', 'mid_range')
        scraping_method = data.get('scraping_method', 'requests')  # Default to 'requests'
        crawl_limit = data.get('crawl_limit', 50)
        
        logger.info(f"Received scrape request data: {data}")
        
        if not brand_name or not website:
            return jsonify({'error': 'Brand name and website are required'}), 400
        
        logger.info(f"🔍 SCRAPING METHOD: {scraping_method.upper()}")
        logger.info(f"Starting unified scrape for {brand_name} ({website})")
        
        # Check Selenium availability
        try:
            from utils.selenium_scraper import SELENIUM_AVAILABLE
        except ImportError:
            SELENIUM_AVAILABLE = False
        
        # PRIORITY: Check if this is an Architonic URL first (use specialized scraper)
        from utils.architonic_scraper import ArchitonicScraper
        use_selenium = scraping_method != 'requests'
        architonic_scraper = ArchitonicScraper(use_selenium=use_selenium and SELENIUM_AVAILABLE)
        
        is_architonic = architonic_scraper.is_architonic_url(website)
        logger.info(f"🔍 Architonic URL check: {website} → is_architonic={is_architonic}")
        
        if is_architonic:
            method_used = "Selenium" if architonic_scraper.use_selenium else "Requests"
            logger.info(f"🏛️ Detected Architonic URL - Using specialized ArchitonicScraper")
            logger.info(f"   📋 Scraping Method: {scraping_method} → Using {method_used} for Architonic")
            scraped_data = architonic_scraper.scrape_collection(
                url=website,
                brand_name=brand_name
            )
            # Convert collections to category_tree format
            if 'collections' in scraped_data and not scraped_data.get('category_tree'):
                scraped_data['category_tree'] = architonic_scraper._convert_collections_to_category_tree(scraped_data['collections'])
        # UNIFIED SCRAPER: Requests with automatic Selenium fallback
        elif scraping_method == 'requests' or scraping_method == 'universal':
            logger.info(f"🚀 Using Unified Scraper (Requests → Selenium fallback)")
            from utils.requests_brand_scraper import RequestsBrandScraper
            scraper = RequestsBrandScraper(delay=0.5, fetch_descriptions=True)
            scraped_data = scraper.scrape_brand_website(
                website=website,
                brand_name=brand_name,
                limit=crawl_limit
            )
            
            # Check if we should trigger Selenium fallback
            # Trigger if: no products, error, or categories found but no products in them (JS-rendered)
            total_products = scraped_data.get('total_products', 0)
            category_tree = scraped_data.get('category_tree', {})
            collections = scraped_data.get('collections', {})
            requires_javascript = scraped_data.get('requires_javascript', False)
            categories_found = scraped_data.get('categories_found', 0)
            
            logger.info(f"📊 Requests scraper results:")
            logger.info(f"   - Success: {scraped_data.get('success')}")
            logger.info(f"   - Total products: {total_products}")
            logger.info(f"   - Category tree size: {len(category_tree)}")
            logger.info(f"   - Collections size: {len(collections)}")
            logger.info(f"   - Requires JavaScript: {requires_javascript}")
            
            # Check if categories exist but have no products (indicates JS-rendered content)
            categories_with_products = 0
            if category_tree:
                for cat_data in category_tree.values():
                    for subcat_data in cat_data.get('subcategories', {}).values():
                        products = subcat_data.get('products', [])
                        if products and len(products) > 0:
                            categories_with_products += 1
            elif collections:
                for coll_data in collections.values():
                    products = coll_data.get('products', [])
                    if products and len(products) > 0:
                        categories_with_products += 1
            
            # ALWAYS trigger fallback if:
            # 1. total_products == 0 (no products found)
            # 2. requires_javascript flag is set (site needs JS)
            # 3. Categories found but have very few products (< 5 per category) - indicates JS-rendered content
            # 4. Categories exist but categories_with_products == 0
            # 5. Success but empty results
            
            # Calculate average products per category
            avg_products_per_category = 0
            if categories_found > 0:
                avg_products_per_category = total_products / categories_found if categories_found > 0 else 0
            
            should_fallback = (
                scraped_data.get('success') == False or 
                total_products == 0 or  # PRIMARY CONDITION - always trigger if no products
                'error' in scraped_data or
                requires_javascript or  # If JS required, always use Selenium
                # If categories found but very few products per category (< 5), likely JS-rendered
                (categories_found > 0 and avg_products_per_category < 5 and requires_javascript) or
                (categories_found > 0 and categories_with_products == 0) or
                (len(category_tree) > 0 or len(collections) > 0) and categories_with_products == 0 or
                # If requests scraper succeeded but found nothing, likely needs Selenium
                (scraped_data.get('success') == True and total_products == 0 and len(category_tree) == 0 and len(collections) == 0)
            )
            
            logger.info(f"🔍 Fallback check: should_fallback={should_fallback}")
            logger.info(f"   - Average products per category: {avg_products_per_category:.1f}")
            
            if should_fallback:
                logger.warning(f"⚠️  Requests scraper issues detected:")
                logger.warning(f"   - Success: {scraped_data.get('success')}")
                logger.warning(f"   - Total products: {total_products}")
                logger.warning(f"   - Categories found: {len(category_tree) + len(collections)}")
                logger.warning(f"   - Categories with products: {categories_with_products}")
                logger.warning(f"   - Requires JavaScript: {requires_javascript}")
                logger.warning(f"   - Category tree empty: {len(category_tree) == 0}")
                logger.info(f"🔄 Attempting automatic fallback to Selenium scraper...")
                logger.info(f"   ⏳ This may take 1-2 minutes for JavaScript-heavy sites like LAS.it...")
                if SELENIUM_AVAILABLE:
                    try:
                        from utils.universal_brand_scraper import UniversalBrandScraper
                        selenium_scraper = UniversalBrandScraper()
                        logger.info(f"🌐 Loading page with Selenium (this may take a minute)...")
                        selenium_data = selenium_scraper.scrape_brand_website(website, brand_name, use_selenium=True)
                        
                        # If Selenium found products or better structure, use that
                        selenium_products = selenium_data.get('total_products', 0)
                        selenium_collections = selenium_data.get('collections', {})
                        selenium_category_tree = selenium_data.get('category_tree', {})
                        
                        if selenium_products > 0:
                            logger.info(f"✅ Selenium fallback successful! Found {selenium_products} products across {len(selenium_collections) + len(selenium_category_tree)} collections")
                            scraped_data = selenium_data
                        elif len(selenium_collections) > 0 or len(selenium_category_tree) > 0:
                            logger.info(f"✅ Selenium found {len(selenium_collections) + len(selenium_category_tree)} categories (products may require additional scraping)")
                            scraped_data = selenium_data
                        else:
                            logger.warning(f"⚠️  Selenium fallback also returned no categories/products")
                    except Exception as e:
                        logger.error(f"❌ Selenium fallback failed: {e}")
                        import traceback
                        logger.error(traceback.format_exc())
                else:
                    logger.warning("⚠️  Selenium not available for fallback. Install Selenium for JavaScript-heavy sites: pip install selenium webdriver-manager")
        elif scraping_method == 'firecrawl':
            logger.info(f"🔥 Using FirecrawlBrandScraper (AI-powered, API limits)")
            from utils.firecrawl_brand_scraper import FirecrawlBrandScraper
            scraper = FirecrawlBrandScraper()
            scraped_data = scraper.scrape_brand_website(
                website=website,
                brand_name=brand_name,
                limit=crawl_limit
            )
        else:  # universal (legacy)
            logger.info(f"⚡ Using UniversalBrandScraper (Legacy)")
            from utils.universal_brand_scraper import UniversalBrandScraper
            scraper = UniversalBrandScraper()
            scraped_data = scraper.scrape_brand_website(website, brand_name)
        
        if 'error' in scraped_data:
            return jsonify({'error': scraped_data['error']}), 400
        
        # Load existing brands from brands_dynamic.json
        brands_file = os.path.join('brands_data', 'brands_dynamic.json')
        if os.path.exists(brands_file):
            with open(brands_file, 'r', encoding='utf-8') as f:
                brands_data = json.load(f)
        else:
            brands_data = {'brands': []}
        
        # Handle both collections (old format) and category_tree (new format)
        # Convert UniversalBrandScraper collections format to category_tree if needed
        if scraped_data.get('collections') and not scraped_data.get('category_tree'):
            # Convert collections format to category_tree format
            category_tree = {}
            collections = scraped_data.get('collections', {})
            logger.info(f"Converting {len(collections)} collections to category_tree format...")
            
            for coll_name, coll_data in collections.items():
                # Use category from coll_data, or fall back to coll_name if category not set
                category = coll_data.get('category') or coll_name
                subcategory = coll_data.get('subcategory') or 'General'
                products = coll_data.get('products', [])
                
                logger.info(f"  Converting collection '{coll_name}': category='{category}', subcategory='{subcategory}', products={len(products)}")
                
                if category not in category_tree:
                    category_tree[category] = {'subcategories': {}}
                
                if subcategory not in category_tree[category]['subcategories']:
                    category_tree[category]['subcategories'][subcategory] = {'products': []}
                
                category_tree[category]['subcategories'][subcategory]['products'].extend(products)
            
            scraped_data['category_tree'] = category_tree
            logger.info(f"✅ Converted collections format to category_tree: {len(category_tree)} categories with {scraped_data.get('total_products', 0)} total products")
        
        categories_data = scraped_data.get('collections', {}) or scraped_data.get('category_tree', {})
        
        # Check if brand already exists
        brand_exists = False
        for brand in brands_data['brands']:
            if brand['name'].lower() == brand_name.lower():
                brand_exists = True
                # Update existing brand
                brand['website'] = website
                brand['country'] = country
                brand['tier'] = tier
                brand['categories'] = scraped_data.get('collections', {})
                brand['category_tree'] = scraped_data.get('category_tree', {})
                brand['last_scraped_at'] = datetime.now().isoformat()
                logger.info(f"Updated existing brand: {brand_name}")
                break
        
        if not brand_exists:
            # Create new brand entry
            new_brand = {
                'name': brand_name,
                'website': website,
                'country': country,
                'tier': tier,
                'categories': scraped_data.get('collections', {}),
                'category_tree': scraped_data.get('category_tree', {}),
                'added_date': datetime.now().isoformat(),
                'last_scraped_at': datetime.now().isoformat()
            }
            brands_data['brands'].append(new_brand)
            logger.info(f"Added new brand: {brand_name}")
        
        # Save to brands_dynamic.json
        os.makedirs('brands_data', exist_ok=True)
        with open(brands_file, 'w', encoding='utf-8') as f:
            json.dump(brands_data, f, indent=2, ensure_ascii=False)
        
        # Also save to individual brand file (e.g., OTTIMO_budgetary.json)
        save_individual_brand_file(brand_name, website, country, tier, scraped_data)
        
        # Extract categories for response
        categories = list(categories_data.keys())
        products_count = scraped_data.get('total_products', 0)
        
        return jsonify({
            'success': True,
            'message': f'Successfully scraped and added {brand_name}',
            'products_count': products_count,
            'categories': categories,
            'scraping_method': scraping_method
        })
        
    except Exception as e:
        logger.exception('Error in scrape_and_add_brand')
        return jsonify({'error': str(e)}), 500

def _scrape_single_brand(brand_info):
    """
    Helper function to scrape a single brand (used for parallel scraping)
    Args:
        brand_info: Dict with 'brand_name', 'website', 'country', 'tier'
    Returns:
        Dict with 'success', 'brand_name', 'data' or 'error'
    """
    try:
        brand_name = brand_info['brand_name']
        website = brand_info['website']
        country = brand_info.get('country', 'Unknown')
        tier = brand_info.get('tier', 'mid_range')
        
        # Normalize tier names
        tier_map = {
            'budgetary': 'budgetary',
            'mid-range': 'mid_range',
            'mid_range': 'mid_range',
            'high-end': 'high_end',
            'high_end': 'high_end'
        }
        tier = tier_map.get(tier.lower(), tier.lower())
        
        # Use unified scraper (Requests → Selenium fallback)
        from utils.requests_brand_scraper import RequestsBrandScraper
        from utils.universal_brand_scraper import UniversalBrandScraper
        
        logger.info(f"[Parallel] Scraping {brand_name} ({website}) for tier {tier}")
        
        # Simplified retry logic - stop after first successful scrape with any products
        max_retries = 2
        retry_count = 0
        scraped_data = None
        last_error = None
        
        while retry_count < max_retries:
            retry_count += 1
            logger.info(f"[Parallel] Scraping attempt {retry_count}/{max_retries} for {brand_name}")
            
            try:
                # Check if Architonic URL first
                from utils.architonic_scraper import ArchitonicScraper
                architonic_scraper = ArchitonicScraper(use_selenium=True)
                
                if architonic_scraper.is_architonic_url(website):
                    logger.info(f"[Parallel] Using Architonic scraper for {brand_name}")
                    scraped_data = architonic_scraper.scrape_collection(website, brand_name)
                else:
                    # Try Requests scraper first
                    requests_scraper = RequestsBrandScraper(delay=0.5, fetch_descriptions=True)
                    scraped_data = requests_scraper.scrape_brand_website(website, brand_name, limit=100)
                    
                    # Check if we need Selenium fallback
                    total_products = scraped_data.get('total_products', 0)
                    requires_javascript = scraped_data.get('requires_javascript', False)
                    
                    if total_products == 0 or requires_javascript:
                        logger.info(f"[Parallel] Falling back to Selenium for {brand_name}")
                        selenium_scraper = UniversalBrandScraper()
                        scraped_data = selenium_scraper.scrape_brand_website(website, brand_name, use_selenium=True)
                
                if 'error' in scraped_data:
                    last_error = scraped_data['error']
                    if retry_count >= max_retries:
                        logger.error(f"[Parallel] Scraping failed after {max_retries} attempts for {brand_name}: {last_error}")
                        return {
                            'success': False,
                            'brand_name': brand_name,
                            'error': f'Scraping failed after {max_retries} attempts. Last error: {last_error}',
                            'details': last_error
                        }
                    logger.warning(f"[Parallel] Scraping error on attempt {retry_count} for {brand_name}: {last_error}. Retrying...")
                    time.sleep(2)
                    continue
            except Exception as e:
                last_error = str(e)
                logger.exception(f"[Parallel] Exception during scraping attempt {retry_count} for {brand_name}: {e}")
                if retry_count >= max_retries:
                    return {
                        'success': False,
                        'brand_name': brand_name,
                        'error': f'Scraping failed after {max_retries} attempts due to exception: {last_error}',
                        'details': last_error
                    }
                time.sleep(2)
                continue
            
            # Count total products - handle both collections format and regular format
            total_products = 0
            is_collections_format = 'collections' in scraped_data and 'total_products' in scraped_data
            
            # Check if it's collections format (Architonic)
            if is_collections_format:
                total_products = scraped_data.get('total_products', 0)
                total_collections = scraped_data.get('total_collections', 0)
                logger.info(f"[Parallel] Found {total_products} products in {total_collections} collections for {brand_name} on attempt {retry_count}")
                # Accept if we got any products - no minimum requirement
                if total_products > 0:
                    logger.info(f"[Parallel] Successfully scraped collections with {total_products} products for {brand_name}. Accepting result.")
                    break
                elif retry_count < max_retries:
                    logger.warning(f"[Parallel] No products found in collections format for {brand_name}. Retrying...")
                    time.sleep(2)
                    continue
                else:
                    logger.error(f"[Parallel] No products found after {max_retries} attempts for {brand_name}")
                    break
            else:
                # Regular format - count from products and categories
                total_products = len(scraped_data.get('products', []))
                for category_products in scraped_data.get('categories', {}).values():
                    if isinstance(category_products, list):
                        total_products += len(category_products)
                    elif isinstance(category_products, dict):
                        for subcat_products in category_products.values():
                            if isinstance(subcat_products, list):
                                total_products += len(subcat_products)
                logger.info(f"[Parallel] Found {total_products} products for {brand_name} on attempt {retry_count}")
                
                # Accept if we got any products - no minimum requirement, stop immediately
                if total_products > 0:
                    logger.info(f"[Parallel] Successfully scraped {total_products} products for {brand_name}. Accepting result.")
                    break
                elif retry_count < max_retries:
                    logger.warning(f"[Parallel] No products found for {brand_name}. Retrying...")
                    time.sleep(2)
                    continue
                else:
                    logger.error(f"[Parallel] No products found after {max_retries} attempts for {brand_name}")
                    break
        
        if not scraped_data:
            update_scrape_status(job_id, 'failed', 'Scraping failed: No data returned', 100)
            return {
                'success': False,
                'brand_name': brand_name,
                'error': 'Scraping failed: No data returned',
                'job_id': job_id
            }
        
        if 'error' in scraped_data:
            update_scrape_status(job_id, 'failed', f'Error: {scraped_data["error"]}', 100)
            return {
                'success': False,
                'brand_name': brand_name,
                'error': scraped_data['error'],
                'job_id': job_id
            }
        
        # Check if data is in collections format (from Architonic)
        is_collections_format = 'collections' in scraped_data and 'total_collections' in scraped_data
        
        update_scrape_status(job_id, 'organizing', 'Organizing product data by categories...', 75)
        
        # Organize data by categories and subcategories
        organized_data = {
            'brand': brand_name,
            'website': website,
            'country': country,
            'tier': tier,
            'categories': {}
        }
        
        # If collections format, preserve it and also convert to categories format
        if is_collections_format:
            # Preserve original collections structure
            organized_data['collections'] = scraped_data.get('collections', {})
            organized_data['all_products'] = scraped_data.get('all_products', [])
            organized_data['source'] = scraped_data.get('source', 'Architonic Collections')
            organized_data['scraped_at'] = scraped_data.get('scraped_at')
            organized_data['total_products'] = scraped_data.get('total_products', 0)
            organized_data['total_collections'] = scraped_data.get('total_collections', 0)
            
            # Convert collections to categories format with intelligent hierarchy and deduplication
            
            # First pass: Identify all products in specific subcategories
            products_by_cat_subcat = {} # {Category: {Subcategory: {model_name: product_data}}}
            
            for collection_name, collection_data in scraped_data.get('collections', {}).items():
                if not collection_data or not collection_data.get('products'):
                    continue
                
                # Skip if collection_name is None
                if not collection_name:
                    continue
                
                clean_name = collection_name.split('\n')[0].strip()
                
                # Determine Category and Subcategory
                if ' > ' in clean_name:
                    parts = clean_name.split(' > ')
                    category = parts[0].strip()
                    subcategory = parts[1].strip()
                else:
                    category = clean_name
                    subcategory = 'General' # Default, will be filtered later
                
                if category not in products_by_cat_subcat:
                    products_by_cat_subcat[category] = {}
                if subcategory not in products_by_cat_subcat[category]:
                    products_by_cat_subcat[category][subcategory] = {}
                
                for product in collection_data.get('products', []):
                    model_name = product.get('model') or product.get('title') or product.get('name')
                    if not model_name: 
                        continue
                    
                    # Clean model name
                    model_name = model_name.strip()
                    
                    # Skip garbage
                    if model_name.lower() in ['select options', 'contact for price', 'read more', 'add to cart']:
                        continue
                    if 'contact for price' in model_name.lower():
                        model_name = model_name.replace('(Contact for price)', '').replace('Contact for price', '').strip()
                    
                    # Store product
                    products_by_cat_subcat[category][subcategory][model_name] = product

            # Second pass: Build organized_data, handling deduplication
            # If a product exists in a specific subcategory, remove it from 'General'
            
            for category, subcats in products_by_cat_subcat.items():
                organized_data['categories'][category] = {}
                
                # Get all models in specific subcategories (excluding General)
                specific_models = set()
                for subcat, models in subcats.items():
                    if subcat != 'General':
                        specific_models.update(models.keys())
                
                for subcat, models in subcats.items():
                    final_models_list = []
                    
                    for model_name, product in models.items():
                        # If we are in General, and this model exists in a specific subcategory, SKIP it
                        if subcat == 'General' and model_name in specific_models:
                            continue
                        
                        # Convert to app's model format
                        model_entry = {
                            'model': model_name,
                            'price': product.get('price'),
                            'price_range': product.get('price_range', "Contact for price"),
                            'features': product.get('features', []),
                            'image_url': product.get('image_url'),
                            'description': product.get('description', ''),
                            'source_url': product.get('url', product.get('source_url', website)),
                            'product_id': product.get('product_id', ''),
                            'collection': category + (' > ' + subcat if subcat != 'General' else '')
                        }
                        final_models_list.append(model_entry)
                    
                    # Only add subcategory if it has products
                    if final_models_list:
                        organized_data['categories'][category][subcat] = final_models_list
        
        # Process category-based products (unlimited) - for non-collections format
        elif 'categories' in scraped_data:
            for category_name, products in scraped_data.get('categories', {}).items():
                if not products:
                    continue
                
                # Organize by subcategories
                subcategories = {}
                for product in products:
                    # Determine subcategory
                    product_text = (product.get('model', '') + ' ' + product.get('description', '')).lower()
                    subcategory = 'general'
                    
                    if any(word in product_text for word in ['chair', 'seating', 'sofa', 'bench', 'stool']):
                        subcategory = 'seating'
                    elif any(word in product_text for word in ['desk', 'table', 'workstation']):
                        subcategory = 'desking'
                    elif any(word in product_text for word in ['cabinet', 'storage', 'shelf', 'drawer']):
                        subcategory = 'storage'
                    elif any(word in product_text for word in ['lamp', 'light']):
                        subcategory = 'lighting'
                    
                    if subcategory not in subcategories:
                        subcategories[subcategory] = []
                    
                    model_entry = {
                        'model': product.get('model', 'Unknown Model'),
                        'price': product.get('price'),
                        'price_range': f"{int(product.get('price', 0))}-{int(product.get('price', 0) * 1.5)}" if product.get('price') else "Contact for price",
                        'features': product.get('features', []),
                        'image_url': product.get('image_url'),
                        'description': product.get('description', ''),
                        'source_url': product.get('source_url', website)
                    }
                    subcategories[subcategory].append(model_entry)
                
                organized_data['categories'][category_name] = subcategories
        
        # Process standalone products (unlimited) - for non-collections format
        if not is_collections_format:
            for product in scraped_data.get('products', []):
                product_text = (product.get('model', '') + ' ' + product.get('description', '')).lower()
                category = 'General'
                subcategory = 'general'
                
                if any(word in product_text for word in ['chair', 'seating', 'sofa']):
                    category = 'Seating'
                    subcategory = 'seating'
                elif any(word in product_text for word in ['desk', 'table']):
                    category = 'Desking'
                    subcategory = 'desking'
                
                if category not in organized_data['categories']:
                    organized_data['categories'][category] = {}
                
                if subcategory not in organized_data['categories'][category]:
                    organized_data['categories'][category][subcategory] = []
                
                model_entry = {
                    'model': product.get('model', 'Unknown Model'),
                    'price': product.get('price'),
                    'price_range': f"{int(product.get('price', 0))}-{int(product.get('price', 0) * 1.5)}" if product.get('price') else "Contact for price",
                    'features': product.get('features', []),
                    'image_url': product.get('image_url'),
                    'description': product.get('description', ''),
                    'source_url': product.get('source_url', website)
                }
                organized_data['categories'][category][subcategory].append(model_entry)
        
        # Save to brands_data folder as separate JSON file
        filepath = save_brand_data_to_file(organized_data, tier, output_dir='brands_data')
        
        # Count total products
        if is_collections_format:
            total_products = organized_data.get('total_products', 0)
            total_collections = organized_data.get('total_collections', 0)
            categories_list = list(organized_data['categories'].keys())
        else:
            total_products = 0
            for category_data in organized_data['categories'].values():
                for subcat_data in category_data.values():
                    total_products += len(subcat_data)
            total_collections = 0
            categories_list = list(organized_data['categories'].keys())
        
        logger.info(f"[Parallel] Brand {brand_name} saved with {total_products} products. File: {filepath}")
        
        # Update brands_dynamic.json
        scraped_at_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        update_brands_dynamic_json(
            brand_name=brand_name,
            website=website,
            country=country,
            tier=tier,
            categories=organized_data.get('categories', {}),
            source='brand_website',
            scraped_at=scraped_at_str
        )
        
        result = {
            'success': True,
            'brand_name': brand_name,
            'message': f'Successfully scraped and saved {brand_name} with {total_products} products. Existing data has been replaced.',
            'products_count': total_products,
            'categories': categories_list,
            'filepath': filepath
        }
        
        if is_collections_format:
            result['collections_count'] = total_collections
            result['message'] = f'Successfully scraped and saved {brand_name} with {total_products} products in {total_collections} collections. Existing data has been replaced.'
        
        return result
        
    except Exception as e:
        logger.exception(f"[Parallel] Error scraping brand {brand_info.get('brand_name', 'Unknown')}: {e}")
        return {
            'success': False,
            'brand_name': brand_info.get('brand_name', 'Unknown'),
            'error': str(e)
        }


@app.route('/api/brands/scrape-multiple', methods=['POST'])
def scrape_multiple_brands():
    """Scrape multiple brands in parallel"""
    try:
        data = request.get_json()
        brands = data.get('brands', [])  # List of brand objects: [{brand_name, website, country, tier}, ...]
        max_workers = data.get('max_workers', 3)  # Number of parallel workers (default: 3)
        
        if not brands or not isinstance(brands, list):
            return jsonify({'error': 'Brands list is required and must be a list'}), 400
        
        if len(brands) == 0:
            return jsonify({'error': 'At least one brand is required'}), 400
        
        # Validate each brand has required fields
        for brand in brands:
            if not brand.get('brand_name') or not brand.get('website'):
                return jsonify({'error': 'Each brand must have brand_name and website'}), 400
        
        logger.info(f"Starting parallel scraping of {len(brands)} brands with {max_workers} workers")
        
        # Use ThreadPoolExecutor for parallel scraping
        from concurrent.futures import ThreadPoolExecutor, as_completed
        import threading
        
        results = []
        completed_count = 0
        total_brands = len(brands)
        
        # Create a lock for thread-safe logging
        log_lock = threading.Lock()
        
        with ThreadPoolExecutor(max_workers=max_workers) as executor:
            # Submit all scraping tasks
            future_to_brand = {
                executor.submit(_scrape_single_brand, brand): brand['brand_name'] 
                for brand in brands
            }
            
            # Collect results as they complete
            for future in as_completed(future_to_brand):
                brand_name = future_to_brand[future]
                try:
                    result = future.result()
                    results.append(result)
                    completed_count += 1
                    
                    with log_lock:
                        status = "✅ SUCCESS" if result.get('success') else "❌ FAILED"
                        logger.info(f"[Parallel] [{completed_count}/{total_brands}] {brand_name}: {status}")
                        if result.get('success'):
                            logger.info(f"[Parallel] {brand_name}: {result.get('products_count', 0)} products scraped")
                        else:
                            logger.error(f"[Parallel] {brand_name}: {result.get('error', 'Unknown error')}")
                            
                except Exception as e:
                    logger.exception(f"[Parallel] Exception for {brand_name}: {e}")
                    results.append({
                        'success': False,
                        'brand_name': brand_name,
                        'error': str(e)
                    })
                    completed_count += 1
        
        # Count successes and failures
        successful = [r for r in results if r.get('success')]
        failed = [r for r in results if not r.get('success')]
        
        total_products = sum(r.get('products_count', 0) for r in successful)
        
        logger.info(f"[Parallel] Completed: {len(successful)} successful, {len(failed)} failed, {total_products} total products")
        
        return jsonify({
            'success': True,
            'message': f'Parallel scraping completed: {len(successful)} successful, {len(failed)} failed',
            'total_brands': total_brands,
            'successful_count': len(successful),
            'failed_count': len(failed),
            'total_products': total_products,
            'results': results
        })
        
    except Exception as e:
        logger.exception('Error in parallel scraping')
        return jsonify({'error': str(e)}), 500


@app.route('/api/brands/scrape-status/<job_id>', methods=['GET'])
def get_scrape_status(job_id):
    """Get current scraping status and events for a job"""
    try:
        with scraping_status_lock:
            status = scraping_status.get(job_id, {
                'status': 'unknown',
                'events': [],
                'progress': 0,
                'message': 'Job not found'
            })
        
        return jsonify({
            'success': True,
            'status': status.get('status', 'unknown'),
            'events': status.get('events', []),
            'progress': status.get('progress', 0),
            'message': status.get('message', ''),
            'timestamp': status.get('timestamp', datetime.now().isoformat())
        })
    except Exception as e:
        logger.exception(f"Error getting scrape status for {job_id}: {e}")
        return jsonify({'error': str(e)}), 500

def update_scrape_status(job_id, status, message=None, progress=None):
    """Update scraping status for real-time preview"""
    try:
        with scraping_status_lock:
            if job_id not in scraping_status:
                scraping_status[job_id] = {
                    'status': 'running',
                    'events': [],
                    'progress': 0,
                    'message': '',
                    'timestamp': datetime.now().isoformat()
                }
            
            scraping_status[job_id]['status'] = status
            scraping_status[job_id]['timestamp'] = datetime.now().isoformat()
            
            if message:
                scraping_status[job_id]['message'] = message
                # Add event (keep last 50 events)
                scraping_status[job_id]['events'].append({
                    'timestamp': datetime.now().isoformat(),
                    'message': message,
                    'status': status
                })
                # Keep only last 50 events
                if len(scraping_status[job_id]['events']) > 50:
                    scraping_status[job_id]['events'] = scraping_status[job_id]['events'][-50:]
            
            if progress is not None:
                scraping_status[job_id]['progress'] = progress
    except Exception as e:
        logger.warning(f"Error updating scrape status: {e}")

def cleanup_scrape_status(job_id):
    """Clean up scraping status after completion (keep for 5 minutes)"""
    def cleanup_after_delay():
        time.sleep(300)  # 5 minutes
        try:
            with scraping_status_lock:
                if job_id in scraping_status:
                    del scraping_status[job_id]
        except:
            pass
    
    Thread(target=cleanup_after_delay, daemon=True).start()

@app.route('/api/brands/download-excel', methods=['GET'])
def download_brand_excel():
    """Download brand data as Excel file"""
    brand = request.args.get('brand', '')
    tier = request.args.get('tier', 'mid_range')
    
    if not brand:
        return jsonify({'error': 'Brand name is required'}), 400
    
    try:
        import pandas as pd
        from io import BytesIO
        import json
        import re
        
        # Normalize tier name
        tier_map = {
            'budgetary': 'budgetary',
            'mid-range': 'mid_range',
            'mid_range': 'mid_range',
            'high-end': 'high_end',
            'high_end': 'high_end'
        }
        tier = tier_map.get(tier.lower(), tier.lower())
        
        # Load brand data
        safe_brand_name = re.sub(r'[^\w\-_]', '', brand.replace(' ', '_'))
        filename = f"{safe_brand_name}_{tier}.json"
        filepath = os.path.join('brands_data', filename)
        
        if not os.path.exists(filepath):
            # Try case-insensitive search
            brands_data_dir = 'brands_data'
            if os.path.exists(brands_data_dir):
                for f in os.listdir(brands_data_dir):
                    if f.lower() == filename.lower():
                        filepath = os.path.join(brands_data_dir, f)
                        break
        
        if not os.path.exists(filepath):
            return jsonify({'error': f'Brand data not found for {brand} ({tier})'}), 404
        
        with open(filepath, 'r', encoding='utf-8') as f:
            brand_data = json.load(f)
        
        # Prepare data for Excel - flatten structure
        rows = []
        
        # Ensure brand_data is a dict
        if not isinstance(brand_data, dict):
            logger.error(f"Brand data is not a dictionary: {type(brand_data)}")
            return jsonify({'error': 'Invalid brand data format'}), 500
        
        logger.info(f"Processing brand data for {brand} ({tier}). Available keys: {list(brand_data.keys())}")
        
        categories_data = brand_data.get('categories', {})
        
        # Handle different data structures
        if not isinstance(categories_data, dict):
            categories_data = {}
        
        # Log what we found
        has_collections = 'collections' in brand_data and isinstance(brand_data.get('collections'), dict) and len(brand_data.get('collections', {})) > 0
        has_category_tree = 'category_tree' in brand_data and isinstance(brand_data.get('category_tree'), dict) and len(brand_data.get('category_tree', {})) > 0
        has_all_products = 'all_products' in brand_data and isinstance(brand_data.get('all_products'), list) and len(brand_data.get('all_products', [])) > 0
        has_categories = isinstance(categories_data, dict) and len(categories_data) > 0
        
        logger.info(f"Data structure check - categories: {has_categories}, collections: {has_collections}, category_tree: {has_category_tree}, all_products: {has_all_products}")
        
        for category_name, category_info in categories_data.items():
            # Handle case where category_info might not be a dict
            if not isinstance(category_info, dict):
                continue
            
            # Check if this category has products directly (Architonic format)
            if 'products' in category_info:
                products = category_info.get('products', [])
                if isinstance(products, list):
                    subcategory_name = category_info.get('subcategory') or 'General'
                    # Clean category name (remove product count if present)
                    clean_category_name = category_name.split('\n')[0].strip() if '\n' in category_name else category_name
                    
                    for product in products:
                        if not isinstance(product, dict):
                            logger.warning(f"Skipping non-dict product in category {category_name}: {type(product)}")
                            continue
                        
                        # Safely extract product data
                        model = ''
                        if 'model' in product and product['model'] is not None:
                            model = str(product['model'])
                        elif 'name' in product and product['name'] is not None:
                            model = str(product['name'])
                        
                        description = ''
                        if 'description' in product and product['description'] is not None:
                            description = str(product['description'])
                        
                        price = product.get('price') if 'price' in product else None
                        price_range = product.get('price_range', 'Contact for price') if 'price_range' in product else 'Contact for price'
                        product_id = str(product.get('product_id', '')) if 'product_id' in product else ''
                        
                        url = ''
                        if 'source_url' in product and product['source_url']:
                            url = str(product['source_url'])
                        elif 'url' in product and product['url']:
                            url = str(product['url'])
                        
                        image_url = str(product.get('image_url', '')) if 'image_url' in product else ''
                        
                        features = ''
                        if 'features' in product:
                            if isinstance(product['features'], list):
                                features = ', '.join(str(f) for f in product['features'])
                            elif product['features'] is not None:
                                features = str(product['features'])
                        
                        rows.append({
                            'Brand': brand_data.get('brand', brand),
                            'Category': clean_category_name,
                            'Sub-Category': subcategory_name if subcategory_name else 'General',
                            'Model': model,
                            'Price': price,
                            'Price Range': price_range,
                            'Description': description,
                            'Product ID': product_id,
                            'URL': url,
                            'Image URL': image_url,
                            'Features': features
                        })
                continue
            
            # Handle nested subcategories structure (legacy format)
            # Skip metadata keys that are not subcategories
            skip_keys = {'url', 'category', 'subcategory', 'product_count', 'products', 'category_tree', 'collections', 'all_products'}
            
            for subcategory_name, models in category_info.items():
                # Skip metadata keys
                if subcategory_name in skip_keys:
                    continue
                
                # Skip non-dict/list values (like strings, numbers, etc.)
                if not isinstance(models, (dict, list)):
                    continue
                
                # If models is a list directly
                if isinstance(models, list):
                    for model in models:
                        if not isinstance(model, dict):
                            if isinstance(model, str):
                                rows.append({
                                    'Brand': brand_data.get('brand', brand),
                                    'Category': category_name,
                                    'Sub-Category': 'General',
                                    'Model': model,
                                    'Price': None,
                                    'Price Range': 'Contact for price',
                                    'Description': '',
                                    'Product ID': '',
                                    'URL': '',
                                    'Image URL': '',
                                    'Features': ''
                                })
                            continue
                        
                        # Safely extract model data
                        model_name = ''
                        if 'model' in model and model['model'] is not None:
                            model_name = str(model['model'])
                        elif 'name' in model and model['name'] is not None:
                            model_name = str(model['name'])
                        
                        description = ''
                        if 'description' in model and model['description'] is not None:
                            description = str(model['description'])
                        
                        price = model.get('price') if 'price' in model else None
                        price_range = model.get('price_range', 'Contact for price') if 'price_range' in model else 'Contact for price'
                        product_id = str(model.get('product_id', '')) if 'product_id' in model else ''
                        
                        url = ''
                        if 'source_url' in model and model['source_url']:
                            url = str(model['source_url'])
                        elif 'url' in model and model['url']:
                            url = str(model['url'])
                        
                        image_url = str(model.get('image_url', '')) if 'image_url' in model else ''
                        
                        features = ''
                        if 'features' in model:
                            if isinstance(model['features'], list):
                                features = ', '.join(str(f) for f in model['features'])
                            elif model['features'] is not None:
                                features = str(model['features'])
                        
                        rows.append({
                            'Brand': brand_data.get('brand', brand),
                            'Category': category_name,
                            'Sub-Category': 'General',
                            'Model': model_name,
                            'Price': price,
                            'Price Range': price_range,
                            'Description': description,
                            'Product ID': product_id,
                            'URL': url,
                            'Image URL': image_url,
                            'Features': features
                        })
                # If models is a dict (nested subcategories)
                elif isinstance(models, dict):
                    if 'products' in models:
                        products = models.get('products', [])
                        if isinstance(products, list):
                            for product in products:
                                if not isinstance(product, dict):
                                    continue
                                
                                # Safely extract product data
                                model = ''
                                if 'model' in product and product['model'] is not None:
                                    model = str(product['model'])
                                elif 'name' in product and product['name'] is not None:
                                    model = str(product['name'])
                                
                                description = ''
                                if 'description' in product and product['description'] is not None:
                                    description = str(product['description'])
                                
                                price = product.get('price') if 'price' in product else None
                                price_range = product.get('price_range', 'Contact for price') if 'price_range' in product else 'Contact for price'
                                product_id = str(product.get('product_id', '')) if 'product_id' in product else ''
                                
                                url = ''
                                if 'source_url' in product and product['source_url']:
                                    url = str(product['source_url'])
                                elif 'url' in product and product['url']:
                                    url = str(product['url'])
                                
                                image_url = str(product.get('image_url', '')) if 'image_url' in product else ''
                                
                                features = ''
                                if 'features' in product:
                                    if isinstance(product['features'], list):
                                        features = ', '.join(str(f) for f in product['features'])
                                    elif product['features'] is not None:
                                        features = str(product['features'])
                                
                    rows.append({
                        'Brand': brand_data.get('brand', brand),
                        'Category': category_name,
                        'Sub-Category': subcategory_name,
                                    'Model': model,
                                    'Price': price,
                                    'Price Range': price_range,
                                    'Description': description,
                                    'Product ID': product_id,
                                    'URL': url,
                                    'Image URL': image_url,
                                    'Features': features
                    })
        
        # If collections format exists, also include those
        if 'collections' in brand_data:
            collections_data = brand_data.get('collections', {})
            if isinstance(collections_data, dict):
                for collection_name, collection_data in collections_data.items():
                    if not isinstance(collection_data, dict):
                        continue
                    
                    clean_collection_name = collection_name.split('\n')[0].strip() if isinstance(collection_name, str) else str(collection_name)
                    products = collection_data.get('products', [])
                    if not isinstance(products, list):
                        continue
                    
                    for product in products:
                        # Handle case where product might not be a dict
                        if not isinstance(product, dict):
                            logger.warning(f"Skipping non-dict product in collection {collection_name}: {type(product)}")
                            continue
                        
                        # Safely extract product data
                        model = ''
                        if 'name' in product and product['name'] is not None:
                            model = str(product['name'])
                        elif 'model' in product and product['model'] is not None:
                            model = str(product['model'])
                        
                        description = ''
                        if 'description' in product and product['description'] is not None:
                            description = str(product['description'])
                        
                        price = product.get('price') if 'price' in product else None
                        price_range = product.get('price_range', 'Contact for price') if 'price_range' in product else 'Contact for price'
                        product_id = str(product.get('product_id', '')) if 'product_id' in product else ''
                        
                        url = ''
                        if 'url' in product and product['url']:
                            url = str(product['url'])
                        elif 'source_url' in product and product['source_url']:
                            url = str(product['source_url'])
                        
                        image_url = str(product.get('image_url', '')) if 'image_url' in product else ''
                        
                        features = ''
                        if 'features' in product:
                            if isinstance(product['features'], list):
                                features = ', '.join(str(f) for f in product['features'])
                            elif product['features'] is not None:
                                features = str(product['features'])
                        
                    rows.append({
                        'Brand': brand_data.get('brand', brand),
                        'Category': clean_collection_name,
                        'Sub-Category': 'general',
                            'Model': model,
                            'Price': price,
                            'Price Range': price_range,
                            'Description': description,
                            'Product ID': product_id,
                            'URL': url,
                            'Image URL': image_url,
                            'Features': features
                        })
        
        # Also check all_products array (check even if rows exist)
        if 'all_products' in brand_data:
            all_products = brand_data.get('all_products', [])
            if isinstance(all_products, list) and len(all_products) > 0:
                logger.info(f"Processing all_products array with {len(all_products)} products")
                for product in all_products:
                    if not isinstance(product, dict):
                        continue
                    
                    # Use direct access with type checking
                    category = str(product.get('category', 'General')) if 'category' in product else 'General'
                    subcategory = str(product.get('subcategory', 'General')) if 'subcategory' in product else 'General'
                    
                    model = ''
                    if 'name' in product:
                        model = str(product['name']) if product['name'] is not None else ''
                    elif 'model' in product:
                        model = str(product['model']) if product['model'] is not None else ''
                    
                    description = ''
                    if 'description' in product:
                        description = str(product['description']) if product['description'] is not None else ''
                    
                    price = product.get('price') if 'price' in product else None
                    price_range = product.get('price_range', 'Contact for price') if 'price_range' in product else 'Contact for price'
                    product_id = str(product.get('product_id', '')) if 'product_id' in product else ''
                    url = product.get('url', product.get('source_url', '')) if 'url' in product or 'source_url' in product else ''
                    image_url = str(product.get('image_url', '')) if 'image_url' in product else ''
                    
                    features = ''
                    if 'features' in product and isinstance(product['features'], list):
                        features = ', '.join(product['features'])
                    elif 'features' in product:
                        features = str(product['features'])
                    
                    rows.append({
                        'Brand': brand_data.get('brand', brand),
                        'Category': category,
                        'Sub-Category': subcategory,
                        'Model': model,
                        'Price': price,
                        'Price Range': price_range,
                        'Description': description,
                        'Product ID': product_id,
                        'URL': url,
                        'Image URL': image_url,
                        'Features': features
                    })
        
        # Also check category_tree format (check even if rows exist, as it might have more products)
        if 'category_tree' in brand_data:
            category_tree = brand_data.get('category_tree', {})
            if isinstance(category_tree, dict) and len(category_tree) > 0:
                logger.info(f"Processing category_tree with {len(category_tree)} categories")
                for category_name, category_info in category_tree.items():
                    if not isinstance(category_info, dict):
                        continue
                    
                    subcategories = category_info.get('subcategories', {})
                    if not isinstance(subcategories, dict):
                        continue
                    
                    for subcategory_name, subcategory_info in subcategories.items():
                        if not isinstance(subcategory_info, dict):
                            continue
                        
                        products = subcategory_info.get('products', [])
                        if not isinstance(products, list):
                            continue
                        
                        for product in products:
                            if not isinstance(product, dict):
                                continue
                            
                            # Use direct access with type checking
                            model = ''
                            if 'name' in product:
                                model = str(product['name']) if product['name'] is not None else ''
                            elif 'model' in product:
                                model = str(product['model']) if product['model'] is not None else ''
                            
                            description = ''
                            if 'description' in product:
                                description = str(product['description']) if product['description'] is not None else ''
                            
                            price = product.get('price') if 'price' in product else None
                            price_range = product.get('price_range', 'Contact for price') if 'price_range' in product else 'Contact for price'
                            product_id = str(product.get('product_id', '')) if 'product_id' in product else ''
                            
                            url = ''
                            if 'url' in product and product['url']:
                                url = str(product['url'])
                            elif 'source_url' in product and product['source_url']:
                                url = str(product['source_url'])
                            
                            image_url = str(product.get('image_url', '')) if 'image_url' in product else ''
                            
                            features = ''
                            if 'features' in product and isinstance(product['features'], list):
                                features = ', '.join(product['features'])
                            elif 'features' in product:
                                features = str(product['features'])
                            
                            rows.append({
                                'Brand': brand_data.get('brand', brand),
                                'Category': str(category_name) if category_name else 'General',
                                'Sub-Category': str(subcategory_name) if subcategory_name else 'General',
                                'Model': model,
                                'Price': price,
                                'Price Range': price_range,
                                'Description': description,
                                'Product ID': product_id,
                                'URL': url,
                                'Image URL': image_url,
                                'Features': features
                    })
        
        if not rows:
            logger.error(f"No products found in brand data for {brand} ({tier})")
            logger.error(f"Brand data keys: {list(brand_data.keys()) if isinstance(brand_data, dict) else 'Not a dict'}")
            logger.error(f"Categories data type: {type(categories_data)}, length: {len(categories_data) if isinstance(categories_data, dict) else 'N/A'}")
            logger.error(f"Has collections: {'collections' in brand_data}, type: {type(brand_data.get('collections')) if 'collections' in brand_data else 'N/A'}")
            logger.error(f"Has category_tree: {'category_tree' in brand_data}, type: {type(brand_data.get('category_tree')) if 'category_tree' in brand_data else 'N/A'}")
            if 'category_tree' in brand_data:
                cat_tree = brand_data.get('category_tree')
                if isinstance(cat_tree, dict):
                    logger.error(f"Category tree has {len(cat_tree)} categories")
                    for cat_name, cat_info in list(cat_tree.items())[:3]:  # Log first 3
                        logger.error(f"  Category: {cat_name}, type: {type(cat_info)}")
                        if isinstance(cat_info, dict) and 'subcategories' in cat_info:
                            subcats = cat_info.get('subcategories', {})
                            logger.error(f"    Has {len(subcats) if isinstance(subcats, dict) else 0} subcategories")
            logger.error(f"Has all_products: {'all_products' in brand_data}, type: {type(brand_data.get('all_products')) if 'all_products' in brand_data else 'N/A'}")
            return jsonify({'error': 'No products found in brand data'}), 404
        
        logger.info(f"Successfully extracted {len(rows)} products for {brand} ({tier})")
        
        # Validate all rows are dicts
        valid_rows = []
        for idx, row in enumerate(rows):
            if not isinstance(row, dict):
                logger.warning(f"Row {idx} is not a dict: {type(row)}, value: {row}")
                continue
            valid_rows.append(row)
        
        if not valid_rows:
            logger.error("No valid rows after validation")
            return jsonify({'error': 'No valid product data found'}), 404
        
        rows = valid_rows
        
        # Create Excel file with formatting using openpyxl directly
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = 'Products'
        
        # Define styles
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        price_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')  # Light yellow
        price_range_fill = PatternFill(start_color='E7F3FF', end_color='E7F3FF', fill_type='solid')  # Light blue
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        wrap_align = Alignment(wrap_text=True, vertical='top')
        
        # Simplified column order - only Product Name, Description, and PRICE
        column_order = ['Product Name', 'Description', 'PRICE']
        
        # Prepare simplified rows with only the required columns
        simplified_rows = []
        for idx, row in enumerate(rows):
            if not isinstance(row, dict):
                logger.warning(f"Row {idx} is not a dict, skipping: {type(row)}")
                continue
            
            try:
                # Get product name (try model, then name) - ensure it's a string
                product_name = ''
                if 'Model' in row:
                    product_name = str(row['Model']) if row['Model'] is not None else ''
                elif 'model' in row:
                    product_name = str(row['model']) if row['model'] is not None else ''
                elif 'name' in row:
                    product_name = str(row['name']) if row['name'] is not None else ''
                
                # Get description - ensure it's a string
                description = ''
                if 'Description' in row:
                    description = str(row['Description']) if row['Description'] is not None else ''
                elif 'description' in row:
                    description = str(row['description']) if row['description'] is not None else ''
                
                # Get price - can be number or string
                price = ''
                if 'Price' in row:
                    price = row['Price'] if row['Price'] is not None else ''
                elif 'price' in row:
                    price = row['price'] if row['price'] is not None else ''
                
                # Store original identifiers for matching on upload
                product_id = ''
                if 'Product ID' in row:
                    product_id = str(row['Product ID']) if row['Product ID'] is not None else ''
                elif 'product_id' in row:
                    product_id = str(row['product_id']) if row['product_id'] is not None else ''
                
                url = ''
                if 'URL' in row:
                    url = str(row['URL']) if row['URL'] is not None else ''
                elif 'url' in row:
                    url = str(row['url']) if row['url'] is not None else ''
                elif 'source_url' in row:
                    url = str(row['source_url']) if row['source_url'] is not None else ''
                
                # Only add if we have at least a product name
                if product_name.strip():
                    simplified_rows.append({
                        'Product Name': product_name.strip(),
                        'Description': description.strip(),
                        'PRICE': price,
                        '_product_id': product_id,
                        '_url': url
                    })
            except Exception as e:
                logger.error(f"Error processing row {idx}: {e}, row type: {type(row)}")
                continue
        
        if not simplified_rows:
            logger.error(f"No valid products found after processing {len(rows)} rows")
            return jsonify({'error': 'No valid products found'}), 404
        
        available_columns = column_order
        
        # Add header row
        ws.append(available_columns)
        
        # Style header row
        for col_idx, col_name in enumerate(available_columns, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border_style
            
            # Set column widths (simplified)
            if col_name == 'Product Name':
                ws.column_dimensions[get_column_letter(col_idx)].width = 40
            elif col_name == 'Description':
                ws.column_dimensions[get_column_letter(col_idx)].width = 60
            elif col_name == 'PRICE':
                ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Add data rows (using simplified_rows)
        rows = simplified_rows  # Use simplified rows for Excel generation
        for row_idx, row_data in enumerate(rows):
            # Ensure row_data is a dict
            if not isinstance(row_data, dict):
                logger.error(f"Row {row_idx} is not a dict: {type(row_data)}, value: {row_data}")
                continue
            
            try:
                row_values = [row_data.get(col, '') for col in available_columns]
                ws.append(row_values)
            except Exception as e:
                logger.error(f"Error processing row {row_idx}: {e}, row_data type: {type(row_data)}")
                continue
            
            # Style data row
            current_row = ws.max_row
            for col_idx, col_name in enumerate(available_columns, 1):
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = border_style
                
                # Highlight PRICE column prominently
                if col_name == 'PRICE':
                    cell.fill = price_fill  # Light yellow background
                    cell.alignment = center_align
                    cell.font = Font(bold=True, size=11)  # Make price bold
                    # Format as currency if it's a number
                    if isinstance(cell.value, (int, float)) and cell.value:
                        cell.number_format = '#,##0.00'
                    elif cell.value == '' or cell.value is None:
                        cell.value = ''  # Empty cells for user to fill
                elif col_name == 'Description':
                    cell.alignment = wrap_align
                else:
                    cell.alignment = Alignment(vertical='top', horizontal='left')
        
        # Freeze header row
        ws.freeze_panes = 'A2'
        
        # Auto-filter
        ws.auto_filter.ref = ws.dimensions
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename
        safe_filename = f"{safe_brand_name}_{tier}.xlsx"
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=safe_filename
        )
        
    except Exception as e:
        logger.exception('Error generating Excel download for brand')
        return jsonify({'error': str(e)}), 500

@app.route('/api/brands/upload-excel', methods=['POST'])
def upload_brand_excel():
    """Upload Excel file and update PRICE field in existing brand JSON"""
    try:
        import pandas as pd
        import json
        import re
        
        if 'file' not in request.files:
            return jsonify({'error': 'No file provided'}), 400
        
        file = request.files['file']
        brand = request.form.get('brand', '')
        tier = request.form.get('tier', 'mid_range')
        
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
        
        if not file.filename.endswith(('.xlsx', '.xls')):
            return jsonify({'error': 'Invalid file format. Please upload an Excel file (.xlsx or .xls)'}), 400
        
        if not brand:
            return jsonify({'error': 'Brand name is required'}), 400
        
        # Normalize tier name
        tier_map = {
            'budgetary': 'budgetary',
            'mid-range': 'mid_range',
            'mid_range': 'mid_range',
            'high-end': 'high_end',
            'high_end': 'high_end'
        }
        tier = tier_map.get(tier.lower(), tier.lower())
        
        # Read Excel file
        try:
            df = pd.read_excel(file)
        except Exception as e:
            return jsonify({'error': f'Error reading Excel file: {str(e)}'}), 400
        
        # Validate required columns
        required_columns = ['Product Name', 'PRICE']
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            return jsonify({'error': f'Missing required columns: {", ".join(missing_columns)}. Excel must have "Product Name" and "PRICE" columns.'}), 400
        
        # Load existing brand data
        safe_brand_name = re.sub(r'[^\w\-_]', '', brand.replace(' ', '_'))
        filename = f"{safe_brand_name}_{tier}.json"
        filepath = os.path.join('brands_data', filename)
        
        if not os.path.exists(filepath):
            # Try case-insensitive search
            brands_data_dir = 'brands_data'
            if os.path.exists(brands_data_dir):
                for f in os.listdir(brands_data_dir):
                    if f.lower() == filename.lower():
                        filepath = os.path.join(brands_data_dir, f)
                        break
        
        if not os.path.exists(filepath):
            return jsonify({'error': f'Brand data not found for {brand} ({tier}). Please download the database first.'}), 404
        
        # Load existing JSON
        with open(filepath, 'r', encoding='utf-8') as f:
            brand_data = json.load(f)
        
        # Create a mapping from Excel: Product Name -> PRICE
        price_updates = {}
        for _, row in df.iterrows():
            product_name = str(row.get('Product Name', '')).strip()
            price = row.get('PRICE', '')
            
            # Handle empty or NaN prices
            if pd.isna(price) or price == '' or price is None:
                continue
            
            # Convert price to number if possible
            try:
                if isinstance(price, str):
                    # Remove currency symbols and commas
                    price_clean = re.sub(r'[^\d.]', '', str(price))
                    if price_clean:
                        price = float(price_clean)
                    else:
                        continue
                elif isinstance(price, (int, float)):
                    price = float(price)
                else:
                    continue
            except (ValueError, TypeError):
                continue
            
            if product_name:
                price_updates[product_name.lower()] = price
        
        # Update prices in brand_data by matching product names
        updated_count = 0
        
        def update_product_price(product, product_name_key):
            """Update price in a product dict if name matches"""
            nonlocal updated_count
            if not isinstance(product, dict):
                return False
            
            # Try to match by product name (model or name field)
            product_name = str(product.get('model', product.get('name', product.get('Model', '')))).strip().lower()
            
            if product_name in price_updates:
                product['price'] = price_updates[product_name]
                updated_count += 1
                return True
            return False
        
        # Update prices in categories structure
        categories_data = brand_data.get('categories', {})
        if isinstance(categories_data, dict):
            for category_name, category_info in categories_data.items():
                if not isinstance(category_info, dict):
                    continue
                
                # Check if category has products directly
                if 'products' in category_info:
                    products = category_info.get('products', [])
                    if isinstance(products, list):
                        for product in products:
                            update_product_price(product, 'model')
                
                # Check nested subcategories
                for key, value in category_info.items():
                    if key in {'url', 'category', 'subcategory', 'product_count', 'products'}:
                        continue
                    
                    if isinstance(value, list):
                        for item in value:
                            if isinstance(item, dict):
                                update_product_price(item, 'model')
                    elif isinstance(value, dict) and 'products' in value:
                        products = value.get('products', [])
                        if isinstance(products, list):
                            for product in products:
                                update_product_price(product, 'model')
        
        # Update prices in collections structure
        if 'collections' in brand_data:
            collections_data = brand_data.get('collections', {})
            if isinstance(collections_data, dict):
                for collection_name, collection_info in collections_data.items():
                    if not isinstance(collection_info, dict):
                        continue
                    
                    products = collection_info.get('products', [])
                    if isinstance(products, list):
                        for product in products:
                            update_product_price(product, 'name')
        
        # Update prices in category_tree structure
        if 'category_tree' in brand_data:
            category_tree = brand_data.get('category_tree', {})
            if isinstance(category_tree, dict):
                for category_name, category_info in category_tree.items():
                    if not isinstance(category_info, dict):
                        continue
                    
                    subcategories = category_info.get('subcategories', {})
                    if isinstance(subcategories, dict):
                        for subcategory_name, subcategory_info in subcategories.items():
                            if not isinstance(subcategory_info, dict):
                                continue
                            
                            products = subcategory_info.get('products', [])
                            if isinstance(products, list):
                                for product in products:
                                    update_product_price(product, 'name')
        
        # Update prices in all_products array
        if 'all_products' in brand_data:
            all_products = brand_data.get('all_products', [])
            if isinstance(all_products, list):
                for product in all_products:
                    update_product_price(product, 'name')
        
        # Save updated JSON
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(brand_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Updated {updated_count} product prices for {brand} ({tier})")
        
        return jsonify({
            'success': True,
            'message': f'Successfully updated {updated_count} product prices for {brand}',
            'updated_count': updated_count,
            'total_in_excel': len(price_updates),
            'filepath': filepath
        })
        
    except Exception as e:
        logger.exception('Error uploading Excel file')
        return jsonify({'error': str(e)}), 500
        organized_data = {
            'brand': brand,
            'website': request.form.get('website', ''),
            'country': request.form.get('country', 'Unknown'),
            'tier': tier,
            'categories': {}
        }
        
        # Group by category and subcategory
        for _, row in df.iterrows():
            category = str(row.get('Category', 'General')).strip()
            subcategory = str(row.get('Sub-Category', 'general')).strip() if pd.notna(row.get('Sub-Category')) else 'general'
            model = str(row.get('Model', '')).strip()
            
            if not category or not model:
                continue
            
            # Initialize category if needed
            if category not in organized_data['categories']:
                organized_data['categories'][category] = {}
            
            # Initialize subcategory if needed
            if subcategory not in organized_data['categories'][category]:
                organized_data['categories'][category][subcategory] = []
            
            # Create model entry
            model_entry = {
                'model': model,
                'price': row.get('Price') if pd.notna(row.get('Price')) else None,
                'price_range': str(row.get('Price Range', 'Contact for price')) if pd.notna(row.get('Price Range')) else 'Contact for price',
                'description': str(row.get('Description', '')) if pd.notna(row.get('Description')) else '',
                'product_id': str(row.get('Product ID', '')) if pd.notna(row.get('Product ID')) else '',
                'source_url': str(row.get('URL', '')) if pd.notna(row.get('URL')) else '',
                'image_url': str(row.get('Image URL', '')) if pd.notna(row.get('Image URL')) else '',
                'features': []
            }
            
            # Parse features if provided
            if 'Features' in df.columns and pd.notna(row.get('Features')):
                features_str = str(row.get('Features', ''))
                if features_str:
                    # Split by comma or semicolon
                    model_entry['features'] = [f.strip() for f in re.split(r'[,;]', features_str) if f.strip()]
            
            organized_data['categories'][category][subcategory].append(model_entry)
        
        # Save to brands_data folder
        filepath = save_brand_data_to_file(organized_data, tier, output_dir='brands_data')
        
        # Count total products
        total_products = 0
        for category_data in organized_data['categories'].values():
            for subcat_data in category_data.values():
                total_products += len(subcat_data)
        
        logger.info(f"Brand {brand} uploaded with {total_products} products from Excel. File: {filepath}")
        
        # Update brands_dynamic.json
        website = organized_data.get('website', '')
        country = organized_data.get('country', 'Unknown')
        update_brands_dynamic_json(
            brand_name=brand,
            website=website,
            country=country,
            tier=tier,
            categories=organized_data.get('categories', {}),
            source='excel_upload',
            scraped_at=datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        )
        
        return jsonify({
            'success': True,
            'message': f'Successfully uploaded and saved {brand} with {total_products} products. Existing data has been replaced.',
            'brand': organized_data,
            'products_count': total_products,
            'categories': list(organized_data['categories'].keys()),
            'filepath': filepath
        })
        
    except Exception as e:
        logger.exception('Error uploading Excel file')
        return jsonify({'error': str(e)}), 500

def save_brand_database(database):
    """Save brand database to file"""
    try:
        import json
        db_file = os.path.join('utils', 'brand_database_custom.json')
        with open(db_file, 'w', encoding='utf-8') as f:
            json.dump(database, f, indent=2, ensure_ascii=False)
        logger.info(f"Brand database saved to {db_file}")
    except Exception as e:
        logger.error(f"Error saving brand database: {e}")

def save_brand_data_to_file(brand_data: dict, tier: str, output_dir: str = 'brands_data') -> str:
    """
    Save brand data to separate JSON file in brands_data folder
    Utility function to replace BrandScraper.save_brand_data()
    
    Args:
        brand_data: Dictionary with brand products data
        tier: Budget tier (budgetary, mid_range, high_end)
        output_dir: Output directory (default: brands_data)
        
    Returns:
        Path to saved file
    """
    try:
        import re
        # Create output directory if it doesn't exist
        os.makedirs(output_dir, exist_ok=True)
        
        # Generate filename from brand name and tier
        brand_name = brand_data.get('brand', 'unknown').replace(' ', '_').replace('/', '_')
        safe_brand_name = re.sub(r'[^\w\-_]', '', brand_name)
        safe_tier = tier.replace('-', '_').lower()
        
        filename = f"{safe_brand_name}_{safe_tier}.json"
        filepath = os.path.join(output_dir, filename)
        
        # Save to JSON file
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(brand_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Brand data saved to {filepath}")
        return filepath
        
    except Exception as e:
        logger.error(f"Error saving brand data: {e}")
        raise

def save_individual_brand_file(brand_name: str, website: str, country: str, tier: str, scraped_data: dict):
    """
    Save brand data to individual brand file (e.g., OTTIMO_budgetary.json)
    
    Args:
        brand_name: Name of the brand
        website: Brand website URL
        country: Country of origin
        tier: Budget tier (budgetary, mid_range, high_end)
        scraped_data: Complete scraped data including categories/category_tree
    """
    try:
        # Create safe filename
        import re
        safe_brand_name = re.sub(r'[^\w\-_]', '', brand_name.replace(' ', '_'))
        filename = f"{safe_brand_name}_{tier}.json"
        filepath = os.path.join('brands_data', filename)
        
        # Build individual brand file structure
        brand_file_data = {
            "brand": brand_name,
            "website": website,
            "country": country,
            "tier": tier,
            "categories": scraped_data.get('collections', {}),
            "category_tree": scraped_data.get('category_tree', {}),
            "source": scraped_data.get('source', 'brand_website'),
            "scraped_at": scraped_data.get('scraped_at', datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            "added_date": datetime.now().isoformat(),
            "updated_date": datetime.now().isoformat()
        }
        
        # Add metadata if available
        if 'total_products' in scraped_data:
            brand_file_data['total_products'] = scraped_data['total_products']
        if 'includes_descriptions' in scraped_data:
            brand_file_data['includes_descriptions'] = scraped_data['includes_descriptions']
        
        # Save to individual file
        os.makedirs('brands_data', exist_ok=True)
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(brand_file_data, f, indent=2, ensure_ascii=False)
        
        logger.info(f"✅ Saved individual brand file: {filename}")
        
    except Exception as e:
        logger.error(f"Error saving individual brand file for {brand_name}: {e}")
        # Don't fail the entire scraping operation if individual file save fails
        logger.exception("Full traceback:")

def update_brands_dynamic_json(brand_name: str, website: str, country: str, tier: str, 
                                categories: dict = None, source: str = 'brand_website', 
                                scraped_at: str = None):
    """
    Update brands_dynamic.json with brand information
    
    Args:
        brand_name: Name of the brand
        website: Brand website URL
        country: Country of origin
        tier: Budget tier (budgetary, mid_range, high_end)
        categories: Dictionary of categories and products (optional)
        source: Source of data (default: 'brand_website')
        scraped_at: Timestamp when scraping was completed (optional)
    """
    try:
        brands_dynamic_path = os.path.join('brands_data', 'brands_dynamic.json')
        
        # Load existing brands_dynamic.json or create new structure
        if os.path.exists(brands_dynamic_path):
            with open(brands_dynamic_path, 'r', encoding='utf-8') as f:
                brands_dynamic = json.load(f)
        else:
            brands_dynamic = {
                "brands": [],
                "last_updated": None,
                "version": "2.0"
            }
        
        # Ensure brands list exists
        if 'brands' not in brands_dynamic:
            brands_dynamic['brands'] = []
        
        # Normalize tier name
        tier_map = {
            'budgetary': 'budgetary',
            'mid-range': 'mid_range',
            'mid_range': 'mid_range',
            'high-end': 'high_end',
            'high_end': 'high_end'
        }
        tier = tier_map.get(tier.lower(), tier.lower())
        
        # Check if brand already exists (by name and tier)
        brand_found = False
        current_time = datetime.now().isoformat()
        
        for i, brand in enumerate(brands_dynamic['brands']):
            if brand.get('name', '').lower() == brand_name.lower() and brand.get('tier') == tier:
                # Update existing brand
                brand['website'] = website
                brand['country'] = country
                brand['tier'] = tier
                brand['updated_date'] = current_time
                brand['source'] = source
                
                if categories is not None:
                    brand['categories'] = categories
                
                if scraped_at:
                    brand['scraped_at'] = scraped_at
                else:
                    brand['scraped_at'] = current_time.split('T')[0] + ' ' + current_time.split('T')[1].split('.')[0]
                
                brands_dynamic['brands'][i] = brand
                brand_found = True
                logger.info(f"Updated existing brand {brand_name} ({tier}) in brands_dynamic.json")
                break
        
        # Add new brand if not found
        if not brand_found:
            new_brand = {
                "name": brand_name,
                "website": website,
                "country": country,
                "tier": tier,
                "categories": categories if categories is not None else {},
                "added_date": current_time,
                "updated_date": current_time,
                "source": source,
                "scraped_at": scraped_at if scraped_at else (current_time.split('T')[0] + ' ' + current_time.split('T')[1].split('.')[0])
            }
            brands_dynamic['brands'].append(new_brand)
            logger.info(f"Added new brand {brand_name} ({tier}) to brands_dynamic.json")
        
        # Update metadata
        brands_dynamic['last_updated'] = current_time
        if 'version' not in brands_dynamic:
            brands_dynamic['version'] = "2.0"
        
        # Save updated file
        os.makedirs('brands_data', exist_ok=True)
        with open(brands_dynamic_path, 'w', encoding='utf-8') as f:
            json.dump(brands_dynamic, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Successfully updated brands_dynamic.json with {brand_name}")
        
    except Exception as e:
        logger.error(f"Error updating brands_dynamic.json: {e}")
        logger.exception("Full traceback:")

@app.route('/download/<file_type>/<file_id>', methods=['GET'])
def download(file_type, file_id):
    """Download generated files"""
    format_type = request.args.get('format', 'pdf')
    
    try:
        from utils.download_manager import DownloadManager
        manager = DownloadManager()
        file_path = manager.prepare_download(file_id, file_type, format_type, session)
        
        return send_file(file_path, as_attachment=True)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/download/extracted/<file_id>', methods=['GET'])
def download_extracted(file_id):
    """Download extracted table data as Excel"""
    format_type = request.args.get('format', 'excel')
    
    try:
        import pandas as pd
        from io import BytesIO
        
        # Get file info
        uploaded_files = session.get('uploaded_files', [])
        file_info = None
        
        for f in uploaded_files:
            if f['id'] == file_id:
                file_info = f
                break
        
        if not file_info or 'extraction_result' not in file_info:
            return jsonify({'error': 'Extraction result not found'}), 404
        
        extraction_result = file_info['extraction_result']
        
        # Create Excel file
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        for page_idx, layout_result in enumerate(extraction_result.get('layoutParsingResults', [])):
            markdown_text = layout_result.get('markdown', {}).get('text', '')
            
            # Try to parse HTML tables from markdown
            if '<table' in markdown_text:
                try:
                    # Use pandas to read HTML tables
                    tables = pd.read_html(markdown_text)
                    for table_idx, df in enumerate(tables):
                        sheet_name = f"Page{page_idx+1}_Table{table_idx+1}"[:31]  # Excel sheet name limit
                        df.to_excel(writer, sheet_name=sheet_name, index=False)
                except Exception as e:
                    logger.error(f"Error parsing HTML table: {e}")
                    # Fallback: create a sheet with raw text
                    df = pd.DataFrame({'Extracted Text': [markdown_text]})
                    sheet_name = f"Page{page_idx+1}"[:31]
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
            else:
                # No HTML tables, save as text
                df = pd.DataFrame({'Extracted Text': [markdown_text]})
                sheet_name = f"Page{page_idx+1}"[:31]
                df.to_excel(writer, sheet_name=sheet_name, index=False)
        
        writer.close()
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"extracted_{file_info['original_name'].rsplit('.', 1)[0]}.xlsx"
        )
    except Exception as e:
        logger.exception('Error generating Excel download')
        return jsonify({'error': str(e)}), 500

@app.route('/download/stitched/<file_id>', methods=['GET'])
def download_stitched(file_id):
    """Download stitched table as Excel"""
    format_type = request.args.get('format', 'excel')
    
    try:
        import pandas as pd
        from io import BytesIO
        
        # Get file info
        uploaded_files = session.get('uploaded_files', [])
        file_info = None
        
        for f in uploaded_files:
            if f['id'] == file_id:
                file_info = f
                break
        
        if not file_info or 'stitched_table' not in file_info:
            return jsonify({'error': 'Stitched table not found. Please stitch tables first.'}), 404
        
        stitched_html = file_info['stitched_table']['html']
        
        # Create Excel file
        output = BytesIO()
        writer = pd.ExcelWriter(output, engine='openpyxl')
        
        try:
            # Use pandas to read the stitched HTML table
            tables = pd.read_html(stitched_html)
            if tables:
                df = tables[0]
                df.to_excel(writer, sheet_name='Stitched_Table', index=False)
            else:
                return jsonify({'error': 'No table found in stitched data'}), 404
        except Exception as e:
            logger.error(f"Error parsing stitched HTML table: {e}")
            return jsonify({'error': f'Error parsing table: {str(e)}'}), 500
        
        writer.close()
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f"stitched_{file_info['original_name'].rsplit('.', 1)[0]}.xlsx"
        )
    except Exception as e:
        logger.exception('Error generating stitched Excel download')
        return jsonify({'error': str(e)}), 500

@app.route('/download/costed/<file_id>', methods=['GET'])
def download_costed_excel(file_id):
    """Download costed table as Excel with proper formatting (uses DownloadManager)"""
    try:
        from utils.download_manager import DownloadManager
        
        # Get file info
        uploaded_files = session.get('uploaded_files', [])
        file_info = None
        
        for f in uploaded_files:
            if f['id'] == file_id:
                file_info = f
                break
        
        if not file_info or 'costed_data' not in file_info:
            return jsonify({'error': 'Costed data not found. Please apply costing first.'}), 404
        
        # Use DownloadManager to create properly formatted Excel (same as manual costing)
        manager = DownloadManager()
        session_id = session.get('session_id', '')
        costed_data = file_info['costed_data']
        
        output_dir = os.path.join('outputs', session_id, 'downloads')
        os.makedirs(output_dir, exist_ok=True)
        
        file_path = manager.create_offer_excel(costed_data, output_dir, file_id)
        
        return send_file(
            file_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=os.path.basename(file_path)
        )
    except Exception as e:
        logger.exception('Error generating costed Excel download')
        return jsonify({'error': str(e)}), 500

@app.route('/admin/cleanup', methods=['POST'])
def admin_cleanup():
    """Manual trigger for cleanup (admin endpoint)"""
    hours = request.json.get('hours', 24) if request.is_json else 24
    try:
        cleaned = cleanup_old_files(hours=hours)
        return jsonify({
            'success': True,
            'message': f'Cleanup completed',
            'cleaned': cleaned
        })
    except Exception as e:
        logger.exception('Error in manual cleanup')
        return jsonify({'error': str(e)}), 500

@app.route('/api/cleanup-session', methods=['POST'])
def cleanup_session_api():
    """API endpoint for cleaning up current session data"""
    try:
        cleanup_session_files()
        return jsonify({
            'success': True,
            'message': 'Session data cleaned'
        })
    except Exception as e:
        logger.exception('Error in session cleanup')
        return jsonify({'error': str(e)}), 500

@app.route('/api/cleanup-all', methods=['POST'])
def cleanup_all_api():
    """API endpoint for cleaning all sessions (triggered on page load)"""
    try:
        session_id = session.get('session_id')
        if session_id:
            cleaned = cleanup_other_sessions(session_id)
            return jsonify({
                'success': True,
                'message': 'Other sessions cleaned',
                'cleaned': cleaned
            })
        return jsonify({'success': False, 'message': 'No active session'}), 400
    except Exception as e:
        logger.exception('Error in cleanup all')
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    # Run cleanup on server start (only files older than 2 hours to avoid deleting active work)
    try:
        logger.info("Running startup cleanup (files older than 2 hours)...")
        cleaned = cleanup_old_files(hours=2)
        logger.info(f"Startup cleanup completed: {cleaned}")
    except Exception as e:
        logger.error(f"Error in startup cleanup: {e}")
    
    # Start background cleanup thread (runs every hour, cleans files older than 24h)
    cleanup_thread = Thread(target=periodic_cleanup, daemon=True)
    cleanup_thread.start()
    logger.info("Started periodic cleanup thread (runs every hour, cleans files older than 24h)")
    
    # Get port from environment variable (Railway provides this) or default to 5000
    port = int(os.environ.get('PORT', 5000))
    # Disable debug mode in production
    debug = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    
    app.run(debug=debug, host='0.0.0.0', port=port)
